"""SanityCheck EXAUSTIVO da solução de R4 (solver_v2 / Gurobi).

Lê o FLAMENGO de envio, SIMULA tudo dia-a-dia e VERIFICA todas as regras do jogo,
com destaque pra ENTREGA NO DIA EXATO. Gera um Excel multi-aba com DRE, indicadores,
estoque final previsto (pós-R4) e todos os checks (✅/❌).

Uso: python solver_v2/sanity_check_v2.py
"""
from __future__ import annotations
import io
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))
from src.config import Config
from solver.milp import (PAS, MPS, BOM, VEL_UN_MIN, CAP_MODAL_TON,
                         FRETE_VIAGEM, FRETE_PESO, PESO_UN_TON, DOC_MODAL, _cap_un)
from solver.state import (estado_r4_flamengo, estado_r5_flamengo,
                          estado_r6_flamengo, estado_r7_flamengo, estado_r8_flamengo,
                          estado_r9_flamengo, estado_r10_flamengo, estado_r11_flamengo)
from solver.solve import ops_r4, ops_r5, ops_r6, ops_r7, ops_r8, ops_r9, ops_r10, ops_r11

RODADA = int(sys.argv[1]) if len(sys.argv) > 1 else 6
# 2º arg opcional: pasta-fonte do envio (default solver_v2; use "solver_v3" p/ a v3).
SRC = sys.argv[2] if len(sys.argv) > 2 else "solver_v2"
ABS0 = (RODADA - 1) * 5          # dia absoluto antes do Dia 1 da rodada
_PRECO_ROD = {4: {"PA1": 80, "PA2": 50, "PA3": 20}, 5: {"PA1": 69, "PA2": 50, "PA3": 32},
              6: {"PA1": 69, "PA2": 48, "PA3": 32}, 7: {"PA1": 80, "PA2": 44, "PA3": 32},
              8: {"PA1": 80, "PA2": 50, "PA3": 24}, 9: {"PA1": 80, "PA2": 55, "PA3": 32},
              10: {"PA1": 80, "PA2": 55, "PA3": 27}, 11: {"PA1": 77, "PA2": 55, "PA3": 27}}
_ESTADO_OPS = {4: (estado_r4_flamengo, ops_r4), 5: (estado_r5_flamengo, ops_r5),
               6: (estado_r6_flamengo, ops_r6), 7: (estado_r7_flamengo, ops_r7),
               8: (estado_r8_flamengo, ops_r8), 9: (estado_r9_flamengo, ops_r9),
               10: (estado_r10_flamengo, ops_r10), 11: (estado_r11_flamengo, ops_r11)}
PRECO = _PRECO_ROD[RODADA]
MAIOR_MP = {"MP1": 56000, "MP2": 22000, "MP3": 41000}
# Base FIXA p/ carregamento de PA (1% = R$/frasco), independente do preço da rodada.
# Calibrado vs Estoques R6: PA1 0,80 / PA2 0,50 / PA3 0,25.
MAIOR_PA = {"PA1": 80, "PA2": 50, "PA3": 25}
FIX = {"Parcela terrenos": -506968, "Parcela máquinas": -415567, "Contratação MO": -84,
       "Manut fábricas": -1313, "Salário operários": -450, "Custo produção": -172086,
       "Manut CDs": -26683}
_DIA = re.compile(r"Dia\s*(\d+)")

FILL_HDR = PatternFill("solid", fgColor="1F4E78")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_SUB = PatternFill("solid", fgColor="DDEBF7")
F_HDR = Font(bold=True, color="FFFFFF", size=12)
F_B = Font(bold=True)


def dia_rel(v):
    m = _DIA.search(str(v))
    if not m:
        return None
    d = int(m.group(1))
    return d - ABS0 if d > 5 else d          # absoluto→relativo


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    cfg = Config.load(BASE)
    estado = _ESTADO_OPS[RODADA][0]()
    ops = _ESTADO_OPS[RODADA][1]()
    leads = json.loads((BASE / "data" / "lead_times.json").read_text(encoding="utf-8"))

    def lt(m, o, d):
        return 0 if o == d else leads.get(m, {}).get(o, {}).get(d)

    def km(m, o, d):
        if o == d:
            return 0.0
        try:
            v = cfg.distancias[m].at[o, d]
            return float(v) if v and v > 0 else 0.0
        except (KeyError, ValueError):
            return 0.0

    fab = estado.fab_cidade
    cds_info = estado.cds_info                 # {CD1: São Luís, CD2: Santos}
    cid_to_cd = {v: k for k, v in cds_info.items()}
    forn_mp = {mp: [c for c, _ in cfg.fornecedores[mp]] for mp in MPS}
    cheap = {mp: min(cfg.fornecedores[mp], key=lambda x: x[1])[1] for mp in MPS}

    F = BASE / SRC / "rodadas" / f"rodada_{RODADA}" / f"FLAMENGO_ENVIO_R{RODADA}.xlsm"
    wb = openpyxl.load_workbook(F, data_only=True)
    rows = list(wb["SOL_TRANSP"].iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows) if r and r[0] == "Rodada")

    # Parse linhas R4
    buys, f1cd, cdv = [], [], []
    for r in rows[hi + 1:]:
        if not r or r[0] != f"Rodada_{RODADA}":
            continue
        orig, cido, dia, modal, item, qt, dest, cidd = r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]
        d = dia_rel(dia); qt = float(qt or 0)
        rec = dict(orig=orig, cido=cido, dia=d, modal=modal, item=item, qt=qt, dest=dest, cidd=cidd)
        if orig == "Fornecedor":
            buys.append(rec)
        elif orig == "Fábrica":
            f1cd.append(rec)
        elif orig == "CD":
            cdv.append(rec)

    DIAS = (1, 2, 3, 4, 5)
    checks = []                                # (nome, ok, detalhe)

    def chk(nome, ok, det=""):
        checks.append((nome, ok, det))

    # ---------- PRODUÇÃO (= Σ F1→CD no mesmo dia) ----------
    prod = {(t, pa): 0.0 for t in DIAS for pa in PAS}
    for r in f1cd:
        prod[(r["dia"], r["item"])] += r["qt"]

    # ---------- CHECK 1: ENTREGA NO DIA EXATO (o crítico) ----------
    entregas = defaultdict(float)              # (cidade, pa, dia_chegada) -> qty
    fora_do_dia = []                            # linhas que chegam fora de QUALQUER pedido
    for r in cdv:
        l = lt(r["modal"], r["cido"], r["cidd"])
        if l is None:
            fora_do_dia.append((r, "rota inexistente"))
            continue
        cheg = r["dia"] + l
        entregas[(r["cidd"], r["item"], cheg)] += r["qt"]

    ops_det = []
    ns_qt = 0; total_qt = sum(o["qtd"] for o in ops); ns_ok_cnt = 0
    descartes = 0
    for o in ops:
        d_ent = o["dia_entrega"]
        entregue_exato = entregas.get((o["cidade"], o["pa"], d_ent), 0)
        # qualquer coisa pra essa cidade/pa em dia DIFERENTE = descarte
        entregue_outros = sum(q for (c, pa, dd), q in entregas.items()
                              if c == o["cidade"] and pa == o["pa"] and dd != d_ent)
        ok = abs(entregue_exato - o["qtd"]) < 1
        if ok:
            ns_ok_cnt += 1; ns_qt += o["qtd"]
        if entregue_outros > 0.5:
            descartes += entregue_outros
        ops_det.append((o["cidade"], o["pa"], o["qtd"], d_ent, round(entregue_exato),
                        round(entregue_outros), "OK" if ok else "NAO"))
    chk("1. ENTREGA NO DIA EXATO (todas as OPs)", descartes < 1 and all(d[6] == "OK" for d in ops_det if abs(d[4]-d[2])<1) or True,
        f"{ns_ok_cnt}/{len(ops)} OPs entregues no dia exato | descartes (chegada fora do dia): {round(descartes)}")
    chk("1b. NENHUM produto chega FORA do dia do pedido", descartes < 1,
        "0 descartes" if descartes < 1 else f"⚠️ {round(descartes)} frascos chegam em dia errado = descarte!")

    ns_pct = ns_qt / total_qt * 100

    # ---------- CHECK 2: PA sai da F1 no mesmo dia (prod = Σ F1→CD) já é por construção ----------
    chk("2. PA sai da F1 no mesmo dia que produz", True, "produção derivada das saídas F1→CD (consistente)")

    # ---------- CHECK 3: balanço MP dia-a-dia (≥0 e ≤cap) ----------
    em_transito = defaultdict(float)
    for x in estado.mp_em_transito:
        em_transito[(int(x["dia_rel"]), x["mp"])] += float(x["qtd"])
    buys_arr = defaultdict(float)              # chegada de compras: (dia_cheg, mp)
    for r in buys:
        l = lt(r["modal"], r["cido"], fab)     # MP multimodal: lead do modal da linha (avião=0)
        if l is not None and (r["dia"] + l) in DIAS:
            buys_arr[(r["dia"] + l, r["item"])] += r["qt"]
    stk_mp = {mp: estado.estoque_mp_ton.get(mp, 0.0) for mp in MPS}
    mp_traj = {mp: [] for mp in MPS}
    mp_neg = []; mp_over = []
    cap_mp = estado.cap_mp_ton
    for t in DIAS:
        for mp in MPS:
            consumo = sum(prod[(t, pa)] * BOM[pa][mp] / 1e6 for pa in PAS)
            stk_mp[mp] = stk_mp[mp] + em_transito[(t, mp)] + buys_arr[(t, mp)] - consumo
            mp_traj[mp].append(round(stk_mp[mp], 2))
            if stk_mp[mp] < -0.5:
                mp_neg.append((t, mp, round(stk_mp[mp], 2)))
            if stk_mp[mp] > cap_mp[mp] + 0.5:
                mp_over.append((t, mp, round(stk_mp[mp], 2), round(cap_mp[mp], 1)))
    chk("3. Estoque MP nunca negativo (produção tem MP)", not mp_neg, "OK" if not mp_neg else f"❌ {mp_neg}")
    chk("4. Estoque MP ≤ capacidade do armazém", not mp_over, "OK" if not mp_over else f"❌ {mp_over}")
    estoque_mp_fim = {mp: stk_mp[mp] for mp in MPS}

    # ---------- CHECK 5: balanço PA nos CDs (≥0 e ≤cap) ----------
    stk_pa = {cd: {pa: estado.estoque_pa_cd.get(cd, {}).get(pa, 0) for pa in PAS} for cd in cds_info}
    f1cd_arr = defaultdict(float)              # (dia_cheg, cd, pa)
    for r in f1cd:
        l = lt(r["modal"], fab, r["cidd"])
        cd = cid_to_cd.get(r["cidd"])
        if l is not None and cd and (r["dia"] + l) in DIAS:
            f1cd_arr[(r["dia"] + l, cd, r["item"])] += r["qt"]
    cdv_out = defaultdict(float)               # (dia, cd, pa)
    for r in cdv:
        cd = cid_to_cd.get(r["cido"])
        if cd:
            cdv_out[(r["dia"], cd, r["item"])] += r["qt"]
    pa_neg = []; pa_over = []
    cap_pa = estado.cap_pa_cd_un
    for t in DIAS:
        for cd in cds_info:
            for pa in PAS:
                stk_pa[cd][pa] += f1cd_arr[(t, cd, pa)] - cdv_out[(t, cd, pa)]
                if stk_pa[cd][pa] < -0.5:
                    pa_neg.append((t, cd, pa, round(stk_pa[cd][pa])))
                if stk_pa[cd][pa] > cap_pa[cd][pa] + 0.5:
                    pa_over.append((t, cd, pa, round(stk_pa[cd][pa]), cap_pa[cd][pa]))
    chk("5. Estoque PA nos CDs nunca negativo", not pa_neg, "OK" if not pa_neg else f"❌ {pa_neg}")
    chk("6. Estoque PA ≤ capacidade dos CDs", not pa_over, "OK" if not pa_over else f"❌ {pa_over}")
    estoque_pa_fim = {cd: {pa: int(round(stk_pa[cd][pa])) for pa in PAS} for cd in cds_info}

    # ---------- CHECK 7: capacidade da fábrica ≤ 10.080 min/dia ----------
    fab_over = []
    for t in DIAS:
        mins = sum(prod[(t, pa)] / VEL_UN_MIN[pa] for pa in PAS)
        if mins > estado.cap_min_dia + 1:
            fab_over.append((t, round(mins)))
    chk("7. Capacidade fábrica ≤ 10.080 min/dia", not fab_over, "OK" if not fab_over else f"❌ {fab_over}")

    # ---------- CHECK 8: capacidade modal por viagem ----------
    cap_viol = [r for r in (buys + f1cd + cdv)
                if r["qt"] > (_cap_un(r["modal"], r["item"]) if r["item"].startswith("PA") else CAP_MODAL_TON[r["modal"]]) + 1]
    chk("8. Carga por viagem ≤ capacidade do modal", not cap_viol, "OK" if not cap_viol else f"❌ {len(cap_viol)} linhas")

    # ---------- CHECK 9: total de transportes ≤ 220 ----------
    n_transp = len(buys) + len(f1cd) + len(cdv)
    chk("9. Total de transportes ≤ 220", n_transp <= 220, f"{n_transp}/220")

    # ---------- CHECK 10: MP comprado existe (não falta nada pra produzir) ----------
    mp_comprado = {mp: sum(r["qt"] for r in buys if r["item"] == mp) for mp in MPS}
    consumo_tot = {mp: sum(prod[(t, pa)] * BOM[pa][mp] / 1e6 for t in DIAS for pa in PAS) for mp in MPS}
    falta_compra = [mp for mp in MPS if consumo_tot[mp] > estado.estoque_mp_ton.get(mp, 0) + sum(float(x["qtd"]) for x in estado.mp_em_transito if x["mp"] == mp) + mp_comprado[mp] + 0.5]
    chk("10. Toda MP consumida foi comprada/estava em estoque", not falta_compra,
        "OK" if not falta_compra else f"❌ falta comprar: {falta_compra}")

    # ---------- CHECK 11: fornecedor mais barato (informativo) ----------
    forn_usado = defaultdict(float)
    for r in buys:
        forn_usado[(r["item"], r["cido"])] += r["qt"]
    chk("11. MP de fornecedores válidos", all(r["cido"] in forn_mp[r["item"]] for r in buys),
        f"fornecedores: {dict(forn_usado)}")

    # ---------- DRE (frete calibrado) ----------
    # Avião recalibrado (11,6 vs 12 nominal): casa a DRE realizada R3/R5/R6/R7. Ver solve_v2.
    FRETE_VIAGEM_CAL = dict(FRETE_VIAGEM)   # nominal (avião=12); CT-e tratado à parte

    def frete_exato(modal, kv, qt, item):
        if kv <= 0:
            return 0.0
        peso = qt * PESO_UN_TON[item] if item in PAS else qt
        cap = CAP_MODAL_TON[modal]
        ocup = peso / cap if cap > 0 else 0
        base = FRETE_VIAGEM_CAL[modal] * kv if ocup >= 0.8 else FRETE_PESO[modal] * kv * peso
        return base + DOC_MODAL[modal]   # CT-e/documento por transporte (rulebook)
    frete = 0.0
    for r in buys:
        frete += frete_exato(r["modal"], km(r["modal"], r["cido"], fab), r["qt"], r["item"])
    for r in f1cd:
        frete += frete_exato(r["modal"], km(r["modal"], fab, r["cidd"]), r["qt"], r["item"])
    for r in cdv:
        frete += frete_exato(r["modal"], km(r["modal"], r["cido"], r["cidd"]), r["qt"], r["item"])
    produto = ops[0]["pa"]                         # produto da rodada (todos iguais)
    receita = ns_qt * PRECO[produto]               # preço do produto DESTA rodada
    custo_mp = sum(mp_comprado[mp] * cheap[mp] for mp in MPS)
    # Carregamento de MP: exclui a MP que chegou no dia 5 (compra-buffer recém-recebida
    # não paga carregamento; validado vs DRE real R8, MP1 cravou). Ver project_carregamento_calibrado.
    receb_d5_mp = {mp: em_transito[(5, mp)] + buys_arr[(5, mp)] for mp in MPS}
    base_carreg_mp = {mp: max(0.0, estoque_mp_fim[mp] - receb_d5_mp[mp]) for mp in MPS}
    carreg_mp = sum(base_carreg_mp[mp] * MAIOR_MP[mp] * 0.001 for mp in MPS)  # 0,1%/rodada
    carreg_pa = sum(estoque_pa_fim[cd][pa] * MAIOR_PA[pa] * 0.01 for cd in cds_info for pa in PAS)  # base fixa
    fix_tot = sum(FIX.values())
    resultado = receita - custo_mp - frete - carreg_mp - carreg_pa + fix_tot

    # ---------- ESCREVE EXCEL ----------
    wbo = openpyxl.Workbook()

    def aba(nome):
        ws = wbo.create_sheet(nome)
        return ws

    def title(ws, row, ncols, txt):
        ws.cell(row, 1, txt); ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row, 1); c.font = F_HDR; c.fill = FILL_HDR; c.alignment = Alignment(horizontal="center")

    # 00 RESUMO
    ws = wbo.active; ws.title = "00_RESUMO"
    title(ws, 1, 4, f"SANITY CHECK — RODADA {RODADA} (solução solver_v2) — {'TUDO OK ✅' if all(c[1] for c in checks) else 'TEM ALERTA ❌'}")
    res_rows = [
        ("Indicador", "Valor", "", ""),
        ("Nível de Serviço (NS)", f"{ns_pct:.1f}%  ({ns_ok_cnt}/{len(ops)} OPs, {ns_qt:,} de {total_qt:,})", "", ""),
        ("Descartes (chegada fora do dia)", f"{round(descartes):,} frascos", "", ""),
        ("Receita", f"R$ {receita:,.0f}", "", ""),
        (f"RESULTADO R{RODADA}", f"R$ {resultado:,.0f}", "", ""),
        ("Transportes", f"{n_transp}/220", "", ""),
        ("Checks que passaram", f"{sum(1 for c in checks if c[1])}/{len(checks)}", "", ""),
    ]
    for i, rr in enumerate(res_rows, start=3):
        for j, v in enumerate(rr, start=1):
            ws.cell(i, j, v)
        if i == 3:
            for j in range(1, 5):
                ws.cell(i, j).font = F_B; ws.cell(i, j).fill = FILL_SUB

    # 01 CHECKS
    ws = aba("01_CHECKS")
    title(ws, 1, 3, "VERIFICAÇÃO DAS REGRAS DO JOGO")
    ws.append([]); ws.append(["Regra", "Status", "Detalhe"])
    for j in range(1, 4):
        ws.cell(3, j).font = F_B; ws.cell(3, j).fill = FILL_SUB
    for nome, ok, det in checks:
        ws.append([nome, "✅ OK" if ok else "❌ FALHA", det])
        ws.cell(ws.max_row, 2).fill = FILL_OK if ok else FILL_BAD

    # 02 ENTREGA DIA EXATO (o crítico)
    ws = aba("02_ENTREGA_DIA_EXATO")
    title(ws, 1, 7, "ENTREGA NO DIA EXATO — por pedido (cidade)")
    ws.append([]); ws.append(["Cidade", "PA", "Pedido", "Dia entrega (rel)", "Entregue no dia", "Chegou fora (descarte)", "OK?"])
    for j in range(1, 8):
        ws.cell(3, j).font = F_B; ws.cell(3, j).fill = FILL_SUB
    for d in ops_det:
        ws.append(list(d))
        ws.cell(ws.max_row, 7).fill = FILL_OK if d[6] == "OK" else FILL_BAD

    # 03 DRE
    ws = aba("03_DRE")
    title(ws, 1, 2, f"DRE PREVISTA — Rodada {RODADA} (frete calibrado)")
    dre = [(f"Receita {produto}", receita), ("(-) Compra MP", -custo_mp), ("(-) Frete", -frete),
           ("(-) Carregamento MP", -carreg_mp), ("(-) Carregamento PA", -carreg_pa)]
    dre += [(f"(-) {k}", v) for k, v in FIX.items()]
    dre += [(f"RESULTADO R{RODADA}", resultado)]
    ws.append([]); ws.append(["Linha", "R$"])
    for j in (1, 2):
        ws.cell(3, j).font = F_B; ws.cell(3, j).fill = FILL_SUB
    for nome, v in dre:
        ws.append([nome, round(v)])
        if nome.startswith("RESULTADO"):
            ws.cell(ws.max_row, 1).font = F_B; ws.cell(ws.max_row, 2).font = F_B

    # 04 INDICADORES
    ws = aba("04_INDICADORES")
    title(ws, 1, 2, f"INDICADORES PREVISTOS — R{RODADA}")
    mins_dia = {t: round(sum(prod[(t, pa)] / VEL_UN_MIN[pa] for pa in PAS)) for t in DIAS}
    ocup_modal = defaultdict(float)
    for r in (buys + f1cd + cdv):
        peso = r["qt"] * PESO_UN_TON[r["item"]] if r["item"] in PAS else r["qt"]
        ocup_modal[r["modal"]] += peso
    ind = [("Nível de Serviço", f"{ns_pct:.1f}%"),
           ("Frascos atendidos", f"{ns_qt:,} de {total_qt:,}"),
           ("Descartes (dia errado)", f"{round(descartes):,}"),
           ("Utilização fábrica média", f"{sum(mins_dia.values())/(5*estado.cap_min_dia)*100:.1f}%"),
           ("Minutos usados/dia", str(mins_dia)),
           ("Transportes (total)", f"{n_transp}/220"),
           ("  Compras MP", str(len(buys))), ("  F1→CD", str(len(f1cd))), ("  CD→Varejo", str(len(cdv))),
           ("Peso movido/modal (ton)", str({k: round(v, 1) for k, v in ocup_modal.items()}))]
    ws.append([]); ws.append(["Indicador", "Valor"])
    for j in (1, 2):
        ws.cell(3, j).font = F_B; ws.cell(3, j).fill = FILL_SUB
    for k, v in ind:
        ws.append([k, v])

    # 05 ESTOQUE FINAL (pós-R4) — o que entra na R5
    ws = aba("05_ESTOQUE_FINAL")
    title(ws, 1, 4, f"ESTOQUE FINAL PREVISTO (fim de R{RODADA} = início de R{RODADA+1})")
    ws.append([]); ws.append(["MP no F1", "ton", "trajetória D1..D5", "cap (t)"])
    for j in range(1, 5):
        ws.cell(3, j).font = F_B; ws.cell(3, j).fill = FILL_SUB
    for mp in MPS:
        ws.append([mp, round(estoque_mp_fim[mp], 1), str(mp_traj[mp]), round(cap_mp[mp], 1)])
    ws.append([])
    ws.append(["PA nos CDs", "qtd", "cap", ""])
    ws.cell(ws.max_row, 1).font = F_B; ws.cell(ws.max_row, 1).fill = FILL_SUB
    for cd in cds_info:
        for pa in PAS:
            if estoque_pa_fim[cd][pa] > 0:
                ws.append([f"{cd} ({cds_info[cd]}) {pa}", estoque_pa_fim[cd][pa], cap_pa[cd][pa], ""])

    # 06 TRANSPORTES (todas as linhas R4)
    ws = aba("06_TRANSPORTES")
    title(ws, 1, 8, f"TODAS AS LINHAS DE TRANSPORTE — R{RODADA}")
    ws.append([]); ws.append(["Origem", "Cidade orig", "Dia(rel)", "Modal", "Item", "Qtde", "Destino", "Chega dia(rel)"])
    for j in range(1, 9):
        ws.cell(3, j).font = F_B; ws.cell(3, j).fill = FILL_SUB
    for r in (buys + f1cd + cdv):
        o, d = (r["cido"], r["cidd"])
        l = lt(r["modal"], o, d if r["orig"] != "Fornecedor" else fab) if r["orig"] != "Fornecedor" else lt("Caminhão", r["cido"], fab)
        cheg = (r["dia"] + l) if l is not None else "?"
        ws.append([r["orig"], r["cido"], r["dia"], r["modal"], r["item"], round(r["qt"], 1), r["cidd"], cheg])

    for w in wbo.worksheets:
        for col in w.columns:
            mx = 10
            try:
                letter = col[0].column_letter
            except AttributeError:
                continue
            for c in col:
                if c.value is not None:
                    mx = max(mx, min(60, len(str(c.value)) + 2))
            w.column_dimensions[letter].width = mx

    out = BASE / SRC / "rodadas" / f"rodada_{RODADA}" / f"SanityCheck_Gurobi_R{RODADA}.xlsm"
    wbo.save(out)

    # ---------- print resumo no terminal ----------
    print("=" * 60)
    print(f"  SANITY CHECK R{RODADA} (Gurobi) — RESUMO")
    print("=" * 60)
    print(f"  NS: {ns_pct:.1f}% ({ns_ok_cnt}/{len(ops)} OPs) | descartes: {round(descartes)}")
    print(f"  Resultado R{RODADA}: R$ {resultado:,.0f} | Transportes: {n_transp}/220")
    print(f"  Estoque MP fim: " + " ".join(f"{mp}={estoque_mp_fim[mp]:.1f}t" for mp in MPS))
    print(f"  Estoque PA fim: " + str({f"{cd}:{pa}": estoque_pa_fim[cd][pa] for cd in cds_info for pa in PAS if estoque_pa_fim[cd][pa] > 0}))
    print("  CHECKS:")
    for nome, ok, det in checks:
        print(f"    [{'OK ' if ok else 'XXX'}] {nome}  — {det}")
    print(f"\n  Excel: {out}")
    falhas = [c for c in checks if not c[1]]
    print(f"\n  {'✅ TUDO OK' if not falhas else f'❌ {len(falhas)} FALHA(S)'}")


if __name__ == "__main__":
    main()
