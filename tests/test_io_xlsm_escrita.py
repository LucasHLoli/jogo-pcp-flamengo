import shutil
from pathlib import Path
import openpyxl
from src.io_xlsm import escrever_plano, ler_sol_transp
from src.domain import PlanoTransporte, PlanoProducao

BASE = Path(__file__).resolve().parents[1]
FLAMENGO_ORIG = BASE / "rodadas" / "FLAMENGO.xlsm"


def test_escrever_sol_transp_e_op_fabricas(tmp_path):
    dst = tmp_path / "FLAMENGO.xlsm"
    shutil.copy(FLAMENGO_ORIG, dst)

    planos_t = [
        PlanoTransporte(rodada=2, origem_tipo="Fornecedor", origem_cidade="Manaus",
                        dia_coleta=1, modal="Caminhão", item="MP1", qtd=24.0,
                        destino_tipo="Fábrica", destino_cidade="Joinville"),
        PlanoTransporte(rodada=2, origem_tipo="Fábrica", origem_cidade="Joinville",
                        dia_coleta=3, modal="Navio", item="PA1", qtd=80000,
                        destino_tipo="CD", destino_cidade="Santos"),
    ]
    planos_p = [
        PlanoProducao(rodada=2, fabrica="F1", dia=1, pa="PA1", qtd=10000),
        PlanoProducao(rodada=2, fabrica="F1", dia=2, pa="PA2", qtd=20000),
    ]
    escrever_plano(dst, planos_t, planos_p, rodada_n=2)

    items = ler_sol_transp(dst, rodada=2)
    assert len(items) == 2
    assert items[0].item == "MP1"
    assert items[1].destino_cidade == "Santos"

    wb = openpyxl.load_workbook(dst, keep_vba=True, data_only=False)
    ws = wb["OP_FABRICAS"]
    assert ws.cell(7, 2).value == 10000
    assert ws.cell(8, 3).value == 20000

    # VBA preservado
    assert wb.vba_archive is not None


def test_escrita_preserva_formulas_calculadas(tmp_path):
    dst = tmp_path / "FLAMENGO.xlsm"
    shutil.copy(FLAMENGO_ORIG, dst)
    escrever_plano(dst, [], [], rodada_n=2)

    wb = openpyxl.load_workbook(dst, keep_vba=True, data_only=False)
    ws = wb["SOL_TRANSP"]
    cell = ws.cell(5, 10)
    # célula J5 é fórmula de "Dias úteis viagem"
    assert cell.value is not None and str(cell.value).startswith("=")


def test_escrita_rodada2_preserva_rodada1(tmp_path):
    """Cenário real: SOL_TRANSP já tem 13 linhas de Rodada_1; escrevemos Rodada_2 sem
    sobrescrever as anteriores."""
    src = BASE / "rodadas" / "Rodada 1.xlsm"  # tem 13 linhas pré-preenchidas
    dst = tmp_path / "FLAMENGO.xlsm"
    shutil.copy(src, dst)

    planos_t = [
        PlanoTransporte(rodada=2, origem_tipo="Fábrica", origem_cidade="Joinville",
                        dia_coleta=2, modal="Caminhão", item="PA1", qtd=80000,
                        destino_tipo="CD", destino_cidade="São Luís"),
    ]
    escrever_plano(dst, planos_t, [], rodada_n=2)

    # Rodada_1 preservada
    items_r1 = ler_sol_transp(dst, rodada=1)
    assert len(items_r1) >= 10
    # Rodada_2 escrita
    items_r2 = ler_sol_transp(dst, rodada=2)
    assert len(items_r2) == 1
    assert items_r2[0].modal == "Caminhão"
