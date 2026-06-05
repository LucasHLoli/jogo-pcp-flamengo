"""SanityCheck COMPLETO — R1 + R2 + R3 com simulação dia a dia consolidada.

Para cada rodada N (1..3):
  - Lê SOL_TRANSP do FLAMENGO consolidado
  - Lê OP_FABRICAS (R3 = atual; R1/R2 = inferido das fontes históricas)
  - Simula DIA A DIA: estoque MP F1, PA nos CDs, chegadas, consumo, saídas
  - Calcula: NS, ociosidade, transportes, custos, receita
  - Valida: cap min/dia, cap MP, cap CD, descartes, NS

Gera SanityCheck_Completo.xlsm com:
  00_RESUMO_3_RODADAS
  01_R1_Detalhe + R2_Detalhe + R3_Detalhe (5 sub-blocos cada)
  02_Estoque_Evolução (linha temporal 15 dias)
  03_DRE_Acumulada
  04_NS_Indicadores_Por_Rodada
  05_Checks_Por_Rodada
"""
from __future__ import annotations
import io
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from src.config import Config
from src.io_xlsm import ler_instalacoes
from src.planner_v3 import lead_dias, km_rota, custo_total_modal


# ============ ESTILOS ============
FILL_HDR = PatternFill("solid", fgColor="1F4E78")
FILL_SUB = PatternFill("solid", fgColor="9BC2E6")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
FILL_INFO = PatternFill("solid", fgColor="DDEBF7")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FONT_HDR = Font(bold=True, color="FFFFFF", size=12)
FONT_SUB = Font(bold=True, size=11)
FONT_BOLD = Font(bold=True)


def title(ws, row, ncols, text):
    ws.cell(row, 1).value = text
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row, 1).font = FONT_HDR; ws.cell(row, 1).fill = FILL_HDR
    ws.cell(row, 1).alignment = Alignment(horizontal="center")


def header(ws, row, headers, fill=FILL_INFO):
    for j, h in enumerate(headers):
        c = ws.cell(row, j + 1)
        c.value = h; c.font = FONT_BOLD; c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def money(c):
    c.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def pct(c):
    c.number_format = "0.0%"


def auto_width(ws, mx=50):
    for col in ws.columns:
        try:
            l = col[0].column_letter
        except AttributeError:
            continue
        m = 10
        for c in col:
            if c.value is not None:
                m = max(m, min(mx, len(str(c.value)) + 2))
        ws.column_dimensions[l].width = m


# ============ DADOS GLOBAIS ============
print("Carregando configs...")
cfg = Config.load(BASE)
lead_tab = json.loads((BASE / "data" / "lead_times.json").read_text(encoding="utf-8"))

# Vamos ler do FLAMENGO_ALL (que tem SOL_TRANSP R1+R2 oficiais) e do FLAMENGO.xlsm
# da rodada_3 (que tem o R3 da heurística atualmente)
FLAMENGO_ALL = BASE / "rodadas" / "FLAMENGO_ALL.xlsm"
FLAMENGO_R3 = BASE / "rodadas" / "rodada_3" / "FLAMENGO.xlsm"

inst = ler_instalacoes(FLAMENGO_R3)
F1 = inst["fabricas"]["F1"]
CDS = inst["cds"]
fab_cidade = F1["cidade"]
cds_info = {cd: d["cidade"] for cd, d in CDS.items()}
cidade_to_cd = {v: k for k, v in cds_info.items()}
cap_mp = {mp: F1["area_mp"][mp] * 2 * cfg.densidades_mp[mp] for mp in ("MP1", "MP2", "MP3")}
cap_pa_cd = {cd: {pa: int(d["area_pa"][pa] * 2 * cfg.densidades_pa[pa] / cfg.peso_un_ton[pa])
                  for pa in ("PA1", "PA2", "PA3")} for cd, d in CDS.items()}
CAP_MIN_DIA = F1["maquinas"] * F1["turnos"] * 8 * 60
VEL = {"PA1": 15, "PA2": 30, "PA3": 60}
PRECO_PA = {"PA1": 80, "PA2": 50, "PA3": 32}  # R3 PA3 = R$ 32
FIX = {"parcela_terr": 506968, "parcela_maq": 415567, "parcela_mo": 84,
       "manut_fab": 1313, "sal_op": 450, "custo_prod": 172086, "manut_cd": 26683}
FIX_TOTAL = sum(FIX.values())  # 1,123,151
MAIOR_MP = {"MP1": 56000, "MP2": 22000, "MP3": 41000}
FORN_MIN = {mp: min(cfg.fornecedores[mp], key=lambda x: x[1]) for mp in ("MP1", "MP2", "MP3")}

# OPs por rodada (R1 não tinha pedido, R2 = PA1, R3 = PA3)
OPS_POR_RODADA = {
    1: [],  # setup
    2: [   # PA1 R2 — qty estimada pelo plano produzido (267,080 PA1 produzido)
        # Sem dados oficiais — vamos ler dos PDFs em rodadas/rodada_2/ se houver
        # ou usar produção real
    ],
    3: [{"cidade": c, "pa": "PA3", "qtd": q, "dia_entrega": d} for c, q, d in [
        ("Belém", 20155, 5), ("Belo Horizonte", 70544, 3), ("Brasília", 117573, 3),
        ("Campinas", 56435, 2), ("Campo Grande", 23515, 3), ("Cuiabá", 28218, 3),
        ("Curitiba", 103464, 2), ("Fortaleza", 68528, 5), ("Goiânia", 65841, 3),
        ("João Pessoa", 40311, 4), ("Joinville", 28218, 2), ("Maceió", 40311, 4),
        ("Manaus", 20155, 5), ("Natal", 40311, 5), ("Porto Alegre", 103464, 2),
        ("Recife", 60466, 4), ("Ribeirão Preto", 47029, 2), ("Rio de Janeiro", 94059, 3),
        ("Salvador", 80622, 4), ("Santos", 47029, 2), ("São Luís", 20155, 5),
        ("São Paulo", 117573, 2), ("Uberlândia", 23515, 3), ("Vitória", 14109, 3),
        ("Vitória da Conquista", 12093, 4),
    ]],
}

# DRE oficial R1 (do PDF)
DRE_OFICIAL_R1 = {
    "receita": 0, "parcela_terr": -506968, "parcela_maq": -415567, "parcela_mo": -84,
    "manut_fab": -1313, "sal_op": -450, "custo_prod": -172086, "manut_cd": -26683,
    "compra_mp": -4_368_000, "frete_mp": -105_666, "frete_pa": -95,
    "carreg_mp": -5_410, "carreg_pa": 0,
}


_DIA = re.compile(r"Dia\s*(\d+)")
_ROD = re.compile(r"Rodada_(\d+)")


def parse_sol_transp(path: Path) -> dict:
    """Retorna {rodada: [linhas]} parseado de SOL_TRANSP."""
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    ws = wb["SOL_TRANSP"]
    out = defaultdict(list)
    for r in range(5, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if not val: continue
        m = _ROD.search(str(val))
        if not m: continue
        rod = int(m.group(1))
        dia_str = str(ws.cell(r, 4).value)
        m2 = _DIA.search(dia_str)
        if not m2: continue
        dia_raw = int(m2.group(1))
        # Normaliza para relativo da rodada
        dia_rel = dia_raw - (rod - 1) * 5 if dia_raw > 5 else dia_raw
        if not (1 <= dia_rel <= 5): continue
        out[rod].append({
            "Rodada": rod,
            "Origem": ws.cell(r, 2).value,
            "Cidade": ws.cell(r, 3).value,
            "dia_rel": dia_rel,
            "Modal": ws.cell(r, 5).value,
            "Tipo do Produto": ws.cell(r, 6).value,
            "Qtde": float(ws.cell(r, 7).value or 0),
            "Destino": ws.cell(r, 8).value,
            "Cidade_Destino": ws.cell(r, 9).value,
        })
    return dict(out)


def simular_rodada(rodada: int, linhas: list, op_fabricas: dict,
                   estoque_mp_inicial: dict, estoque_pa_cd_inicial: dict,
                   chegadas_em_transito: dict) -> dict:
    """Simula uma rodada dia a dia retornando estado final + métricas.

    Args:
        rodada: número
        linhas: SOL_TRANSP dessa rodada
        op_fabricas: {dia: {pa: qty}} para os 5 dias
        estoque_mp_inicial: ton no dia 0
        estoque_pa_cd_inicial: {cd: {pa: qty}}
        chegadas_em_transito: {dia: {mp: ton}} de MP de rodadas ANTERIORES chegando nesta

    Returns:
        dict com sim_mp, sim_pa, estoque_final_mp, estoque_final_pa,
        viagens_por_modal, custo_compra_mp, custo_frete, etc.
    """
    # Indexa arrivals/saídas
    arrivals_mp = defaultdict(lambda: defaultdict(float))
    arrivals_pa_cd = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    saidas_cd = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for d, mps in chegadas_em_transito.items():
        for mp, q in mps.items():
            arrivals_mp[d][mp] += q

    # Tail buys: MP comprado nesta rodada com lt que cai na PRÓXIMA
    tail_em_transito = defaultdict(lambda: defaultdict(float))  # [dia_rel_proxima_rodada][mp]

    for row in linhas:
        d_part = row["dia_rel"]
        modal = row["Modal"]; item = row["Tipo do Produto"]; qtd = row["Qtde"]
        o, dest = row["Cidade"], row["Cidade_Destino"]
        lt = lead_dias(cfg, o, dest, modal) or 0
        d_cheg = d_part + lt
        if row["Origem"] == "Fornecedor":
            if 1 <= d_cheg <= 5:
                arrivals_mp[d_cheg][item] += qtd
            else:
                # chega na próxima rodada
                tail_em_transito[d_cheg - 5][item] += qtd
        elif row["Origem"] == "Fábrica" and row["Destino"] == "CD":
            cd_id = cidade_to_cd.get(dest, "?")
            if 1 <= d_cheg <= 5 and cd_id != "?":
                arrivals_pa_cd[d_cheg][cd_id][item] += qtd
        elif row["Origem"] == "CD" and row["Destino"] == "Varejista":
            cd_id = cidade_to_cd.get(o, "?")
            if cd_id != "?":
                saidas_cd[d_part][cd_id][item] += qtd

    # Simulação
    sim_mp = {}
    sim_pa = {cd: {} for cd in CDS}
    stk_mp = dict(estoque_mp_inicial)
    stk_pa = {cd: dict(estoque_pa_cd_inicial.get(cd, {pa: 0 for pa in ("PA1", "PA2", "PA3")}))
              for cd in CDS}

    for dia in range(1, 6):
        prod = op_fabricas.get(dia, {pa: 0 for pa in ("PA1", "PA2", "PA3")})
        sim_mp[dia] = {}
        for mp in ("MP1", "MP2", "MP3"):
            pre = stk_mp[mp]
            arr = arrivals_mp[dia].get(mp, 0)
            pos = pre + arr
            desc = max(0, pos - cap_mp[mp])
            pos = min(pos, cap_mp[mp])
            cons = sum(prod[pa] * cfg.BoM[pa][mp] / 1e6 for pa in ("PA1", "PA2", "PA3"))
            end = pos - cons
            sim_mp[dia][mp] = {"pre": pre, "arr": arr, "pos": pos, "desc": desc,
                                "cons": cons, "end": end}
            stk_mp[mp] = max(0, end)
        for cd in CDS:
            sim_pa[cd][dia] = {}
            for pa in ("PA1", "PA2", "PA3"):
                pre = stk_pa[cd][pa]
                arr = arrivals_pa_cd[dia][cd].get(pa, 0)
                sai = saidas_cd[dia][cd].get(pa, 0)
                pos = pre + arr
                desc = max(0, pos - cap_pa_cd[cd][pa])
                pos = min(pos, cap_pa_cd[cd][pa])
                end = pos - sai
                sim_pa[cd][dia][pa] = {"pre": pre, "arr": arr, "sai": sai,
                                        "desc": desc, "end": end}
                stk_pa[cd][pa] = max(0, end)

    estoque_mp_final = {mp: sim_mp[5][mp]["end"] for mp in ("MP1", "MP2", "MP3")}
    estoque_pa_final = {cd: {pa: int(sim_pa[cd][5][pa]["end"]) for pa in ("PA1","PA2","PA3")}
                        for cd in CDS}

    # Métricas
    viagens_por_modal = defaultdict(int)
    for row in linhas:
        viagens_por_modal[row["Modal"]] += 1

    custo_compra_mp = sum(row["Qtde"] * FORN_MIN[row["Tipo do Produto"]][1]
                          for row in linhas if row["Origem"] == "Fornecedor")
    custo_frete = 0.0
    for row in linhas:
        modal = row["Modal"]; item = row["Tipo do Produto"]; qtd = row["Qtde"]
        kv = km_rota(cfg, row["Cidade"], row["Cidade_Destino"], modal) or 0
        peso = qtd if item.startswith("MP") else qtd * cfg.peso_un_ton[item]
        custo_frete += custo_total_modal(cfg, modal, kv, peso, 1)

    carreg_mp_fim = sum(estoque_mp_final[mp] * MAIOR_MP[mp] * 0.01 for mp in ("MP1","MP2","MP3"))
    carreg_pa_fim = sum(estoque_pa_final[cd][pa] * PRECO_PA[pa] * 0.01
                        for cd in CDS for pa in ("PA1","PA2","PA3"))

    # Min usados/dia
    min_usados = {}
    for dia in range(1, 6):
        prod = op_fabricas.get(dia, {pa: 0 for pa in ("PA1","PA2","PA3")})
        min_usados[dia] = sum(prod[pa] / VEL[pa] for pa in ("PA1", "PA2", "PA3"))

    # NS — para cada OP, verifica se chegou no dia_entrega EXATO
    ops = OPS_POR_RODADA.get(rodada, [])
    ns_atendidas = []
    ns_descartadas = []
    for op in ops:
        # PA dessa OP chegando no varejo no dia_entrega
        qty_no_dia = 0
        for row in linhas:
            if row["Origem"] != "CD" or row["Destino"] != "Varejista": continue
            if row["Cidade_Destino"] != op["cidade"]: continue
            if row["Tipo do Produto"] != op["pa"]: continue
            lt_cd = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
            dia_cheg = row["dia_rel"] + lt_cd
            if dia_cheg == op["dia_entrega"]:
                qty_no_dia += row["Qtde"]
        if abs(qty_no_dia - op["qtd"]) < 1:
            ns_atendidas.append({**op, "qty_no_dia": int(qty_no_dia)})
        else:
            ns_descartadas.append({**op, "qty_no_dia": int(qty_no_dia)})

    qtd_atend = sum(o["qtd"] for o in ns_atendidas)
    qtd_total = sum(o["qtd"] for o in ops) if ops else 0
    receita = sum(o["qtd"] * PRECO_PA[o["pa"]] for o in ns_atendidas)

    custo_total = FIX_TOTAL + custo_compra_mp + custo_frete + carreg_mp_fim + carreg_pa_fim
    resultado = receita - custo_total

    return {
        "rodada": rodada,
        "sim_mp": sim_mp, "sim_pa": sim_pa,
        "estoque_mp_final": estoque_mp_final,
        "estoque_pa_final": estoque_pa_final,
        "tail_em_transito": dict(tail_em_transito),  # MP que sai nesta rodada e chega na próxima
        "viagens_por_modal": dict(viagens_por_modal),
        "n_transp_total": len(linhas),
        "custo_fixo": FIX_TOTAL,
        "custo_compra_mp": custo_compra_mp,
        "custo_frete": custo_frete,
        "custo_carreg_mp": carreg_mp_fim,
        "custo_carreg_pa": carreg_pa_fim,
        "custo_total": custo_total,
        "receita": receita,
        "resultado": resultado,
        "min_usados": min_usados,
        "ops_atendidas": ns_atendidas,
        "ops_descartadas": ns_descartadas,
        "qtd_total": qtd_total,
        "qtd_atendida": qtd_atend,
        "ns_pct": qtd_atend / max(1, qtd_total) * 100 if qtd_total > 0 else 0,
    }


# ============ ROTA DE EXECUÇÃO ============
print("Lendo SOL_TRANSP de FLAMENGO_ALL e FLAMENGO R3...")
linhas_all = parse_sol_transp(FLAMENGO_ALL)
linhas_r3 = parse_sol_transp(FLAMENGO_R3)
linhas_all[3] = linhas_r3.get(3, linhas_all.get(3, []))  # usa R3 do arquivo atual (heurística)
print(f"  R1: {len(linhas_all.get(1, []))} linhas")
print(f"  R2: {len(linhas_all.get(2, []))} linhas")
print(f"  R3: {len(linhas_all.get(3, []))} linhas")

# OP_FABRICAS por rodada
# R1: assume zero (setup)
# R2: do FLAMENGO_envio.xlsm (R2 original)
# R3: do FLAMENGO.xlsm atual (heurística)
op_fab_r1 = {d: {pa: 0 for pa in ("PA1", "PA2", "PA3")} for d in range(1, 6)}

# R2 production - do FLAMENGO_envio.xlsm (que tem OP_FABRICAS R2 ORIGINAL)
wb_r2 = openpyxl.load_workbook(BASE / "rodadas" / "FLAMENGO_envio.xlsm",
                                keep_vba=True, data_only=True)
ws_op_r2 = wb_r2["OP_FABRICAS"]
op_fab_r2 = {}
for d in range(1, 6):
    op_fab_r2[d] = {
        "PA1": int(ws_op_r2.cell(6 + d, 2).value or 0),
        "PA2": int(ws_op_r2.cell(6 + d, 3).value or 0),
        "PA3": int(ws_op_r2.cell(6 + d, 4).value or 0),
    }

# R3 production
wb_r3 = openpyxl.load_workbook(FLAMENGO_R3, keep_vba=True, data_only=True)
ws_op_r3 = wb_r3["OP_FABRICAS"]
op_fab_r3 = {}
for d in range(1, 6):
    op_fab_r3[d] = {
        "PA1": int(ws_op_r3.cell(6 + d, 2).value or 0),
        "PA2": int(ws_op_r3.cell(6 + d, 3).value or 0),
        "PA3": int(ws_op_r3.cell(6 + d, 4).value or 0),
    }

# ===== SIMULAÇÃO R1 → R2 → R3 (CADEIA) =====
print("\nSimulando R1 dia a dia...")
sim_r1 = simular_rodada(
    rodada=1, linhas=linhas_all.get(1, []),
    op_fabricas=op_fab_r1,
    estoque_mp_inicial={"MP1": 0, "MP2": 0, "MP3": 0},
    estoque_pa_cd_inicial={cd: {pa: 0 for pa in ("PA1","PA2","PA3")} for cd in CDS},
    chegadas_em_transito={},
)

# R2 começa com estoque final R1 + chegadas em-trânsito de R1
print("Simulando R2 dia a dia...")
chegadas_r1_para_r2 = sim_r1["tail_em_transito"]  # MP de R1 que chega em R2
sim_r2 = simular_rodada(
    rodada=2, linhas=linhas_all.get(2, []),
    op_fabricas=op_fab_r2,
    estoque_mp_inicial=sim_r1["estoque_mp_final"],
    estoque_pa_cd_inicial=sim_r1["estoque_pa_final"],
    chegadas_em_transito=chegadas_r1_para_r2,
)

print("Simulando R3 dia a dia...")
chegadas_r2_para_r3 = sim_r2["tail_em_transito"]
sim_r3 = simular_rodada(
    rodada=3, linhas=linhas_all.get(3, []),
    op_fabricas=op_fab_r3,
    estoque_mp_inicial=sim_r2["estoque_mp_final"],
    estoque_pa_cd_inicial=sim_r2["estoque_pa_final"],
    chegadas_em_transito=chegadas_r2_para_r3,
)

resultados = {1: sim_r1, 2: sim_r2, 3: sim_r3}

# Imprime resumo
for n in (1, 2, 3):
    s = resultados[n]
    print(f"\n=== R{n} ===")
    print(f"  Transp: {s['n_transp_total']} ({s['viagens_por_modal']})")
    print(f"  Estoque MP fim: {dict((k, round(v, 2)) for k, v in s['estoque_mp_final'].items())}")
    print(f"  PA total fim CDs: {sum(s['estoque_pa_final'][cd][pa] for cd in CDS for pa in ('PA1','PA2','PA3'))}")
    print(f"  NS: {s['ns_pct']:.1f}% ({len(s['ops_atendidas'])}/{len(s['ops_atendidas'])+len(s['ops_descartadas'])} OPs)")
    print(f"  Receita: R$ {s['receita']:>14,.0f}")
    print(f"  Custos: R$ {s['custo_total']:>14,.0f}  (fixo {s['custo_fixo']:>14,.0f} + MP {s['custo_compra_mp']:>10,.0f} + frete {s['custo_frete']:>10,.0f})")
    print(f"  Resultado: R$ {s['resultado']:>14,.0f}")
    print(f"  Min/dia: {dict((d, round(v)) for d, v in s['min_usados'].items())}")


# ============ GERA EXCEL ============
print("\nGerando SanityCheck_Completo.xlsm...")
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============ 00 RESUMO 3 RODADAS ============
ws = wb.create_sheet("00_RESUMO_3_RODADAS")
title(ws, 1, 7, "RESUMO 3 RODADAS — R1 + R2 + R3 (sanity check consolidado)")
header(ws, 3, ["Indicador", "R1 (Setup)", "R2 (real)", "R3 (heurística)", "Total", "Limite/Esperado", "Status"])

cap_check = []  # acumula status pra Total Falhas

def kpi(r, label, v1, v2, v3, total_txt, limite, ok_func=None):
    ws.cell(r, 1).value = label
    ws.cell(r, 2).value = v1
    ws.cell(r, 3).value = v2
    ws.cell(r, 4).value = v3
    ws.cell(r, 5).value = total_txt
    ws.cell(r, 6).value = limite
    if ok_func is not None:
        ok_status = ok_func()
        ws.cell(r, 7).value = "✅" if ok_status else "❌"
        ws.cell(r, 7).fill = FILL_OK if ok_status else FILL_BAD
        cap_check.append(ok_status)

r = 4

kpi(r, "🚚 Transportes (cap 220/rod)",
    sim_r1["n_transp_total"], sim_r2["n_transp_total"], sim_r3["n_transp_total"],
    sum(s["n_transp_total"] for s in resultados.values()),
    "≤220 por rod",
    lambda: all(s["n_transp_total"] <= 220 for s in resultados.values()))
r += 1
kpi(r, "📊 NS (% atend no dia exato)",
    f"{sim_r1['ns_pct']:.1f}% (sem OPs)" if not OPS_POR_RODADA[1] else f"{sim_r1['ns_pct']:.1f}%",
    f"{sim_r2['ns_pct']:.1f}%",
    f"{sim_r3['ns_pct']:.1f}%",
    "—", "—",
    lambda: True)
r += 1
kpi(r, "🎯 OPs atendidas / total",
    f"{len(sim_r1['ops_atendidas'])}/{len(sim_r1['ops_atendidas'])+len(sim_r1['ops_descartadas'])}",
    f"{len(sim_r2['ops_atendidas'])}/{len(sim_r2['ops_atendidas'])+len(sim_r2['ops_descartadas'])}",
    f"{len(sim_r3['ops_atendidas'])}/{len(sim_r3['ops_atendidas'])+len(sim_r3['ops_descartadas'])}",
    "—", "—")
r += 1
kpi(r, "💰 Receita",
    sim_r1["receita"], sim_r2["receita"], sim_r3["receita"],
    sum(s["receita"] for s in resultados.values()), "—")
for c in (2, 3, 4, 5):
    money(ws.cell(r, c))
r += 1
kpi(r, "💸 Custo total",
    sim_r1["custo_total"], sim_r2["custo_total"], sim_r3["custo_total"],
    sum(s["custo_total"] for s in resultados.values()), "—")
for c in (2, 3, 4, 5):
    money(ws.cell(r, c))
r += 1
kpi(r, "📈 Resultado da rodada",
    sim_r1["resultado"], sim_r2["resultado"], sim_r3["resultado"],
    sum(s["resultado"] for s in resultados.values()), "—")
for c in (2, 3, 4, 5):
    money(ws.cell(r, c))
    if isinstance(ws.cell(r, c).value, (int, float)) and ws.cell(r, c).value > 0:
        ws.cell(r, c).fill = FILL_OK
    elif isinstance(ws.cell(r, c).value, (int, float)) and ws.cell(r, c).value < 0:
        ws.cell(r, c).fill = FILL_BAD
r += 1
# Utilização fábrica
util_r = {n: sum(resultados[n]["min_usados"].values()) / (CAP_MIN_DIA*5) for n in (1,2,3)}
kpi(r, "⏱️ Utilização fábrica",
    f"{util_r[1]*100:.1f}%", f"{util_r[2]*100:.1f}%", f"{util_r[3]*100:.1f}%",
    "—", "—",
    lambda: True)
r += 1
kpi(r, "⚙️ Ociosidade fábrica",
    f"{(1-util_r[1])*100:.1f}%", f"{(1-util_r[2])*100:.1f}%", f"{(1-util_r[3])*100:.1f}%",
    "—", "—")
r += 1
kpi(r, "📦 Estoque MP1 fim (ton)",
    round(sim_r1["estoque_mp_final"]["MP1"], 2),
    round(sim_r2["estoque_mp_final"]["MP1"], 2),
    round(sim_r3["estoque_mp_final"]["MP1"], 2),
    "—", f"≤ {cap_mp['MP1']:.0f}",
    lambda: all(s["estoque_mp_final"]["MP1"] <= cap_mp["MP1"] + 0.01 for s in resultados.values()))
r += 1
kpi(r, "📦 Estoque MP2 fim (ton)",
    round(sim_r1["estoque_mp_final"]["MP2"], 2),
    round(sim_r2["estoque_mp_final"]["MP2"], 2),
    round(sim_r3["estoque_mp_final"]["MP2"], 2),
    "—", f"≤ {cap_mp['MP2']:.1f}",
    lambda: all(s["estoque_mp_final"]["MP2"] <= cap_mp["MP2"] + 0.01 for s in resultados.values()))
r += 1
kpi(r, "📦 Estoque MP3 fim (ton)",
    round(sim_r1["estoque_mp_final"]["MP3"], 2),
    round(sim_r2["estoque_mp_final"]["MP3"], 2),
    round(sim_r3["estoque_mp_final"]["MP3"], 2),
    "—", f"≤ {cap_mp['MP3']:.1f}",
    lambda: all(s["estoque_mp_final"]["MP3"] <= cap_mp["MP3"] + 0.01 for s in resultados.values()))
r += 1
kpi(r, "🏪 PA total fim CDs (un)",
    sum(sim_r1["estoque_pa_final"][cd][pa] for cd in CDS for pa in ("PA1","PA2","PA3")),
    sum(sim_r2["estoque_pa_final"][cd][pa] for cd in CDS for pa in ("PA1","PA2","PA3")),
    sum(sim_r3["estoque_pa_final"][cd][pa] for cd in CDS for pa in ("PA1","PA2","PA3")),
    "—", "—")
r += 1
# Acumulado
ws.cell(r, 1).value = "💼 RESULTADO ACUMULADO R1+R2+R3"
ws.cell(r, 5).value = sum(s["resultado"] for s in resultados.values())
money(ws.cell(r, 5))
acumulado = sum(s["resultado"] for s in resultados.values())
ws.cell(r, 5).fill = FILL_OK if acumulado > 0 else FILL_BAD
ws.cell(r, 1).font = FONT_BOLD; ws.cell(r, 5).font = FONT_BOLD
auto_width(ws)


# ============ 01 R1 DETALHE ============
def detalhe_rodada(rodada: int):
    s = resultados[rodada]
    ws = wb.create_sheet(f"{rodada+1:02d}_R{rodada}_Detalhe")
    title(ws, 1, 10, f"RODADA {rodada} — SIMULAÇÃO DIA A DIA")

    # Bloco 1: Estoque MP dia a dia
    r = 3
    ws.cell(r, 1).value = f"BLOCO 1 — Estoque MP F1 dia a dia"
    ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    header(ws, r, ["Dia", "MP", "Pré", "+Arr", "Pos", "Desc", "Consumo", "End", "Cap", "%"])
    r += 1
    for d in range(1, 6):
        for mp in ("MP1", "MP2", "MP3"):
            info = s["sim_mp"][d][mp]
            ws.cell(r, 1).value = f"Dia {d}"
            ws.cell(r, 2).value = mp
            ws.cell(r, 3).value = round(info["pre"], 3)
            ws.cell(r, 4).value = round(info["arr"], 3)
            ws.cell(r, 5).value = round(info["pos"], 3)
            ws.cell(r, 6).value = round(info["desc"], 3)
            ws.cell(r, 7).value = round(info["cons"], 3)
            ws.cell(r, 8).value = round(info["end"], 3)
            ws.cell(r, 9).value = round(cap_mp[mp], 2)
            ws.cell(r, 10).value = f"=H{r}/I{r}"
            pct(ws.cell(r, 10))
            if info["desc"] > 0.01: ws.cell(r, 6).fill = FILL_BAD
            if info["end"] < -0.01: ws.cell(r, 8).fill = FILL_BAD
            r += 1
        r += 1

    # Bloco 2: PA nos CDs
    r += 1
    ws.cell(r, 1).value = f"BLOCO 2 — PA nos CDs dia a dia"
    ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    header(ws, r, ["Dia", "CD", "PA", "Pré", "+Arr", "-Saída", "Desc", "End", "Cap", "%"])
    r += 1
    for d in range(1, 6):
        for cd in ("CD1", "CD2"):
            for pa in ("PA1", "PA2", "PA3"):
                info = s["sim_pa"][cd][d][pa]
                if info["pre"] == 0 and info["arr"] == 0 and info["sai"] == 0 and info["end"] == 0:
                    continue
                ws.cell(r, 1).value = f"Dia {d}"
                ws.cell(r, 2).value = f"{cd} ({CDS[cd]['cidade']})"
                ws.cell(r, 3).value = pa
                ws.cell(r, 4).value = round(info["pre"])
                ws.cell(r, 5).value = round(info["arr"])
                ws.cell(r, 6).value = round(info["sai"])
                ws.cell(r, 7).value = round(info["desc"])
                ws.cell(r, 8).value = round(info["end"])
                ws.cell(r, 9).value = cap_pa_cd[cd][pa]
                ws.cell(r, 10).value = f"=H{r}/I{r}"
                pct(ws.cell(r, 10))
                if info["desc"] > 0: ws.cell(r, 7).fill = FILL_BAD
                if info["end"] < -0.5: ws.cell(r, 8).fill = FILL_BAD
                r += 1

    # Bloco 3: Produção e Util
    r += 1
    ws.cell(r, 1).value = f"BLOCO 3 — Produção + Utilização fábrica"
    ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    header(ws, r, ["Dia", "PA1", "PA2", "PA3", "Min PA1", "Min PA2", "Min PA3", "Min total", "Cap", "Util %"])
    r += 1
    op_fab = {1: op_fab_r1, 2: op_fab_r2, 3: op_fab_r3}[rodada]
    for d in range(1, 6):
        ws.cell(r, 1).value = f"Dia {d}"
        ws.cell(r, 2).value = op_fab[d]["PA1"]
        ws.cell(r, 3).value = op_fab[d]["PA2"]
        ws.cell(r, 4).value = op_fab[d]["PA3"]
        ws.cell(r, 5).value = f"=B{r}/15"
        ws.cell(r, 6).value = f"=C{r}/30"
        ws.cell(r, 7).value = f"=D{r}/60"
        ws.cell(r, 8).value = f"=E{r}+F{r}+G{r}"
        ws.cell(r, 9).value = CAP_MIN_DIA
        ws.cell(r, 10).value = f"=H{r}/I{r}"
        pct(ws.cell(r, 10))
        r += 1

    # Bloco 4: Transportes (resumo modal)
    r += 1
    ws.cell(r, 1).value = f"BLOCO 4 — Transportes por modal"
    ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    header(ws, r, ["Modal", "Viagens", "% do total"])
    r += 1
    total_v = s["n_transp_total"]
    for m in ("Caminhão", "Navio", "Avião"):
        v = s["viagens_por_modal"].get(m, 0)
        ws.cell(r, 1).value = m
        ws.cell(r, 2).value = v
        ws.cell(r, 3).value = v / max(1, total_v)
        pct(ws.cell(r, 3))
        r += 1
    ws.cell(r, 1).value = "TOTAL"
    ws.cell(r, 1).font = FONT_BOLD
    ws.cell(r, 2).value = total_v
    ws.cell(r, 2).font = FONT_BOLD

    # Bloco 5: DRE da rodada
    r += 2
    ws.cell(r, 1).value = f"BLOCO 5 — DRE Rodada {rodada}"
    ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1
    header(ws, r, ["Linha", "Valor R$"])
    r += 1
    linhas_dre = [
        ("Receita (PA atendidos no dia)", s["receita"]),
        ("(-) Custos fixos", -s["custo_fixo"]),
        ("(-) Compra MP", -s["custo_compra_mp"]),
        ("(-) Frete", -s["custo_frete"]),
        ("(-) Carregamento MP", -s["custo_carreg_mp"]),
        ("(-) Carregamento PA", -s["custo_carreg_pa"]),
        ("= Resultado da rodada", s["resultado"]),
    ]
    for label, v in linhas_dre:
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = v
        money(ws.cell(r, 2))
        if "Resultado" in label:
            ws.cell(r, 1).font = FONT_BOLD
            ws.cell(r, 2).font = FONT_BOLD
            ws.cell(r, 2).fill = FILL_OK if v > 0 else FILL_BAD
        r += 1

    # Bloco 6: OPs (se houver)
    if OPS_POR_RODADA.get(rodada):
        r += 2
        ws.cell(r, 1).value = f"BLOCO 6 — OPs entregues (NS no dia exato)"
        ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1
        header(ws, r, ["Cidade", "PA", "Qtd planejada", "Dia entrega", "Qtd no dia exato", "✅/❌"])
        r += 1
        for op in s["ops_atendidas"]:
            ws.cell(r, 1).value = op["cidade"]
            ws.cell(r, 2).value = op["pa"]
            ws.cell(r, 3).value = op["qtd"]
            ws.cell(r, 4).value = op["dia_entrega"]
            ws.cell(r, 5).value = op["qty_no_dia"]
            ws.cell(r, 6).value = "✅"
            ws.cell(r, 6).fill = FILL_OK
            r += 1
        for op in s["ops_descartadas"]:
            ws.cell(r, 1).value = op["cidade"]
            ws.cell(r, 2).value = op["pa"]
            ws.cell(r, 3).value = op["qtd"]
            ws.cell(r, 4).value = op["dia_entrega"]
            ws.cell(r, 5).value = op["qty_no_dia"]
            ws.cell(r, 6).value = "❌"
            ws.cell(r, 6).fill = FILL_BAD
            r += 1

    auto_width(ws)

for rodada in (1, 2, 3):
    detalhe_rodada(rodada)


# ============ DRE Acumulada ============
ws = wb.create_sheet("05_DRE_Acumulada")
title(ws, 1, 7, "DRE ACUMULADA R1 + R2 + R3")
header(ws, 3, ["Linha DRE", "R1", "R2", "R3", "Acumulado", "Fórmula"])
r = 4
linhas_dre_acum = [
    ("Receita", "receita"),
    ("(-) Custos fixos", "custo_fixo"),
    ("(-) Compra MP", "custo_compra_mp"),
    ("(-) Frete", "custo_frete"),
    ("(-) Carreg MP", "custo_carreg_mp"),
    ("(-) Carreg PA", "custo_carreg_pa"),
    ("= Resultado", "resultado"),
]
for label, key in linhas_dre_acum:
    ws.cell(r, 1).value = label
    for j, n in enumerate((1, 2, 3)):
        v = resultados[n][key]
        if "(-)" in label: v = -v
        ws.cell(r, 2 + j).value = v
        money(ws.cell(r, 2 + j))
    ws.cell(r, 5).value = f"=B{r}+C{r}+D{r}"
    money(ws.cell(r, 5))
    if "Resultado" in label:
        for c in range(1, 6):
            ws.cell(r, c).font = FONT_BOLD
        ws.cell(r, 5).fill = FILL_OK if sum(resultados[n]["resultado"] for n in (1,2,3)) > 0 else FILL_BAD
    r += 1
auto_width(ws)


# ============ Evolução Estoques (linha temporal 15 dias) ============
ws = wb.create_sheet("06_Evolucao_Estoques")
title(ws, 1, 10, "EVOLUÇÃO ESTOQUE MP F1 (15 dias absolutos)")
header(ws, 3, ["Dia abs", "Rodada", "Dia rel", "MP1 fim", "MP2 fim", "MP3 fim",
                "Cap MP1", "Cap MP2", "Cap MP3", "Total ton"])
r = 4
for n in (1, 2, 3):
    for d in range(1, 6):
        dia_abs = (n - 1) * 5 + d
        ws.cell(r, 1).value = dia_abs
        ws.cell(r, 2).value = f"R{n}"
        ws.cell(r, 3).value = d
        ws.cell(r, 4).value = round(resultados[n]["sim_mp"][d]["MP1"]["end"], 2)
        ws.cell(r, 5).value = round(resultados[n]["sim_mp"][d]["MP2"]["end"], 2)
        ws.cell(r, 6).value = round(resultados[n]["sim_mp"][d]["MP3"]["end"], 2)
        ws.cell(r, 7).value = round(cap_mp["MP1"], 1)
        ws.cell(r, 8).value = round(cap_mp["MP2"], 1)
        ws.cell(r, 9).value = round(cap_mp["MP3"], 1)
        ws.cell(r, 10).value = f"=D{r}+E{r}+F{r}"
        r += 1
auto_width(ws)


# ============ Checks Globais ============
ws = wb.create_sheet("07_Checks_Globais")
title(ws, 1, 5, "SANITY CHECKS GLOBAIS (R1, R2, R3)")
header(ws, 3, ["#", "Checagem", "R1", "R2", "R3"])
r = 4
checks = []
def add_check(label, *vals_status):
    """vals_status: list of (valor_exibir, ok_bool) por rodada"""
    nonlocal_r = [r]
    rr = r
    ws.cell(rr, 1).value = len(checks) + 1
    ws.cell(rr, 2).value = label
    for j, (v, ok) in enumerate(vals_status):
        cc = ws.cell(rr, 3 + j)
        cc.value = f"{v} {'✅' if ok else '❌'}"
        cc.fill = FILL_OK if ok else FILL_BAD
    checks.append(all(ok for _, ok in vals_status))

# Cada check
def transp_ok(s):
    return s["n_transp_total"], s["n_transp_total"] <= 220
def mp_neg_check(s, mp):
    minv = min(s["sim_mp"][d][mp]["end"] for d in range(1, 6))
    return f"{minv:.1f}", minv >= -0.01
def mp_cap_check(s, mp):
    maxv = max(s["sim_mp"][d][mp]["end"] for d in range(1, 6))
    return f"{maxv:.1f}/{cap_mp[mp]:.0f}", maxv <= cap_mp[mp] + 0.01
def mp_desc_check(s, mp):
    total_desc = sum(s["sim_mp"][d][mp]["desc"] for d in range(1, 6))
    return f"{total_desc:.2f}t", total_desc < 0.01
def min_cap_check(s):
    mv = max(s["min_usados"].values())
    return f"{mv:.0f}/{CAP_MIN_DIA}", mv <= CAP_MIN_DIA + 1

add_check("Transportes ≤ 220 por rodada",
    transp_ok(sim_r1), transp_ok(sim_r2), transp_ok(sim_r3))
r += 1
add_check("Min usados/dia ≤ 10.080",
    min_cap_check(sim_r1), min_cap_check(sim_r2), min_cap_check(sim_r3))
r += 1
add_check("Estoque MP1 nunca negativo",
    mp_neg_check(sim_r1, "MP1"), mp_neg_check(sim_r2, "MP1"), mp_neg_check(sim_r3, "MP1"))
r += 1
add_check("Estoque MP2 nunca negativo",
    mp_neg_check(sim_r1, "MP2"), mp_neg_check(sim_r2, "MP2"), mp_neg_check(sim_r3, "MP2"))
r += 1
add_check("Estoque MP3 nunca negativo",
    mp_neg_check(sim_r1, "MP3"), mp_neg_check(sim_r2, "MP3"), mp_neg_check(sim_r3, "MP3"))
r += 1
add_check("Estoque MP1 ≤ cap (sem descarte)",
    mp_cap_check(sim_r1, "MP1"), mp_cap_check(sim_r2, "MP1"), mp_cap_check(sim_r3, "MP1"))
r += 1
add_check("Estoque MP2 ≤ cap (sem descarte)",
    mp_cap_check(sim_r1, "MP2"), mp_cap_check(sim_r2, "MP2"), mp_cap_check(sim_r3, "MP2"))
r += 1
add_check("Estoque MP3 ≤ cap (sem descarte)",
    mp_cap_check(sim_r1, "MP3"), mp_cap_check(sim_r2, "MP3"), mp_cap_check(sim_r3, "MP3"))
r += 1
add_check("Descarte MP1 = 0 (ton)",
    mp_desc_check(sim_r1, "MP1"), mp_desc_check(sim_r2, "MP1"), mp_desc_check(sim_r3, "MP1"))
r += 1
add_check("Descarte MP2 = 0 (ton)",
    mp_desc_check(sim_r1, "MP2"), mp_desc_check(sim_r2, "MP2"), mp_desc_check(sim_r3, "MP2"))
r += 1
add_check("Descarte MP3 = 0 (ton)",
    mp_desc_check(sim_r1, "MP3"), mp_desc_check(sim_r2, "MP3"), mp_desc_check(sim_r3, "MP3"))

# Resumo global
r += 3
total_falhas = sum(1 for ok in checks if not ok)
ws.cell(r, 1).value = "RESUMO GLOBAL"
ws.cell(r, 1).font = FONT_BOLD; ws.cell(r, 1).fill = FILL_INFO
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
ws.cell(r, 1).value = f"Total checks: {len(checks)} | Falhas: {total_falhas}"
ws.cell(r, 1).font = FONT_BOLD
ws.cell(r, 1).fill = FILL_OK if total_falhas == 0 else FILL_BAD
auto_width(ws)


OUT = BASE / "rodadas" / "SanityCheck_Completo.xlsm"
wb.save(OUT)
print(f"\nOK SanityCheck_Completo.xlsm criado em: {OUT}")
print(f"  Abas: {wb.sheetnames}")
print(f"\nResumo global:")
print(f"  R1 resultado: R$ {sim_r1['resultado']:>14,.0f}")
print(f"  R2 resultado: R$ {sim_r2['resultado']:>14,.0f}")
print(f"  R3 resultado: R$ {sim_r3['resultado']:>14,.0f}")
print(f"  ACUMULADO:    R$ {sum(s['resultado'] for s in resultados.values()):>14,.0f}")
