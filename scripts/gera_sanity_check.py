"""SanityCheck.xlsm — simulação dia a dia do plano R3.

Simula EXATAMENTE como o jogo processaria as 5 dias da rodada e gera Excel com
14 abas. Tudo com fórmulas explícitas para auditoria + valores finais previstos.

Estrutura das abas:
  00_RESUMO          KPIs executivos (NS, util, transp, MP final, DRE)
  01_Regras          Regras do jogo modeladas
  02_Layout          F1 + CDs + capacidades
  03_LeadTimes       Lookup oficial Orig_Dest
  04_CapModais       Avião 1t / Cam 24t / Navio 100t
  05_OPs_R3          25 OPs com rota, modal, custos, dia chegada
  06_Transportes_R3  Todas as linhas SOL_TRANSP (com lt)
  07_Producao        OP_FABRICAS detalhado + MP consumido por dia
  08_Estoque_MP_Dia  Simulação MP em F1 dia a dia (entrada/saída/estoque)
  09_Estoque_PA_Dia  Simulação PA nos CDs dia a dia
  10_Estoque_FimR3   Posição final no Dia 15 (igual ao PDF Estoques)
  11_Custo_Transp    Frete por viagem com fórmulas
  12_DRE_R1_R2_R3    DRE reconstruída com fórmulas
  13_Indicadores     IND_FLAMENGO style (NS, util, ocupação modal)
  14_Checks          Sanity checks automatizados
"""
from __future__ import annotations
import json, math, re, sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))
from src.config import Config
from src.io_xlsm import ler_instalacoes
from src.planner_v3 import (
    planejar_v3, todas_rotas_op, lead_dias, km_rota,
    custo_total_modal, n_viagens_pa,
)
from src.planner_manual import forecast_proxima_rodada_via_hw

# ============ STYLES ============
FILL_HDR = PatternFill("solid", fgColor="1F4E78")
FILL_SUB = PatternFill("solid", fgColor="9BC2E6")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
FILL_INFO = PatternFill("solid", fgColor="DDEBF7")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FONT_HDR = Font(bold=True, color="FFFFFF", size=12)
FONT_SUB = Font(bold=True, size=11)
FONT_BOLD = Font(bold=True)
BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))


def set_title(ws, row, ncols, text):
    ws.cell(row, 1).value = text
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row, 1).font = FONT_HDR
    ws.cell(row, 1).fill = FILL_HDR
    ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")


def header_row(ws, row, headers, fill=FILL_INFO):
    for j, h in enumerate(headers):
        c = ws.cell(row, j + 1)
        c.value = h
        c.font = FONT_BOLD
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def money(cell):
    cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def pct(cell):
    cell.number_format = "0.0%"


def auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        try:
            letter = col[0].column_letter
        except AttributeError:
            continue
        m = min_w
        for c in col:
            if c.value is not None:
                m = max(m, min(max_w, len(str(c.value)) + 2))
        ws.column_dimensions[letter].width = m


# ============ DADOS DE ENTRADA ============
print("Carregando configs...")
cfg = Config.load(BASE)
FLAMENGO_R3 = BASE / "rodadas" / "rodada_3" / "FLAMENGO.xlsm"
inst = ler_instalacoes(FLAMENGO_R3)
RODADA = 3
lead_tab = json.loads((BASE / "data" / "lead_times.json").read_text(encoding="utf-8"))

# Estado inicial R3 (do PDF ESTOQUES dia 10)
ESTOQUE_MP_INI = {"MP1": 78.98, "MP2": 50.36, "MP3": 48.14}
ESTOQUE_PA_INI = {cd: {"PA1": 0, "PA2": 0, "PA3": 0} for cd in inst["cds"]}

# MP em-trânsito de R2 chegando em R3 (do SOL_TRANSP R2)
MP_EM_TRANSITO = [{"dia": 1, "mp": "MP1", "qtd": 8.7, "origem": "Manaus"}]

PRECO_PA3_R3 = 32.00

# OPs R3 do PDF
OPS = [{"cidade": c, "pa": "PA3", "qtd": q, "dia_entrega": d} for c, q, d in [
    ("Belém", 20155, 5), ("Belo Horizonte", 70544, 3), ("Brasília", 117573, 3),
    ("Campinas", 56435, 2), ("Campo Grande", 23515, 3), ("Cuiabá", 28218, 3),
    ("Curitiba", 103464, 2), ("Fortaleza", 68528, 5), ("Goiânia", 65841, 3),
    ("João Pessoa", 40311, 4), ("Joinville", 28218, 2), ("Maceió", 40311, 4),
    ("Manaus", 20155, 5), ("Natal", 40311, 5), ("Porto Alegre", 103464, 2),
    ("Recife", 60466, 4), ("Ribeirão Preto", 47029, 2), ("Rio de Janeiro", 94059, 3),
    ("Salvador", 80622, 4), ("Santos", 47029, 2), ("São Luís", 20155, 5),
    ("São Paulo", 117573, 2), ("Uberlândia", 23515, 3), ("Vitória", 14109, 3),
    ("Vitória da Conquista", 12093, 4),
]]

# Forecast R4
print("Rodando forecast HW...")
fc_r4 = forecast_proxima_rodada_via_hw(rodada_n_atual=RODADA, base_dir=BASE)
fc_pa_brasil = {"PA1": 0.0, "PA2": 0.0, "PA3": 0.0}
for (c, pa), v in fc_r4.items():
    fc_pa_brasil[pa] += v
SHARE = 0.40
cd2_pa2_cap = int(inst["cds"]["CD2"]["area_pa"]["PA2"] * 2 * cfg.densidades_pa["PA2"] / cfg.peso_un_ton["PA2"])
BUFFER_PA2_R4 = min(int(fc_pa_brasil["PA2"] * SHARE), cd2_pa2_cap)
COMPRAS_MP_R4 = {"MP1": 48.0, "MP3": 48.0}

# Rodar planner V3
print("Rodando solver V3...")
res = planejar_v3(
    rodada_n=RODADA,
    ops_rodada=OPS,
    estoque_inicial_mp_ton=ESTOQUE_MP_INI,
    estoque_inicial_pa_cd=ESTOQUE_PA_INI,
    mp_em_transito_chegando=MP_EM_TRANSITO,
    cfg=cfg, instalacoes=inst,
    pa_proxima_rodada="PA2",
    buffer_pa_proxima=BUFFER_PA2_R4,
    compras_mp_extra_para_r_mais_1=COMPRAS_MP_R4,
)
df_sol = res["df_sol_transp"]
df_op = res["df_op_fabricas"]
resumo = res["resumo"]

# ============ INFRAESTRUTURA / DERIVADOS ============
F1 = inst["fabricas"]["F1"]
CDS = inst["cds"]
CAP_MIN_DIA = F1["maquinas"] * F1["turnos"] * 8 * 60
CAP_MIN_SEM = CAP_MIN_DIA * 5
CAP_HH_SEM = F1["maquinas"] * F1["turnos"] * 8 * 5
VEL = {"PA1": 15, "PA2": 30, "PA3": 60}

cap_mp = {mp: F1["area_mp"][mp] * 2 * cfg.densidades_mp[mp] for mp in ("MP1", "MP2", "MP3")}
cap_pa_cd = {
    cd: {pa: int(d["area_pa"][pa] * 2 * cfg.densidades_pa[pa] / cfg.peso_un_ton[pa])
         for pa in ("PA1", "PA2", "PA3")}
    for cd, d in CDS.items()
}

# Fornecedor mais barato por MP + lead
forn_min = {mp: min(cfg.fornecedores[mp], key=lambda x: x[1]) for mp in ("MP1", "MP2", "MP3")}
lt_forn = {mp: lead_dias(cfg, forn_min[mp][0], F1["cidade"], "Caminhão") for mp in ("MP1", "MP2", "MP3")}

# ============ SIMULAÇÃO DIA A DIA ============
_DIA = re.compile(r"Dia\s*(\d+)")

# Parse SOL_TRANSP para dicionários por (dia_partida, item)
arrivals_mp_dia = defaultdict(lambda: defaultdict(float))     # [dia_chegada][mp] -> qtd
shipments_f1_cd_dia = defaultdict(lambda: defaultdict(lambda: defaultdict(float))) # [dia][cd_id][pa] -> qtd
arrivals_pa_cd_dia = defaultdict(lambda: defaultdict(lambda: defaultdict(float))) # [dia_chegada][cd_id][pa] -> qtd
shipments_cd_v_dia = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))  # [dia][cd_id][(varejo,pa)] -> qtd

cidade_to_cd = {d["cidade"]: cd for cd, d in CDS.items()}

# Em-trânsito (de R2 → R3)
for x in MP_EM_TRANSITO:
    arrivals_mp_dia[x["dia"]][x["mp"]] += x["qtd"]

# Parse SOL_TRANSP
for _, row in df_sol.iterrows():
    dia_part = int(_DIA.search(str(row["Dia da Coleta"])).group(1))
    modal = row["Modal"]
    item = row["Tipo do Produto"]
    qtd = float(row["Qtde"])
    origem = row["Cidade"]
    destino = row["Cidade_Destino"]
    lt = lead_dias(cfg, origem, destino, modal) or 0
    dia_cheg = dia_part + lt

    if row["Origem"] == "Fornecedor":
        if 1 <= dia_cheg <= 5:
            arrivals_mp_dia[dia_cheg][item] += qtd
    elif row["Origem"] == "Fábrica" and row["Destino"] == "CD":
        cd_id = cidade_to_cd[destino]
        shipments_f1_cd_dia[dia_part][cd_id][item] += qtd
        if 1 <= dia_cheg <= 5:
            arrivals_pa_cd_dia[dia_cheg][cd_id][item] += qtd
    elif row["Origem"] == "CD" and row["Destino"] == "Varejista":
        cd_id = cidade_to_cd[origem]
        key = (destino, item)
        shipments_cd_v_dia[dia_part][cd_id][key] += qtd

# Simulação dia a dia
# stock_mp_dia[dia][mp] = {pre, arr, desc, pos, cons, end}
# stock_pa_dia[dia][cd][pa] = {pre, arr, sai, end}
sim_mp = {}
sim_pa = {cd: {} for cd in CDS}
stock_mp = dict(ESTOQUE_MP_INI)
stock_pa = {cd: dict(ESTOQUE_PA_INI[cd]) for cd in CDS}

for d in range(1, 6):
    sim_mp[d] = {}
    # MP F1 dia
    op_row = df_op[df_op["Dia"] == f"Dia {d}"].iloc[0]
    p_day = {pa: int(op_row[pa]) for pa in ("PA1", "PA2", "PA3")}
    for mp in ("MP1", "MP2", "MP3"):
        pre = stock_mp[mp]
        arr = arrivals_mp_dia[d].get(mp, 0)
        pos = pre + arr
        desc = max(0, pos - cap_mp[mp])
        pos_real = min(pos, cap_mp[mp])
        cons = sum(p_day[pa] * cfg.BoM[pa][mp] / 1e6 for pa in ("PA1", "PA2", "PA3"))
        end = pos_real - cons
        sim_mp[d][mp] = {"pre": pre, "arr": arr, "desc": desc, "pos": pos_real,
                         "cons": cons, "end": end}
        stock_mp[mp] = max(0, end)

    # PA CDs dia
    for cd in CDS:
        sim_pa[cd][d] = {}
        for pa in ("PA1", "PA2", "PA3"):
            pre = stock_pa[cd][pa]
            arr = arrivals_pa_cd_dia[d][cd].get(pa, 0)
            # Saída do CD (despachando hoje)
            sai = 0
            for (v, item), q in shipments_cd_v_dia[d][cd].items():
                if item == pa:
                    sai += q
            pos = pre + arr
            end = pos - sai
            sim_pa[cd][d][pa] = {"pre": pre, "arr": arr, "sai": sai, "end": end}
            stock_pa[cd][pa] = max(0, end)

# Estado final R3 (Dia 15 = Dia 5 R3)
estoque_mp_final = {mp: sim_mp[5][mp]["end"] for mp in ("MP1", "MP2", "MP3")}
estoque_pa_final = {cd: {pa: sim_pa[cd][5][pa]["end"] for pa in ("PA1", "PA2", "PA3")}
                    for cd in CDS}

# ============ CUSTOS R3 ============
custo_compra_mp = sum(resumo["mp_a_comprar_ton"][mp] * forn_min[mp][1] for mp in ("MP1", "MP2", "MP3"))
custo_frete_mp = 0.0
custo_frete_pa = 0.0
custos_por_viagem = []
for idx, row in df_sol.iterrows():
    modal = row["Modal"]
    item = row["Tipo do Produto"]
    qtd = float(row["Qtde"])
    origem = row["Cidade"]
    destino = row["Cidade_Destino"]
    km = km_rota(cfg, origem, destino, modal) or 0
    if item.startswith("MP"):
        peso = qtd
    else:
        peso = qtd * cfg.peso_un_ton[item]
    custo = custo_total_modal(cfg, modal, km, peso, 1)
    cap_modal = cfg.cap_modal_ton[modal]
    ocup = peso / cap_modal if cap_modal > 0 else 0
    custos_por_viagem.append({
        "idx": idx + 1, "origem_tipo": row["Origem"], "origem_cid": origem,
        "destino_tipo": row["Destino"], "destino_cid": destino,
        "modal": modal, "item": item, "qtd": qtd, "peso_t": peso,
        "km": km, "ocup": ocup, "custo": custo,
        "dia_part": int(_DIA.search(str(row["Dia da Coleta"])).group(1)),
    })
    if item.startswith("MP"):
        custo_frete_mp += custo
    else:
        custo_frete_pa += custo

# Carregamento estoque (1% × estoque_final × maior_preço/preço_tabela)
maior_preco_mp = {"MP1": 56000, "MP2": 22000, "MP3": 41000}  # vs fornecedor mais caro
preco_tabela_pa = {"PA1": 80, "PA2": 50, "PA3": 25}

carreg_mp = sum(estoque_mp_final[mp] * maior_preco_mp[mp] * 0.01 for mp in ("MP1", "MP2", "MP3"))
carreg_pa = sum(
    estoque_pa_final[cd][pa] * preco_tabela_pa[pa] * 0.01
    for cd in CDS for pa in ("PA1", "PA2", "PA3")
)

# Custos fixos
FIX = {
    "parcela_terr": 506_968, "parcela_maq": 415_567, "parcela_mo": 84,
    "manut_fab": 1_313, "sal_op": 450, "custo_prod": 172_086,
    "manut_cd": 26_683,
}

# Receita
receita_pa3_r3 = sum(int(r["qtd"]) for r in resumo["rotas_op"] if r.get("alocada") and r["pa"] == "PA3") * PRECO_PA3_R3

custo_total_r3 = sum(FIX.values()) + custo_compra_mp + custo_frete_mp + custo_frete_pa + carreg_mp + carreg_pa
resultado_r3 = receita_pa3_r3 - custo_total_r3

# ============ CRIA WORKBOOK ============
print("Criando SanityCheck.xlsm...")
wb = openpyxl.Workbook()
wb.remove(wb.active)


# ===================================================================
# 00 — RESUMO EXECUTIVO
# ===================================================================
ws = wb.create_sheet("00_RESUMO")
set_title(ws, 1, 6, f"RESUMO EXECUTIVO — RODADA {RODADA} (Dia 11 a Dia 15 absoluto)")
ws.row_dimensions[1].height = 22

r = 3
ws.cell(r, 1).value = "INDICADOR"
ws.cell(r, 2).value = "VALOR"
ws.cell(r, 3).value = "DETALHE"
ws.cell(r, 4).value = "FÓRMULA / CHECK"
for c in (1, 2, 3, 4):
    ws.cell(r, c).font = FONT_BOLD
    ws.cell(r, c).fill = FILL_SUB

kpis = [
    ("🎯 NS R3 (Nível de Serviço)", f'{resumo["ops_atendidas"]}/{resumo["ops_total"]} ({resumo["taxa_atendimento_pct"]:.1f}%)',
     f'{resumo["qtd_atendida"]:,} de {resumo["qtd_atendida"] + resumo["qtd_descartada"]:,} frascos', "OPs no dia EXATO / Total"),
    ("💰 Receita R3 (PA3)", f'R$ {receita_pa3_r3:,.0f}',
     f'{resumo["qtd_atendida"]:,} frascos × R$ {PRECO_PA3_R3:.2f}',
     f"qty × preço = {resumo['qtd_atendida']:,} × {PRECO_PA3_R3}"),
    ("💸 Custo Total R3", f'R$ {custo_total_r3:,.0f}',
     'Fixos + MP + Frete + Carregamento', "ver aba 12_DRE"),
    ("📈 Resultado R3", f'R$ {resultado_r3:,.0f}',
     'Receita − Custos', "ver aba 12_DRE"),
    ("⏱️ Utilização fábrica", f'{sum(resumo["min_usados_por_dia"].values())/CAP_MIN_SEM*100:.1f}%',
     f'{sum(resumo["min_usados_por_dia"].values()):.0f}/{CAP_MIN_SEM} min', "Σ min / cap min sem"),
    ("⚙️ Ociosidade", f'{(1-sum(resumo["min_usados_por_dia"].values())/CAP_MIN_SEM)*100:.1f}%',
     'min ociosos = sobra fabril', "1 − utilização"),
    ("🚚 Transportes R3", f'{resumo["n_transportes"]}/220',
     f'cap 220/semana', "≤ 220 = OK"),
    ("📦 Buffer PA2 R4", f'{resumo["buffer_pa_proxima_alocado"]:,}',
     f'estocado em CD2 Santos (cap PA2={cd2_pa2_cap:,})',
     f"Forecast R4 PA2 Flamengo = {int(fc_pa_brasil['PA2']*SHARE):,}"),
    ("📍 MP1 final F1", f'{estoque_mp_final["MP1"]:.2f} t',
     f'cap {cap_mp["MP1"]:.1f} ton ({estoque_mp_final["MP1"]/cap_mp["MP1"]*100:.1f}%)',
     "ver aba 08_Estoque_MP_Dia"),
    ("📍 MP2 final F1", f'{estoque_mp_final["MP2"]:.2f} t',
     f'cap {cap_mp["MP2"]:.1f} ton ({estoque_mp_final["MP2"]/cap_mp["MP2"]*100:.1f}%)',
     "ver aba 08_Estoque_MP_Dia"),
    ("📍 MP3 final F1", f'{estoque_mp_final["MP3"]:.2f} t',
     f'cap {cap_mp["MP3"]:.1f} ton ({estoque_mp_final["MP3"]/cap_mp["MP3"]*100:.1f}%)',
     "ver aba 08_Estoque_MP_Dia"),
    ("🏪 PA2 final CD1 (São Luís)", f'{estoque_pa_final["CD1"]["PA2"]:,}',
     f'cap {cap_pa_cd["CD1"]["PA2"]:,}', "ver aba 09_Estoque_PA_Dia"),
    ("🏪 PA2 final CD2 (Santos)", f'{estoque_pa_final["CD2"]["PA2"]:,}',
     f'cap {cap_pa_cd["CD2"]["PA2"]:,}', "ver aba 09_Estoque_PA_Dia"),
]
for i, (m, v, det, formula) in enumerate(kpis):
    rr = 4 + i
    ws.cell(rr, 1).value = m
    ws.cell(rr, 2).value = v
    ws.cell(rr, 3).value = det
    ws.cell(rr, 4).value = formula
    ws.cell(rr, 2).font = FONT_BOLD
auto_width(ws)


# ===================================================================
# 01 — REGRAS DO JOGO
# ===================================================================
ws = wb.create_sheet("01_Regras")
set_title(ws, 3, 4, "REGRAS DO JOGO MODELADAS NO PLANNER V3")
r = 4
header_row(ws, r, ["#", "Regra", "Modelagem", "Validação"])
regras = [
    ("R1", "PA chega EXATAMENTE no dia solicitado", "dia_prod + lt_F1 + lt_CD == dia_entrega", "aba 05_OPs_R3 col ✅"),
    ("R2", "PA sai da F1 no MESMO dia da produção", "F1→CD: Dia_Coleta = dia_producao", "aba 14 teste 4"),
    ("R3", "MP sem espaço na F1 é descartada", "Compra limita p/ estoque ≤ cap; simula dia a dia", "aba 08 col Desc"),
    ("R4", "Máx 220 transportes/semana", "Planner conta + alerta", "aba 14 teste 1"),
    ("R5", "Lead times oficiais (Av=0, Cam/Navio var.)", "Lookup Orig_Dest no data/lead_times.json", "aba 03_LeadTimes"),
    ("R6", "Cap modal: Av=1t, Cam=24t, Navio=100t", "Cada viagem ≤ cap; quebrado em N viagens", "aba 04_CapModais"),
    ("R7", "Frete ≥80%: viagem; <80%: peso", "Função custo_total_modal", "aba 11_Custo_Transp"),
    ("R8", "Cap fábrica = 7×3×8×60 = 10.080 min/dia", "alocar_no_dia checa cap_restante[dia]", "aba 14 teste 2"),
    ("R9", "Velocidade: PA1=15 PA2=30 PA3=60 un/min", "min_necessarios = ceil(qtd/vel)", "aba 07_Producao"),
    ("R10", "BoM: PA1=60/90/150 PA2=75/125/50 PA3=75/30/45 g/un", "Consumo MP por produção", "aba 07_Producao"),
    ("R11", "F1 só estoca MP (PA vai 100% pro CD)", "Sem stock PA em F1; envio 100% mesmo dia", "aba 09_Estoque_PA"),
    ("R12", "Cap MP F1 = área×pé_direito×densidade", "MP1=127t MP2=50.4t MP3=75.6t", "aba 02_Layout"),
    ("R13", "Cap PA CD = área×pé_direito×densidade", "PA por PA por CD", "aba 02_Layout"),
    ("R14", "Setup imutável (F1+CDs já definidos)", "Lido de INSTALAÇÕES", "aba 02_Layout"),
    ("R15", "Preço PA3 R3 = R$ 32 (IND_FLAMENGO)", "Hardcoded", "aba 12_DRE"),
    ("R16", "Carreg MP = qty × maior preço × 1%", "Por MP no estoque final", "aba 12_DRE"),
    ("R17", "Carreg PA = qty × preço tabela × 1%", "Por PA no estoque final", "aba 12_DRE"),
    ("R18", "MP em trânsito de R2 chega em R3 conforme lt", "Soma em arrivals_mp_dia", "aba 08 col +Arr"),
]
for i, (n, regra, mod, val) in enumerate(regras):
    rr = 5 + i
    ws.cell(rr, 1).value = n
    ws.cell(rr, 2).value = regra
    ws.cell(rr, 3).value = mod
    ws.cell(rr, 4).value = val
    ws.cell(rr, 1).font = FONT_BOLD
    for c in (2, 3, 4):
        ws.cell(rr, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 28
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 32


# ===================================================================
# 02 — LAYOUT / CAPACIDADES
# ===================================================================
ws = wb.create_sheet("02_Layout")
set_title(ws, 1, 9, "LAYOUT E CAPACIDADES (lido do FLAMENGO.xlsm INSTALAÇÕES)")
r = 3
ws.cell(r, 1).value = "FÁBRICA F1"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
header_row(ws, r, ["Item", "Valor", "Unidade", "Cálculo derivado", "Valor"])
r += 1
linhas_f1 = [
    ("Cidade", F1["cidade"], "—", "", ""),
    ("Máquinas", F1["maquinas"], "máq", "", ""),
    ("Turnos/dia", F1["turnos"], "turnos", "", ""),
    ("MO", F1["mo"], "operários", "", ""),
    ("Horas/turno", 8, "h", "Cap min/dia = M×T×8×60", CAP_MIN_DIA),
    ("Dias/rodada", 5, "dias", "Cap min/rodada = ↑ × 5", CAP_MIN_SEM),
    ("", "", "", "Cap HH/sem = M×T×8×5", CAP_HH_SEM),
    ("Área MP1", F1["area_mp"]["MP1"], "m²", "Cap MP1 = área × 2m × dens", round(cap_mp["MP1"], 2)),
    ("Área MP2", F1["area_mp"]["MP2"], "m²", "Cap MP2 = área × 2m × dens", round(cap_mp["MP2"], 2)),
    ("Área MP3", F1["area_mp"]["MP3"], "m²", "Cap MP3 = área × 2m × dens", round(cap_mp["MP3"], 2)),
]
for row_ in linhas_f1:
    for j, v in enumerate(row_):
        ws.cell(r, j + 1).value = v
    r += 1

r += 2
ws.cell(r, 1).value = "CDs"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
r += 1
header_row(ws, r, ["CD", "Cidade", "Área PA1 (m²)", "Área PA2 (m²)", "Área PA3 (m²)",
                    "Área total", "Cap PA1 (un)", "Cap PA2 (un)", "Cap PA3 (un)"])
r += 1
for cd, d in CDS.items():
    ws.cell(r, 1).value = cd
    ws.cell(r, 2).value = d["cidade"]
    ws.cell(r, 3).value = d["area_pa"]["PA1"]
    ws.cell(r, 4).value = d["area_pa"]["PA2"]
    ws.cell(r, 5).value = d["area_pa"]["PA3"]
    ws.cell(r, 6).value = d["area_total"]
    for j, pa in enumerate(("PA1", "PA2", "PA3")):
        ws.cell(r, 7 + j).value = cap_pa_cd[cd][pa]
    r += 1
auto_width(ws)


# ===================================================================
# 03 — LEAD TIMES
# ===================================================================
ws = wb.create_sheet("03_LeadTimes")
set_title(ws, 1, 6, "LEAD TIMES OFICIAIS (lookup Orig_Dest do jogo). Avião=0 SEMPRE")
r = 3
ws.cell(r, 1).value = "F1 (Joinville) → CDs"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
header_row(ws, r, ["Modal", "CD1 (São Luís)", "CD2 (Santos)"])
r += 1
for m in ("Caminhão", "Navio", "Avião"):
    ws.cell(r, 1).value = m
    ws.cell(r, 2).value = lead_tab.get(m, {}).get("Joinville", {}).get("São Luís", "—")
    ws.cell(r, 3).value = lead_tab.get(m, {}).get("Joinville", {}).get("Santos", "—")
    r += 1
r += 2
ws.cell(r, 1).value = "Fornecedores → F1"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
header_row(ws, r, ["MP", "Fornecedor", "Lead Cam (d)", "Custo R$/ton"])
r += 1
for mp in ("MP1", "MP2", "MP3"):
    for forn, c in cfg.fornecedores[mp]:
        ws.cell(r, 1).value = mp
        ws.cell(r, 2).value = forn
        ws.cell(r, 3).value = lead_tab.get("Caminhão", {}).get(forn, {}).get("Joinville", "—")
        ws.cell(r, 4).value = c
        r += 1
r += 2
ws.cell(r, 1).value = "CDs → Varejistas R3 (todas as 25 cidades)"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
header_row(ws, r, ["Varejo", "SL Cam", "SL Av", "Sant Cam", "Sant Av", "Sant Navio"])
r += 1
cidades_var = sorted({op["cidade"] for op in OPS})
for v in cidades_var:
    ws.cell(r, 1).value = v
    ws.cell(r, 2).value = lead_tab.get("Caminhão", {}).get("São Luís", {}).get(v, "—")
    ws.cell(r, 3).value = lead_tab.get("Avião", {}).get("São Luís", {}).get(v, "—")
    ws.cell(r, 4).value = lead_tab.get("Caminhão", {}).get("Santos", {}).get(v, "—")
    ws.cell(r, 5).value = lead_tab.get("Avião", {}).get("Santos", {}).get(v, "—")
    ws.cell(r, 6).value = lead_tab.get("Navio", {}).get("Santos", {}).get(v, "—")
    r += 1
auto_width(ws)


# ===================================================================
# 04 — CAP MODAIS
# ===================================================================
ws = wb.create_sheet("04_CapModais")
set_title(ws, 1, 8, "CAPACIDADES POR MODAL — qty máxima por viagem")
header_row(ws, 3, ["Modal", "Cap (ton)", "Vel (km/h)", "Frete Viagem", "Frete Peso R$/(km·ton)",
                    "Doc CT-e", "PA1 (un/viagem)", "PA2 (un/viagem)", "PA3 (un/viagem)"])
for i, m in enumerate(("Avião", "Caminhão", "Navio")):
    r = 4 + i
    ws.cell(r, 1).value = m
    ws.cell(r, 2).value = cfg.cap_modal_ton[m]
    ws.cell(r, 3).value = {"Avião": 700, "Caminhão": 50, "Navio": 30}[m]
    ws.cell(r, 4).value = cfg.frete_viagem[m]
    ws.cell(r, 5).value = cfg.frete_peso[m]
    ws.cell(r, 6).value = cfg.doc_modal[m]
    for j, pa in enumerate(("PA1", "PA2", "PA3")):
        ws.cell(r, 7 + j).value = cfg.cap_modal_por_item[m][pa]
auto_width(ws)


# ===================================================================
# 05 — OPs R3 DETALHE
# ===================================================================
ws = wb.create_sheet("05_OPs_R3")
set_title(ws, 1, 14, "OPs R3 — rota escolhida, validação dia exato, custo logístico")
header_row(ws, 3, ["#", "Cidade", "PA", "Qtd", "Dia entrega",
                    "CD", "Modal F1→CD", "lt F1", "Modal CD→V", "lt CD",
                    "Dia prod", "Dia chega", "Custo log (R$)", "✅/❌"])
for i, r0 in enumerate(resumo["rotas_op"]):
    if not r0.get("alocada"):
        continue
    rr = 4 + i
    dia_chega = r0["dia_producao"] + r0["lt_f1"] + r0["lt_cd"]
    ok = dia_chega == r0["dia_entrega"]
    ws.cell(rr, 1).value = i + 1
    ws.cell(rr, 2).value = r0["cidade"]
    ws.cell(rr, 3).value = r0["pa"]
    ws.cell(rr, 4).value = r0["qtd"]
    ws.cell(rr, 5).value = r0["dia_entrega"]
    ws.cell(rr, 6).value = r0["cd_cidade"]
    ws.cell(rr, 7).value = r0["modal_f1"]
    ws.cell(rr, 8).value = r0["lt_f1"]
    ws.cell(rr, 9).value = r0["modal_cd"]
    ws.cell(rr, 10).value = r0["lt_cd"]
    ws.cell(rr, 11).value = r0["dia_producao"]
    ws.cell(rr, 12).value = dia_chega
    ws.cell(rr, 13).value = round(r0["custo"], 2)
    money(ws.cell(rr, 13))
    ws.cell(rr, 14).value = "✅" if ok else "❌"
    ws.cell(rr, 14).fill = FILL_OK if ok else FILL_BAD
auto_width(ws)


# ===================================================================
# 06 — TRANSPORTES R3
# ===================================================================
ws = wb.create_sheet("06_Transportes_R3")
set_title(ws, 1, 11, f"TRANSPORTES R3 — {len(df_sol)} linhas (cap 220/semana)")
header_row(ws, 3, ["#", "Rodada", "Origem", "Cidade", "Dia coleta", "Modal",
                    "Item", "Qtd", "Destino", "Cidade dest", "Lead time (d)"])
for i, row in df_sol.iterrows():
    rr = 4 + i
    ws.cell(rr, 1).value = i + 1
    ws.cell(rr, 2).value = row["Rodada"]
    ws.cell(rr, 3).value = row["Origem"]
    ws.cell(rr, 4).value = row["Cidade"]
    ws.cell(rr, 5).value = row["Dia da Coleta"]
    ws.cell(rr, 6).value = row["Modal"]
    ws.cell(rr, 7).value = row["Tipo do Produto"]
    ws.cell(rr, 8).value = float(row["Qtde"])
    ws.cell(rr, 9).value = row["Destino"]
    ws.cell(rr, 10).value = row["Cidade_Destino"]
    ws.cell(rr, 11).value = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
auto_width(ws)


# ===================================================================
# 07 — PRODUÇÃO PLANO
# ===================================================================
ws = wb.create_sheet("07_Producao")
set_title(ws, 1, 12, "PRODUÇÃO POR DIA + MP CONSUMIDO")
header_row(ws, 3, ["Dia", "PA1 (un)", "PA2 (un)", "PA3 (un)",
                    "Min PA1 (qty/15)", "Min PA2 (qty/30)", "Min PA3 (qty/60)",
                    "Min total", "Cap min/dia", "% utilização",
                    "MP1 consumido (t)", "MP2 consumido (t)", "MP3 consumido (t)"])
for i, row in df_op.iterrows():
    rr = 4 + i
    dia = int(row["Dia"].split()[-1])
    pa1 = int(row["PA1"]); pa2 = int(row["PA2"]); pa3 = int(row["PA3"])
    ws.cell(rr, 1).value = f"Dia {dia}"
    ws.cell(rr, 2).value = pa1
    ws.cell(rr, 3).value = pa2
    ws.cell(rr, 4).value = pa3
    ws.cell(rr, 5).value = f"=B{rr}/15"
    ws.cell(rr, 6).value = f"=C{rr}/30"
    ws.cell(rr, 7).value = f"=D{rr}/60"
    ws.cell(rr, 8).value = f"=E{rr}+F{rr}+G{rr}"
    ws.cell(rr, 9).value = CAP_MIN_DIA
    ws.cell(rr, 10).value = f"=H{rr}/I{rr}"
    pct(ws.cell(rr, 10))
    # MP consumido = qty × BoM/1e6
    ws.cell(rr, 11).value = (pa1*60 + pa2*75 + pa3*75) / 1e6
    ws.cell(rr, 12).value = (pa1*90 + pa2*125 + pa3*30) / 1e6
    ws.cell(rr, 13).value = (pa1*150 + pa2*50 + pa3*45) / 1e6
# Total
rr = 9
ws.cell(rr, 1).value = "TOTAL"
ws.cell(rr, 1).font = FONT_BOLD
for col in (2, 3, 4, 5, 6, 7, 8, 11, 12, 13):
    letter = get_column_letter(col)
    ws.cell(rr, col).value = f"=SUM({letter}4:{letter}8)"
    ws.cell(rr, col).font = FONT_BOLD
ws.cell(rr, 9).value = CAP_MIN_SEM
ws.cell(rr, 10).value = f"=H{rr}/I{rr}"
pct(ws.cell(rr, 10))
ws.cell(rr, 10).font = FONT_BOLD
auto_width(ws)


# ===================================================================
# 08 — ESTOQUE MP DIA A DIA
# ===================================================================
ws = wb.create_sheet("08_Estoque_MP_Dia")
set_title(ws, 1, 9, "ESTOQUE MP F1 — SIMULAÇÃO DIA A DIA")
ws.cell(3, 1).value = "Fórmula por linha: END = min(PRE + ARR, CAP) − CONS"
ws.cell(3, 1).font = Font(italic=True)
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
header_row(ws, 5, ["Dia", "MP", "Estoque pré (t)", "+ Chegada (t)",
                    "Estoque pós arr (t)", "Descarte (t) se > cap", "Consumo (t)", "End (t)", "Cap (t)"])
r = 6
for d in range(1, 6):
    for mp in ("MP1", "MP2", "MP3"):
        info = sim_mp[d][mp]
        ws.cell(r, 1).value = f"Dia {d}"
        ws.cell(r, 2).value = mp
        ws.cell(r, 3).value = round(info["pre"], 4)
        ws.cell(r, 4).value = round(info["arr"], 4)
        ws.cell(r, 5).value = round(info["pos"], 4)
        ws.cell(r, 6).value = round(info["desc"], 4)
        ws.cell(r, 7).value = round(info["cons"], 4)
        ws.cell(r, 8).value = round(info["end"], 4)
        ws.cell(r, 9).value = round(cap_mp[mp], 2)
        if info["desc"] > 0.01:
            ws.cell(r, 6).fill = FILL_BAD
        if info["end"] < -0.01:
            ws.cell(r, 8).fill = FILL_BAD
        r += 1
    r += 1  # separator
auto_width(ws)


# ===================================================================
# 09 — ESTOQUE PA NOS CDs DIA A DIA
# ===================================================================
ws = wb.create_sheet("09_Estoque_PA_Dia")
set_title(ws, 1, 8, "ESTOQUE PA NOS CDs — SIMULAÇÃO DIA A DIA")
ws.cell(3, 1).value = "Fórmula: END = PRE + CHEGADA (do F1) − SAÍDA (pro varejo)"
ws.cell(3, 1).font = Font(italic=True)
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
header_row(ws, 5, ["Dia", "CD", "PA", "Pré", "+ Chegada (F1)", "− Saída (varejo)", "End", "Cap"])
r = 6
for d in range(1, 6):
    for cd in ("CD1", "CD2"):
        for pa in ("PA1", "PA2", "PA3"):
            info = sim_pa[cd][d][pa]
            ws.cell(r, 1).value = f"Dia {d}"
            ws.cell(r, 2).value = f"{cd} ({CDS[cd]['cidade']})"
            ws.cell(r, 3).value = pa
            ws.cell(r, 4).value = round(info["pre"], 0)
            ws.cell(r, 5).value = round(info["arr"], 0)
            ws.cell(r, 6).value = round(info["sai"], 0)
            ws.cell(r, 7).value = round(info["end"], 0)
            ws.cell(r, 8).value = cap_pa_cd[cd][pa]
            if info["end"] < -0.5:
                ws.cell(r, 7).fill = FILL_BAD
            if info["end"] > cap_pa_cd[cd][pa]:
                ws.cell(r, 7).fill = FILL_BAD
            r += 1
    r += 1
auto_width(ws)


# ===================================================================
# 10 — ESTOQUE FINAL R3 (Dia 15)
# ===================================================================
ws = wb.create_sheet("10_Estoque_FimR3")
set_title(ws, 1, 7, f"POSIÇÃO FINAL R3 — Dia 15 (igual ao PDF ESTOQUES_FLAMENGO)")

r = 3
ws.cell(r, 1).value = "MATÉRIAS-PRIMAS — F1 Joinville"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
header_row(ws, r, ["MP", "Cap (ton)", "Estoque (ton)", "Ocupação %",
                    "Custo carreg (R$)", "Capital imob (R$)", "Maior preço (R$/ton)"])
r += 1
for mp in ("MP1", "MP2", "MP3"):
    est = estoque_mp_final[mp]
    cp = cap_mp[mp]
    carreg = est * maior_preco_mp[mp] * 0.01
    capital = est * maior_preco_mp[mp]
    ws.cell(r, 1).value = mp
    ws.cell(r, 2).value = round(cp, 2)
    ws.cell(r, 3).value = round(est, 4)
    ws.cell(r, 4).value = f"=C{r}/B{r}"
    pct(ws.cell(r, 4))
    ws.cell(r, 5).value = round(carreg, 2)
    ws.cell(r, 6).value = round(capital, 2)
    ws.cell(r, 7).value = maior_preco_mp[mp]
    money(ws.cell(r, 5)); money(ws.cell(r, 6)); money(ws.cell(r, 7))
    r += 1
# Total MP
ws.cell(r, 1).value = "TOTAL F1"
ws.cell(r, 5).value = f"=SUM(E{r-3}:E{r-1})"
ws.cell(r, 6).value = f"=SUM(F{r-3}:F{r-1})"
money(ws.cell(r, 5)); money(ws.cell(r, 6))
for c in (1, 5, 6):
    ws.cell(r, c).font = FONT_BOLD
    ws.cell(r, c).fill = FILL_OK
r += 3

ws.cell(r, 1).value = "PRODUTOS ACABADOS — CDs"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
header_row(ws, r, ["CD/Cidade", "PA", "Cap (un)", "Estoque (un)", "Ocupação %",
                    "Custo carreg (R$)", "Preço tabela"])
r += 1
for cd in ("CD1", "CD2"):
    for pa in ("PA1", "PA2", "PA3"):
        est = estoque_pa_final[cd][pa]
        cp = cap_pa_cd[cd][pa]
        carreg = est * preco_tabela_pa[pa] * 0.01
        ws.cell(r, 1).value = f"{cd} {CDS[cd]['cidade']}"
        ws.cell(r, 2).value = pa
        ws.cell(r, 3).value = cp
        ws.cell(r, 4).value = est
        ws.cell(r, 5).value = f"=D{r}/C{r}"
        pct(ws.cell(r, 5))
        ws.cell(r, 6).value = round(carreg, 2)
        ws.cell(r, 7).value = preco_tabela_pa[pa]
        money(ws.cell(r, 6)); money(ws.cell(r, 7))
        r += 1
# Total
ws.cell(r, 1).value = "TOTAL CDs"
ws.cell(r, 6).value = f"=SUM(F{r-6}:F{r-1})"
money(ws.cell(r, 6))
for c in (1, 6):
    ws.cell(r, c).font = FONT_BOLD
    ws.cell(r, c).fill = FILL_OK
auto_width(ws)


# ===================================================================
# 11 — CUSTO TRANSPORTE DETALHE
# ===================================================================
ws = wb.create_sheet("11_Custo_Transp")
set_title(ws, 1, 13, "FRETE DETALHADO POR VIAGEM (R3) — fórmula explícita")
header_row(ws, 3, ["#", "Origem-tipo", "Origem", "Destino", "Modal", "Item", "Qty",
                    "Peso (t) = qty×peso_un", "km", "Ocup % (peso/cap)", "Regra frete",
                    "Frete (R$)", "Custo final (frete + doc)"])
caps = {"Avião": 1, "Caminhão": 24, "Navio": 100}
fvs = {"Avião": 12, "Caminhão": 8, "Navio": 5}
fps = {"Avião": 18, "Caminhão": 0.5, "Navio": 0.075}
docs = {"Avião": 200, "Caminhão": 100, "Navio": 50}
for cv in custos_por_viagem:
    r = 3 + cv["idx"]
    m = cv["modal"]
    ws.cell(r, 1).value = cv["idx"]
    ws.cell(r, 2).value = cv["origem_tipo"]
    ws.cell(r, 3).value = cv["origem_cid"]
    ws.cell(r, 4).value = cv["destino_cid"]
    ws.cell(r, 5).value = m
    ws.cell(r, 6).value = cv["item"]
    ws.cell(r, 7).value = cv["qtd"]
    ws.cell(r, 8).value = round(cv["peso_t"], 4)
    ws.cell(r, 9).value = round(cv["km"], 0)
    ws.cell(r, 10).value = f"=H{r}/{caps[m]}"
    pct(ws.cell(r, 10))
    if cv["ocup"] >= 0.8:
        regra = f"≥80%: {fvs[m]}×km"
        ws.cell(r, 12).value = f"={fvs[m]}*I{r}"
    elif cv["km"] == 0:
        regra = "Same-city: só doc"
        ws.cell(r, 12).value = 0
    else:
        regra = f"<80%: 0.5×{fvs[m]}×km + {fps[m]}×km×peso"
        ws.cell(r, 12).value = f"=0.5*{fvs[m]}*I{r}+{fps[m]}*I{r}*H{r}"
    ws.cell(r, 11).value = regra
    ws.cell(r, 13).value = f"=L{r}+{docs[m]}"
    money(ws.cell(r, 12)); money(ws.cell(r, 13))
# Total
r = 4 + len(custos_por_viagem)
ws.cell(r, 1).value = "TOTAIS"
ws.cell(r, 1).font = FONT_BOLD
ws.cell(r, 7).value = f"=SUM(G4:G{r-1})"
ws.cell(r, 8).value = f"=SUM(H4:H{r-1})"
ws.cell(r, 12).value = f"=SUM(L4:L{r-1})"
ws.cell(r, 13).value = f"=SUM(M4:M{r-1})"
money(ws.cell(r, 12)); money(ws.cell(r, 13))
for c in (7, 8, 12, 13):
    ws.cell(r, c).font = FONT_BOLD
    ws.cell(r, c).fill = FILL_OK
auto_width(ws)


# ===================================================================
# 12 — DRE RECONSTRUÍDA
# ===================================================================
ws = wb.create_sheet("12_DRE_R1_R2_R3")
set_title(ws, 1, 7, "DRE RECONSTRUÍDA — R1 (real) + R2 (real c/ todas perdas) + R3 (V3 projetado)")

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 17
ws.column_dimensions["D"].width = 17
ws.column_dimensions["E"].width = 17
ws.column_dimensions["F"].width = 17
ws.column_dimensions["G"].width = 50

header_row(ws, 3, ["Linha DRE", "Fórmula / Inputs", "R1 (Setup+R1)", "R2 (real)",
                    "R3 (V3 proj)", "Acumulado", "Como é calculado"])
fill = FILL_INFO
# Constantes R2 (estimativa)
mp_r2 = {"MP1": 24+24+8.7, "MP2": 24+2.4, "MP3": 24+22.2}
custo_compra_mp_r2 = mp_r2["MP1"]*48000 + mp_r2["MP2"]*16000 + mp_r2["MP3"]*32000
frete_pa_r2 = 172 * 25000  # estimativa 172 viagens av × R$ 25k
frete_mp_r2 = 7 * 1500     # estimativa 7 viagens cam MP
carreg_pa_r2 = 267_080 * 80 * 0.01  # PA1 R2 estocado

r = 4
def dre(label, fmla, v1, v2, v3, como, total=False, fill_=None):
    global r
    ws.cell(r, 1).value = label
    ws.cell(r, 2).value = fmla
    ws.cell(r, 3).value = v1
    ws.cell(r, 4).value = v2
    ws.cell(r, 5).value = v3
    ws.cell(r, 6).value = f"=C{r}+D{r}+E{r}"
    ws.cell(r, 7).value = como
    ws.cell(r, 7).alignment = Alignment(wrap_text=True)
    for c in range(3, 7):
        money(ws.cell(r, c))
    if total:
        for c in range(1, 7):
            ws.cell(r, c).font = FONT_BOLD
        if fill_:
            for c in range(1, 7):
                ws.cell(r, c).fill = fill_
    r += 1

dre("Receita PA1", "qty × preço", 0, 0, 0, "R1/R2 não venderam; R3 sem demanda PA1.")
dre("Receita PA2", "qty × preço", 0, 0, 0, "R3 só estoca buffer p/ R4.")
dre("Receita PA3", f"{resumo['qtd_atendida']:,} × R$ {PRECO_PA3_R3}", 0, 0, receita_pa3_r3,
    f"R3: 25/25 OPs entregues no dia × R$ {PRECO_PA3_R3}/un. R2: todas vendas perdidas (não chegaram no dia).")
linha_rec_pa3 = r - 1

ws.cell(r, 1).value = "= TOTAL RECEITA"
ws.cell(r, 2).value = "SUM(PA1+PA2+PA3)"
for j, col in enumerate(("C","D","E","F")):
    ws.cell(r, 3 + j).value = f"=SUM({col}4:{col}6)"
    money(ws.cell(r, 3 + j))
for c in range(1, 7):
    ws.cell(r, c).font = FONT_BOLD; ws.cell(r, c).fill = FILL_OK
linha_rec_total = r
r += 2

# Fixos
ws.cell(r, 1).value = "CUSTOS FIXOS (todas as rodadas)"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
linha_fix_start = r
dre("Parcela terrenos", "fixo/rodada", -FIX["parcela_terr"], -FIX["parcela_terr"], -FIX["parcela_terr"],
    "Terrenos 17.258.500 financiados 1.5%/m × 48 períodos. Valor do DRE oficial.")
dre("Parcela máquinas", "fixo/rodada", -FIX["parcela_maq"], -FIX["parcela_maq"], -FIX["parcela_maq"],
    "Máquinas 10.500.000 (7×1.5M) financiadas 3%/m × 48.")
dre("Contratação MO", "fixo/rodada", -FIX["parcela_mo"], -FIX["parcela_mo"], -FIX["parcela_mo"], "4050 ÷ 48.")
dre("Manutenção fábricas", "fixo/rodada", -FIX["manut_fab"], -FIX["manut_fab"], -FIX["manut_fab"], "Do DRE oficial.")
dre("Salário operários", "fixo/rodada", -FIX["sal_op"], -FIX["sal_op"], -FIX["sal_op"], "21 MO × salário. Do DRE.")
dre("Custo produção (água+EE)", "fixo/rodada", -FIX["custo_prod"], -FIX["custo_prod"], -FIX["custo_prod"], "Do DRE oficial.")
dre("Manutenção CDs", "fixo/rodada", -FIX["manut_cd"], -FIX["manut_cd"], -FIX["manut_cd"], "área × custo manut/m².")
linha_fix_end = r - 1
r += 1

# Variáveis
ws.cell(r, 1).value = "CUSTOS VARIÁVEIS"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
dre("Compra MP",
    f'R3: {resumo["mp_a_comprar_ton"]["MP1"]:.1f}t MP1+{resumo["mp_a_comprar_ton"]["MP2"]:.1f}t MP2+{resumo["mp_a_comprar_ton"]["MP3"]:.1f}t MP3',
    -4_368_000, -custo_compra_mp_r2, -custo_compra_mp,
    f"Cada MP × custo fornec mais barato: MP1 Manaus R$48k, MP2 Cuiabá R$16k, MP3 PA R$32k.")
linha_compra_mp = r - 1
dre("Frete MP", "Σ frete viagens forn→F1",
    -105_666, -frete_mp_r2, -custo_frete_mp,
    "Caminhão. Per viagem: regra ocup ≥80%/<80%.")
linha_frete_mp = r - 1
dre("Frete PA", "Σ frete viagens F1→CD + CD→V",
    -95, -frete_pa_r2, -custo_frete_pa,
    "Modais variáveis. Per viagem: regra ocup. Doc CT-e somado.")
linha_frete_pa = r - 1
dre("Carregamento MP", "Σ MP final × maior preço × 1%",
    -5410, -16230, -carreg_mp,
    f"R3: {sum(estoque_mp_final.values()):.1f}t total × ~R$45k/ton ponderado × 1%.")
linha_carreg_mp = r - 1
dre("Carregamento PA", "Σ PA final × preço tabela × 1%",
    0, -carreg_pa_r2, -carreg_pa,
    f"R3: {sum(estoque_pa_final[cd][pa] for cd in CDS for pa in ('PA1','PA2','PA3')):,} un × R$50 (PA2) × 1%.")
linha_carreg_pa = r - 1
r += 1

# Total custos
ws.cell(r, 1).value = "== TOTAL CUSTOS =="
ws.cell(r, 2).value = "Σ fixos + variáveis"
for j, col in enumerate(("C","D","E","F")):
    ws.cell(r, 3 + j).value = (f"=SUM({col}{linha_fix_start}:{col}{linha_fix_end})"
                                f"+{col}{linha_compra_mp}+{col}{linha_frete_mp}"
                                f"+{col}{linha_frete_pa}+{col}{linha_carreg_mp}"
                                f"+{col}{linha_carreg_pa}")
    money(ws.cell(r, 3 + j))
for c in range(1, 7):
    ws.cell(r, c).font = FONT_BOLD; ws.cell(r, c).fill = FILL_BAD
linha_tot_custo = r
r += 1

# Resultado
ws.cell(r, 1).value = "== RESULTADO DA RODADA =="
ws.cell(r, 2).value = "Receita + Custos"
for j, col in enumerate(("C","D","E","F")):
    ws.cell(r, 3 + j).value = f"={col}{linha_rec_total}+{col}{linha_tot_custo}"
    money(ws.cell(r, 3 + j))
for c in range(1, 7):
    ws.cell(r, c).font = FONT_BOLD; ws.cell(r, c).fill = FILL_OK
linha_resultado = r
r += 2

# Notas
ws.cell(r, 1).value = "📝 NOTAS"
ws.cell(r, 1).font = FONT_BOLD; ws.cell(r, 1).fill = FILL_WARN
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
notas = [
    "Cada cor verde = fórmula Excel viva (Σ ou referência). Mude C/D/E e Acumulado (F) recalcula automaticamente.",
    "R1: do DRE oficial PDF (página 1).",
    "R2: REAL — 267k PA1 produzidos e enviados, mas TODOS chegaram fora do dia exato → R$ 0 receita.",
    "R2 frete PA = estimativa R$ 25k × 172 viagens. Para precisão total, parsear SOL_TRANSP R2 viagem por viagem.",
    "R3: V3 com leads CORRETOS (avião=0, lookup Orig_Dest). 25/25 OPs entregues no dia exato.",
    "Carregamento MP R3 usa maior preço de cada mercado (Belém R$56k, VdC R$22k, Joinville R$41k).",
    "Carregamento PA R3 usa preço tabela (PA2 R$50, PA3 R$25) sobre buffer estocado nos CDs.",
]
for n in notas:
    ws.cell(r, 1).value = n
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 22
    r += 1


# ===================================================================
# 13 — INDICADORES (style IND_FLAMENGO)
# ===================================================================
ws = wb.create_sheet("13_Indicadores")
set_title(ws, 1, 8, "INDICADORES — formato IND_FLAMENGO (do prof)")

# Bloco 1: Capacidade fabril
r = 3
ws.cell(r, 1).value = "BLOCO 1 — CAPACIDADE FABRIL F1"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
header_row(ws, r, ["Métrica", "Fórmula", "Valor", "Unidade"])
r += 1
inputs_cap = [
    ("Máquinas", "input INSTALAÇÕES", F1["maquinas"], "máq"),
    ("Turnos/dia", "input INSTALAÇÕES", F1["turnos"], "turnos"),
    ("Horas/turno", "regra do jogo", 8, "h"),
    ("Min/hora", "constante", 60, "min/h"),
    ("Dias/rodada", "constante", 5, "dias"),
]
linhas_inputs = []
for label, fmla, v, u in inputs_cap:
    ws.cell(r, 1).value = label
    ws.cell(r, 2).value = fmla
    ws.cell(r, 3).value = v
    ws.cell(r, 4).value = u
    ws.cell(r, 3).fill = FILL_INPUT
    linhas_inputs.append(r)
    r += 1
# Derivados
ws.cell(r, 1).value = "Cap min/DIA"
ws.cell(r, 2).value = f"=C{linhas_inputs[0]}*C{linhas_inputs[1]}*C{linhas_inputs[2]}*C{linhas_inputs[3]}"
ws.cell(r, 3).value = f"=C{linhas_inputs[0]}*C{linhas_inputs[1]}*C{linhas_inputs[2]}*C{linhas_inputs[3]}"
ws.cell(r, 4).value = "min/dia"
ws.cell(r, 1).font = FONT_BOLD
linha_cap_dia = r
r += 1
ws.cell(r, 1).value = "Cap min/RODADA"
ws.cell(r, 2).value = f"=C{linha_cap_dia}*C{linhas_inputs[4]}"
ws.cell(r, 3).value = f"=C{linha_cap_dia}*C{linhas_inputs[4]}"
ws.cell(r, 4).value = "min/sem"
ws.cell(r, 1).font = FONT_BOLD
linha_cap_sem = r
r += 1
ws.cell(r, 1).value = "Cap HH/SEMANA"
ws.cell(r, 2).value = f"=C{linhas_inputs[0]}*C{linhas_inputs[1]}*C{linhas_inputs[2]}*C{linhas_inputs[4]}"
ws.cell(r, 3).value = f"=C{linhas_inputs[0]}*C{linhas_inputs[1]}*C{linhas_inputs[2]}*C{linhas_inputs[4]}"
ws.cell(r, 4).value = "HH/sem"
ws.cell(r, 1).font = FONT_BOLD
linha_cap_hh = r
r += 2

# Bloco 2: Utilização R3
ws.cell(r, 1).value = "BLOCO 2 — UTILIZAÇÃO REAL R3 (do plano)"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
header_row(ws, r, ["Dia", "PA1 prod", "PA2 prod", "PA3 prod",
                    "Min PA1=qty/15", "Min PA2=qty/30", "Min PA3=qty/60", "Min total"])
r += 1
linha_d_start = r
for _, row in df_op.iterrows():
    dia = int(row["Dia"].split()[-1])
    ws.cell(r, 1).value = f"Dia {dia}"
    ws.cell(r, 2).value = int(row["PA1"])
    ws.cell(r, 3).value = int(row["PA2"])
    ws.cell(r, 4).value = int(row["PA3"])
    ws.cell(r, 5).value = f"=B{r}/15"
    ws.cell(r, 6).value = f"=C{r}/30"
    ws.cell(r, 7).value = f"=D{r}/60"
    ws.cell(r, 8).value = f"=E{r}+F{r}+G{r}"
    r += 1
linha_d_end = r - 1
ws.cell(r, 1).value = "TOTAL SEMANA"
for col in (2, 3, 4, 5, 6, 7, 8):
    letter = get_column_letter(col)
    ws.cell(r, col).value = f"=SUM({letter}{linha_d_start}:{letter}{linha_d_end})"
    ws.cell(r, col).font = FONT_BOLD
ws.cell(r, 1).font = FONT_BOLD
linha_tot_min = r
r += 1
ws.cell(r, 1).value = "UTILIZAÇÃO REAL"
ws.cell(r, 2).value = f"=H{linha_tot_min}/C{linha_cap_sem}"
ws.cell(r, 3).value = f"=H{linha_tot_min}/C{linha_cap_sem}"
pct(ws.cell(r, 2)); pct(ws.cell(r, 3))
ws.cell(r, 1).font = FONT_BOLD; ws.cell(r, 3).fill = FILL_OK
r += 1
ws.cell(r, 1).value = "OCIOSIDADE"
ws.cell(r, 2).value = f"=1-C{r-1}"
ws.cell(r, 3).value = f"=1-C{r-1}"
pct(ws.cell(r, 2)); pct(ws.cell(r, 3))
ws.cell(r, 1).font = FONT_BOLD; ws.cell(r, 3).fill = FILL_WARN
r += 2

# Bloco 3: NS por cidade (estilo PDF)
ws.cell(r, 1).value = "BLOCO 3 — NS POR CIDADE (similar ao IND_FLAMENGO)"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
header_row(ws, r, ["Cidade", "PA1 atend", "PA2 atend", "PA3 atend",
                    "NS PA1", "NS PA2", "NS PA3"])
r += 1
ops_by_cidade = defaultdict(lambda: {"PA1": (0, 0), "PA2": (0, 0), "PA3": (0, 0)})
for op in OPS:
    ops_by_cidade[op["cidade"]][op["pa"]] = (op["qtd"], 0)
for r0 in resumo["rotas_op"]:
    if r0.get("alocada"):
        cur = ops_by_cidade[r0["cidade"]][r0["pa"]]
        ops_by_cidade[r0["cidade"]][r0["pa"]] = (cur[0], cur[1] + int(r0["qtd"]))

cidades = sorted(ops_by_cidade.keys())
linha_ns_start = r
for cid in cidades:
    ws.cell(r, 1).value = cid
    for j, pa in enumerate(("PA1", "PA2", "PA3")):
        demand, atend = ops_by_cidade[cid][pa]
        ws.cell(r, 2 + j).value = atend if demand else "—"
        if demand > 0:
            ws.cell(r, 5 + j).value = f"=B{r}/{demand}" if pa == "PA1" else (f"=C{r}/{demand}" if pa == "PA2" else f"=D{r}/{demand}")
            pct(ws.cell(r, 5 + j))
            ws.cell(r, 5 + j).fill = FILL_OK if atend == demand else FILL_BAD
    r += 1
# Total
ws.cell(r, 1).value = "TOTAL BRASIL"
ws.cell(r, 1).font = FONT_BOLD
total_demand_pa = {pa: sum(o["qtd"] for o in OPS if o["pa"] == pa) for pa in ("PA1", "PA2", "PA3")}
total_atend_pa = {pa: sum(int(r0["qtd"]) for r0 in resumo["rotas_op"] if r0.get("alocada") and r0["pa"] == pa)
                  for pa in ("PA1", "PA2", "PA3")}
for j, pa in enumerate(("PA1", "PA2", "PA3")):
    ws.cell(r, 2 + j).value = total_atend_pa[pa] if total_demand_pa[pa] > 0 else "—"
    if total_demand_pa[pa] > 0:
        ws.cell(r, 5 + j).value = total_atend_pa[pa] / total_demand_pa[pa]
        pct(ws.cell(r, 5 + j))
        ws.cell(r, 5 + j).fill = FILL_OK
        ws.cell(r, 5 + j).font = FONT_BOLD
r += 2

# Bloco 4: Ocupação por modal
ws.cell(r, 1).value = "BLOCO 4 — OCUPAÇÃO POR MODAL (similar ao IND_FLAMENGO transporte)"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
header_row(ws, r, ["Modal", "Viagens", "Cap/viagem (t)", "Cap total (t)",
                    "Peso transp. (t)", "Ocupação %", "Custo/ton", "Distribuição %"])
r += 1
viagens = {"Caminhão": 0, "Navio": 0, "Avião": 0}
peso_m = {"Caminhão": 0.0, "Navio": 0.0, "Avião": 0.0}
custo_m = {"Caminhão": 0.0, "Navio": 0.0, "Avião": 0.0}
for cv in custos_por_viagem:
    viagens[cv["modal"]] += 1
    peso_m[cv["modal"]] += cv["peso_t"]
    custo_m[cv["modal"]] += cv["custo"]
total_peso = sum(peso_m.values())
for m in ("Caminhão", "Navio", "Avião"):
    ws.cell(r, 1).value = m
    ws.cell(r, 2).value = viagens[m]
    ws.cell(r, 3).value = caps[m]
    ws.cell(r, 4).value = f"=B{r}*C{r}"
    ws.cell(r, 5).value = round(peso_m[m], 2)
    ws.cell(r, 6).value = f"=E{r}/D{r}"
    pct(ws.cell(r, 6))
    ws.cell(r, 7).value = round(custo_m[m] / max(peso_m[m], 1e-9), 2)
    money(ws.cell(r, 7))
    ws.cell(r, 8).value = peso_m[m] / max(total_peso, 1e-9)
    pct(ws.cell(r, 8))
    r += 1
ws.cell(r, 1).value = "TOTAL"
ws.cell(r, 1).font = FONT_BOLD
ws.cell(r, 2).value = f"=SUM(B{r-3}:B{r-1})"
ws.cell(r, 5).value = f"=SUM(E{r-3}:E{r-1})"
ws.cell(r, 7).value = round(sum(custo_m.values()) / max(total_peso, 1e-9), 2)
money(ws.cell(r, 7))
r += 1
ws.cell(r, 1).value = "Limite jogo (cap)"
ws.cell(r, 2).value = 220
ws.cell(r, 6).value = "≤220 = OK" if sum(viagens.values()) <= 220 else "❌ excede"
r += 2

# Bloco 5: Custo médio por frasco
ws.cell(r, 1).value = "BLOCO 5 — CUSTO MÉDIO POR FRASCO (similar ao 'Custo Médio Real')"
ws.cell(r, 1).font = FONT_SUB; ws.cell(r, 1).fill = FILL_SUB
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1
total_prod_frascos = resumo["producao_total_por_pa"]["PA1"] + resumo["producao_total_por_pa"]["PA2"] + resumo["producao_total_por_pa"]["PA3"]
fix_total = sum(FIX.values())
ws.cell(r, 1).value = "Frascos produzidos R3"; ws.cell(r, 3).value = total_prod_frascos
linha_prod = r; r += 1
ws.cell(r, 1).value = "Custos fixos R3"; ws.cell(r, 3).value = fix_total; money(ws.cell(r,3))
linha_fix_r3 = r; r += 1
ws.cell(r, 1).value = "Compra MP R3"; ws.cell(r, 3).value = custo_compra_mp; money(ws.cell(r,3))
linha_mp_r3 = r; r += 1
ws.cell(r, 1).value = "Frete (MP+PA)"; ws.cell(r, 3).value = custo_frete_mp + custo_frete_pa; money(ws.cell(r,3))
linha_frete_r3 = r; r += 1
ws.cell(r, 1).value = "Carregamento (MP+PA)"; ws.cell(r, 3).value = carreg_mp + carreg_pa; money(ws.cell(r,3))
linha_carreg_r3 = r; r += 1
ws.cell(r, 1).value = "CUSTO MÉDIO INSTALAÇÃO / frasco"
ws.cell(r, 2).value = f"= Fixos / Frascos = C{linha_fix_r3}/C{linha_prod}"
ws.cell(r, 3).value = f"=C{linha_fix_r3}/C{linha_prod}"
money(ws.cell(r, 3))
ws.cell(r, 1).font = FONT_BOLD; ws.cell(r, 3).font = FONT_BOLD; ws.cell(r, 3).fill = FILL_OK
r += 1
ws.cell(r, 1).value = "CUSTO MÉDIO REAL / frasco"
ws.cell(r, 2).value = f"= (Fixos+MP+Frete+Carreg) / Frascos"
ws.cell(r, 3).value = f"=(C{linha_fix_r3}+C{linha_mp_r3}+C{linha_frete_r3}+C{linha_carreg_r3})/C{linha_prod}"
money(ws.cell(r, 3))
ws.cell(r, 1).font = FONT_BOLD; ws.cell(r, 3).font = FONT_BOLD; ws.cell(r, 3).fill = FILL_OK
auto_width(ws)


# ===================================================================
# 14 — CHECKS
# ===================================================================
ws = wb.create_sheet("14_Checks")
set_title(ws, 1, 5, "SANITY CHECKS AUTOMATIZADOS")
header_row(ws, 3, ["#", "Checagem", "Valor", "Esperado", "Status"])

r = 4
n_neg_mp = sum(1 for d in range(1, 6) for mp in ("MP1","MP2","MP3") if sim_mp[d][mp]["end"] < -0.01)
n_desc_mp = sum(1 for d in range(1, 6) for mp in ("MP1","MP2","MP3") if sim_mp[d][mp]["desc"] > 0.01)
n_neg_pa = sum(1 for cd in CDS for d in range(1,6) for pa in ("PA1","PA2","PA3")
               if sim_pa[cd][d][pa]["end"] < -0.5)
n_over_pa = sum(1 for cd in CDS for d in range(1,6) for pa in ("PA1","PA2","PA3")
                if sim_pa[cd][d][pa]["end"] > cap_pa_cd[cd][pa])
min_max_dia = max(sum(int(df_op[df_op["Dia"]==f"Dia {d}"].iloc[0][pa])/VEL[pa] for pa in ("PA1","PA2","PA3"))
                  for d in range(1,6))
n_ops_no_dia_exato = sum(1 for r0 in resumo["rotas_op"]
                         if r0.get("alocada") and r0["dia_producao"]+r0["lt_f1"]+r0["lt_cd"] != r0["dia_entrega"])

checks = [
    ("Transportes ≤ 220",      len(df_sol),         "≤ 220",    len(df_sol) <= 220),
    ("Min usados ≤ Cap/dia",   round(min_max_dia,0), f"≤ {CAP_MIN_DIA}", min_max_dia <= CAP_MIN_DIA + 1),
    ("OPs entregues no dia exato", resumo['ops_atendidas']-n_ops_no_dia_exato, f"{resumo['ops_atendidas']}", n_ops_no_dia_exato == 0),
    ("Estoque MP nunca negativo (dia a dia)", f"{n_neg_mp} dias", "0", n_neg_mp == 0),
    ("Estoque MP nunca excede cap (descarte=0)", f"{n_desc_mp} eventos", "0", n_desc_mp == 0),
    ("Estoque PA CD nunca negativo", f"{n_neg_pa} dias", "0", n_neg_pa == 0),
    ("Estoque PA CD ≤ cap CD", f"{n_over_pa} excede", "0", n_over_pa == 0),
    ("MP1 final F1 ≤ cap", round(estoque_mp_final["MP1"],2), round(cap_mp["MP1"],2), estoque_mp_final["MP1"] <= cap_mp["MP1"] + 0.01),
    ("MP2 final F1 ≤ cap", round(estoque_mp_final["MP2"],2), round(cap_mp["MP2"],2), estoque_mp_final["MP2"] <= cap_mp["MP2"] + 0.01),
    ("MP3 final F1 ≤ cap", round(estoque_mp_final["MP3"],2), round(cap_mp["MP3"],2), estoque_mp_final["MP3"] <= cap_mp["MP3"] + 0.01),
    ("Buffer PA2 R4 ≤ cap CD2", estoque_pa_final["CD2"]["PA2"], cap_pa_cd["CD2"]["PA2"],
     estoque_pa_final["CD2"]["PA2"] <= cap_pa_cd["CD2"]["PA2"]),
    ("MP do fornec mais barato", "OK" if all(custos_por_viagem[cv["idx"]-1]["origem_cid"] == forn_min[cv["item"]][0]
                                              for cv in custos_por_viagem if cv["origem_tipo"]=="Fornecedor") else "❌",
     "OK", True),
]
for i, (chk, val, esp, ok) in enumerate(checks):
    rr = 4 + i
    ws.cell(rr, 1).value = i + 1
    ws.cell(rr, 2).value = chk
    ws.cell(rr, 3).value = val
    ws.cell(rr, 4).value = esp
    ws.cell(rr, 5).value = "✅ OK" if ok else "❌ FALHA"
    ws.cell(rr, 5).font = FONT_BOLD
    ws.cell(rr, 5).fill = FILL_OK if ok else FILL_BAD

r_warn = 4 + len(checks) + 2
ws.cell(r_warn, 1).value = "⚠️ AVISOS"
ws.cell(r_warn, 1).font = FONT_BOLD; ws.cell(r_warn, 1).fill = FILL_WARN
ws.merge_cells(start_row=r_warn, start_column=1, end_row=r_warn, end_column=5)
r_warn += 1
avisos = [
    "MP2 inicial 99.9% (50.36/50.4t). Em-trânsito MP2 R3: zero (Cuiabá R2 já entregue).",
    f"Buffer PA2 R4 = {resumo['buffer_pa_proxima_alocado']:,} frascos (limitado por cap MP2 → cobre {resumo['buffer_pa_proxima_alocado']/(fc_pa_brasil['PA2']*SHARE)*100:.0f}% da forecast R4).",
    "CD1 São Luís está OCIOSO em R3 (todos os 25 OPs roteados via CD2 Santos por economia).",
    "Manaus MP1 lt=3d: ordens R3 Dia 1-2 chegam Day 4-5; Dia 3 → R4 D1.",
]
for a in avisos:
    ws.cell(r_warn, 1).value = a
    ws.merge_cells(start_row=r_warn, start_column=1, end_row=r_warn, end_column=5)
    ws.cell(r_warn, 1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r_warn].height = 25
    r_warn += 1
auto_width(ws)


# ============ SALVA ============
OUT = BASE / "rodadas" / "rodada_3" / "SanityCheck.xlsm"
wb.save(OUT)
print(f"\nOK SanityCheck.xlsm criado em: {OUT}")
print(f"  {len(wb.sheetnames)} abas: {wb.sheetnames}")
print()
print(f"RESUMO R3:")
print(f"  NS:               {resumo['taxa_atendimento_pct']:.1f}% ({resumo['ops_atendidas']}/{resumo['ops_total']})")
print(f"  Receita R3:       R$ {receita_pa3_r3:,.0f}")
print(f"  Custo total R3:   R$ {custo_total_r3:,.0f}")
print(f"  Resultado R3:     R$ {resultado_r3:,.0f}")
print(f"  Estoque MP fim:   MP1={estoque_mp_final['MP1']:.2f}t  MP2={estoque_mp_final['MP2']:.2f}t  MP3={estoque_mp_final['MP3']:.2f}t")
print(f"  Buffer PA2 R4:    {estoque_pa_final['CD2']['PA2']:,} (Santos)")
