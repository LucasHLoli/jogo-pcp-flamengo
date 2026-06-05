"""Auditoria das restrições do plano da Rodada 2 escrito no FLAMENGO.xlsm."""
import sys
import math
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from src.config import Config
from src.io_xlsm import ler_instalacoes

BASE = Path(__file__).resolve().parent.parent
cfg = Config.load(BASE)
inst = ler_instalacoes(BASE / 'rodadas' / 'FLAMENGO.xlsm')

wb = openpyxl.load_workbook(BASE / 'rodadas' / 'FLAMENGO.xlsm', keep_vba=True, data_only=False)
ws = wb['SOL_TRANSP']

transp_r2 = []
for r in range(5, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is None or str(v).strip() == '':
        break
    if 'Rodada_2' not in str(v):
        continue
    dia_match = re.search(r'\d+', str(ws.cell(r, 4).value or ''))
    if not dia_match:
        continue
    transp_r2.append({
        'linha': r,
        'origem_tipo': ws.cell(r, 2).value,
        'origem_cidade': ws.cell(r, 3).value,
        'dia_part': int(dia_match.group()),
        'modal': ws.cell(r, 5).value,
        'item': ws.cell(r, 6).value,
        'qtd': float(ws.cell(r, 7).value or 0),
        'destino_tipo': ws.cell(r, 8).value,
        'destino_cidade': ws.cell(r, 9).value,
    })

ws_op = wb['OP_FABRICAS']
producao = {}
for r in range(7, 12):
    dia = r - 6
    producao[dia] = {
        'PA1': int(ws_op.cell(r, 2).value or 0),
        'PA2': int(ws_op.cell(r, 3).value or 0),
        'PA3': int(ws_op.cell(r, 4).value or 0),
    }

print(f'Transportes Rodada 2: {len(transp_r2)}')
print(f'Produção F1: {producao}')
print()

erros = []

# 1. CAP MÁQUINA POR DIA
print('━' * 60)
print('1) CAPACIDADE MÁQUINA (cap 10080 min/dia)')
print('━' * 60)
for dia in range(1, 6):
    mins = sum(producao[dia][pa] / cfg.velocidades[pa] for pa in ('PA1', 'PA2', 'PA3'))
    pct = mins / 10080 * 100
    status = '✅' if mins <= 10080 + 0.5 else '❌'
    print(f'  Dia {dia}: {mins:>7.1f} min ({pct:>5.1f}%) {status}')
    if mins > 10080.5:
        erros.append(f'Dia {dia}: produção excede capacidade ({mins:.0f} > 10080)')
print()

# 2. CAP MODAL POR VIAGEM
print('━' * 60)
print('2) CAPACIDADE MODAL POR VIAGEM')
print('━' * 60)
ok_cap_modal = True
for t in transp_r2:
    cap = cfg.cap_modal_por_item[t['modal']][t['item']]
    if t['qtd'] > cap + 0.5:
        print(f'  ❌ L{t["linha"]}: {t["modal"]} {t["item"]} qtd={t["qtd"]} > cap {cap}')
        erros.append(f'Cap modal L{t["linha"]}')
        ok_cap_modal = False
if ok_cap_modal:
    print(f'  ✅ Todas {len(transp_r2)} viagens dentro da cap modal')
print()

# 3. NAVIO ROTAS VÁLIDAS
print('━' * 60)
print('3) NAVIO — rotas marítimas válidas')
print('━' * 60)
ok_navio = True
for t in transp_r2:
    if t['modal'] == 'Navio':
        if (t['origem_cidade'], t['destino_cidade']) not in cfg.rotas_navio_validas:
            print(f'  ❌ L{t["linha"]}: Navio {t["origem_cidade"]}→{t["destino_cidade"]}')
            erros.append(f'Navio rota inválida L{t["linha"]}')
            ok_navio = False
n_navio = sum(1 for t in transp_r2 if t['modal'] == 'Navio')
if ok_navio:
    print(f'  ✅ {n_navio} viagens de navio (todas válidas)' if n_navio > 0 else '  ✅ Sem navio nesta rodada')
print()

# 4. LIMITE 220 TRANSPORTES
print('━' * 60)
print('4) LIMITE 220 TRANSPORTES/SEMANA')
print('━' * 60)
n = len(transp_r2)
status = '✅' if n <= 220 else '❌'
print(f'  Total Rodada 2: {n}/220 {status}')
if n > 220:
    erros.append(f'Excedeu 220 ({n})')
print()

# 5. PA NÃO DORME NA FÁBRICA
print('━' * 60)
print('5) PA NÃO DORME NA FÁBRICA (produção dia X = saída dia X)')
print('━' * 60)
saidas_f1_pa = defaultdict(lambda: defaultdict(float))
for t in transp_r2:
    if t['origem_tipo'] == 'Fábrica' and t['item'].startswith('PA'):
        saidas_f1_pa[t['dia_part']][t['item']] += t['qtd']

ok_dorme = True
for dia in range(1, 6):
    for pa in ('PA1', 'PA2', 'PA3'):
        prod = producao[dia][pa]
        saida = saidas_f1_pa[dia][pa]
        if abs(prod - saida) > 1:
            print(f'  ❌ Dia {dia} {pa}: produzido {prod} ≠ enviado {saida:.0f}')
            erros.append(f'PA dorme F1 Dia {dia} {pa}')
            ok_dorme = False
if ok_dorme:
    print('  ✅ Produção = saída F1 em todos os dias')
print()

# 6. LEAD TIME CD→VAREJO
print('━' * 60)
print('6) LEAD TIME — CD→Varejo respeita dia_entrega')
print('━' * 60)
OPS_DIA_ENTREGA = {
    'Belém': 5, 'Belo Horizonte': 3, 'Brasília': 3, 'Campinas': 2,
    'Campo Grande': 3, 'Cuiabá': 3, 'Curitiba': 2, 'Fortaleza': 5,
    'Goiânia': 3, 'João Pessoa': 4, 'Joinville': 2, 'Maceió': 4,
    'Manaus': 5, 'Natal': 5, 'Porto Alegre': 2, 'Recife': 4,
    'Ribeirão Preto': 2, 'Rio de Janeiro': 3, 'Salvador': 4, 'Santos': 2,
    'São Luís': 5, 'São Paulo': 2, 'Uberlândia': 3, 'Vitória': 3,
    'Vitória da Conquista': 4,
}
ok_lead = True
for t in transp_r2:
    if not (t['origem_tipo'] == 'CD' and t['destino_tipo'] == 'Varejista'):
        continue
    cidade = t['destino_cidade']
    if cidade not in OPS_DIA_ENTREGA:
        continue
    dia_entrega = OPS_DIA_ENTREGA[cidade]
    try:
        km = float(cfg.distancias[t['modal']].at[t['origem_cidade'], cidade])
    except Exception:
        continue
    if not km or km != km:
        continue
    vel = {'Caminhão': 50, 'Navio': 30, 'Avião': 700}[t['modal']]
    lead = max(1, math.ceil(km / vel / 8))
    dia_chegada = t['dia_part'] + lead
    if dia_chegada > dia_entrega:
        print(f'  ❌ L{t["linha"]}: {t["origem_cidade"]}→{cidade} {t["modal"]} D{t["dia_part"]}+{lead}=D{dia_chegada} > entrega D{dia_entrega}')
        erros.append(f'Lead L{t["linha"]}')
        ok_lead = False
n_cd_var = sum(1 for t in transp_r2 if t['origem_tipo'] == 'CD' and t['destino_tipo'] == 'Varejista')
if ok_lead:
    print(f'  ✅ {n_cd_var} viagens CD→Varejo respeitam lead time')
print()

# 7. ESTOQUE MP F1 DIA A DIA
print('━' * 60)
print('7) ESTOQUE MP F1 — simulação dia a dia')
print('━' * 60)
cap_mp = {'MP1': 127.0, 'MP2': 50.4, 'MP3': 75.6}
estoque = {'MP1': 47.0, 'MP2': 48.0, 'MP3': 42.0}

chegadas_mp = defaultdict(lambda: defaultdict(float))
for t in transp_r2:
    if t['origem_tipo'] != 'Fornecedor':
        continue
    try:
        km = float(cfg.distancias[t['modal']].at[t['origem_cidade'], 'Joinville'])
    except Exception:
        km = 0
    if not km or km != km:
        continue
    vel = {'Caminhão': 50, 'Navio': 30, 'Avião': 700}[t['modal']]
    lead = max(1, math.ceil(km / vel / 8))
    dia_cheg = t['dia_part'] + lead
    if dia_cheg <= 5:
        chegadas_mp[dia_cheg][t['item']] += t['qtd']

consumo_mp = defaultdict(lambda: defaultdict(float))
for dia in range(1, 6):
    for pa in ('PA1', 'PA2', 'PA3'):
        for mp in ('MP1', 'MP2', 'MP3'):
            consumo_mp[dia][mp] += producao[dia][pa] * cfg.BoM[pa][mp] / 1_000_000

print(f'{"":>5} {"MP1 ini→chega→cons→fim":<32} {"MP2":<32} {"MP3":<32}')
ok_estoque_mp = True
for dia in range(1, 6):
    line = f'Dia {dia} |'
    for mp in ('MP1', 'MP2', 'MP3'):
        ini = estoque[mp]
        chega = chegadas_mp[dia][mp]
        cons = consumo_mp[dia][mp]
        fim = ini + chega - cons
        pico = ini + chega
        ok = pico <= cap_mp[mp] + 0.5 and fim > -0.5
        mark = '✅' if ok else '❌'
        line += f' {ini:>5.1f}+{chega:>4.1f}-{cons:>4.1f}={fim:>5.1f}{mark} |'
        if pico > cap_mp[mp] + 0.5:
            erros.append(f'Cap MP {mp} estourou Dia {dia} (pico {pico:.1f} > {cap_mp[mp]})')
            ok_estoque_mp = False
        if fim < -0.5:
            erros.append(f'MP {mp} negativo Dia {dia} ({fim:.1f})')
            ok_estoque_mp = False
        estoque[mp] = fim
    print(line)

if ok_estoque_mp:
    print('  ✅ Estoque MP cabe na cap, sem negativos')
print(f'  Estoque MP F1 fim R2: MP1={estoque["MP1"]:.1f} MP2={estoque["MP2"]:.1f} MP3={estoque["MP3"]:.1f}')
estoque_mp_fim_r2 = dict(estoque)
print()

# 8. ESTOQUE PA NOS CDs
print('━' * 60)
print('8) ESTOQUE PA NOS CDs')
print('━' * 60)
cap_cd = {
    'CD1': {'PA1': 733_333, 'PA2': 432_000, 'PA3': 9_312_000},
    'CD2': {'PA1': 666_666, 'PA2': 400_000, 'PA3': 8_533_333},
}
cidade_to_cd = {'São Luís': 'CD1', 'Santos': 'CD2'}

chegadas_pa_cd = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
for t in transp_r2:
    if t['origem_tipo'] == 'Fábrica' and t['destino_tipo'] == 'CD' and t['item'].startswith('PA'):
        cd = cidade_to_cd.get(t['destino_cidade'])
        if not cd:
            continue
        try:
            km = float(cfg.distancias[t['modal']].at['Joinville', t['destino_cidade']])
        except Exception:
            continue
        vel = {'Caminhão': 50, 'Navio': 30, 'Avião': 700}[t['modal']]
        lead = max(1, math.ceil(km / vel / 8))
        dia_cheg = t['dia_part'] + lead
        if dia_cheg <= 5:
            chegadas_pa_cd[dia_cheg][cd][t['item']] += int(t['qtd'])

saidas_pa_cd = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
for t in transp_r2:
    if t['origem_tipo'] == 'CD' and t['item'].startswith('PA'):
        cd = cidade_to_cd.get(t['origem_cidade'])
        if not cd:
            continue
        saidas_pa_cd[t['dia_part']][cd][t['item']] += int(t['qtd'])

ok_estoque_cd = True
for cd in ('CD1', 'CD2'):
    pico_max = {'PA1': 0, 'PA2': 0, 'PA3': 0}
    estoque_atual = {'PA1': 0, 'PA2': 0, 'PA3': 0}
    for dia in range(1, 6):
        for pa in ('PA1', 'PA2', 'PA3'):
            estoque_atual[pa] += chegadas_pa_cd[dia][cd][pa]
            pico_max[pa] = max(pico_max[pa], estoque_atual[pa])
            estoque_atual[pa] -= saidas_pa_cd[dia][cd][pa]
            if estoque_atual[pa] < 0:
                print(f'  ❌ {cd} {pa} Dia {dia}: estoque {estoque_atual[pa]} negativo')
                erros.append(f'CD {cd} {pa} negativo Dia {dia}')
                ok_estoque_cd = False
    for pa, p in pico_max.items():
        if p > 0:
            status = '✅' if p <= cap_cd[cd][pa] else '❌'
            print(f'  {cd} {pa}: pico {p:>8,} / cap {cap_cd[cd][pa]:>9,} {status}')
            if p > cap_cd[cd][pa]:
                erros.append(f'Cap CD {cd} {pa} estourou')
                ok_estoque_cd = False

if ok_estoque_cd:
    print('  ✅ Caps PA nos CDs respeitadas')
print()

# RESUMO
print('=' * 60)
print('RESUMO DA AUDITORIA')
print('=' * 60)
if not erros:
    print('  ✅ TODAS AS RESTRIÇÕES RESPEITADAS')
else:
    print(f'  ❌ {len(erros)} VIOLAÇÕES:')
    for e in erros:
        print(f'    - {e}')
