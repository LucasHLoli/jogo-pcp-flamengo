"""State machine multi-rodada: consolida tudo o que o solver precisa saber
sobre Flamengo no início da rodada N.

Para a rodada N, o estado consiste em:
  - estoque MP em F1 (do PDF ESTOQUES da rodada N)
  - estoque PA em cada CD (do PDF ESTOQUES)
  - MP em-trânsito (ordens de Forn→F1 das rodadas 1..N-1 cujo dia_chegada cai
    em N) — POR DIA da rodada N
  - PA em-trânsito (similar para F1→CD e CD→Varejo, mas em geral o jogo trata
    cada rodada de forma independente para PA porque PA descartado se chega
    fora do dia)
  - histórico de DRE (R1, R2, ..., R(N-1) resultados)

Lê:
  - rodadas/rodada_N/FLAMENGO.xlsm (instalações + SOL_TRANSP histórica)
  - rodadas/FLAMENGO_ALL.xlsm (consolidado se existir)
  - rodadas/rodada_N/ESTOQUES_FLAMENGO.pdf (parsing manual no código)
"""
from __future__ import annotations
import json
import re
import sys
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Any, Dict, List

import openpyxl

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))
from src.config import Config
from src.io_xlsm import ler_instalacoes

_DIA = re.compile(r"Dia\s*(\d+)")
_ROD = re.compile(r"Rodada_(\d+)")


@dataclass
class EstadoRodada:
    """Estado conhecido no INÍCIO da rodada N (= fim da rodada N-1)."""

    rodada: int
    # Estoques no dia (N-1)*5 (fim da rodada anterior)
    estoque_mp_ton: Dict[str, float] = field(default_factory=dict)        # MP1/MP2/MP3
    estoque_pa_cd: Dict[str, Dict[str, int]] = field(default_factory=dict) # cd → pa → qty

    # MP em-trânsito chegando DURANTE a rodada N (de ordens de N-1 ou antes)
    # Lista de dicts: {dia_rel, mp, qtd, origem, modal}
    mp_em_transito: List[Dict[str, Any]] = field(default_factory=list)

    # DREs históricas (R1, R2, ..., R(N-1))
    historico_dre: List[float] = field(default_factory=list)

    # Capacidades derivadas (read-only)
    cap_mp_ton: Dict[str, float] = field(default_factory=dict)
    cap_pa_cd_un: Dict[str, Dict[str, int]] = field(default_factory=dict)
    cap_min_dia: int = 0
    fab_principal: str = "F1"
    fab_cidade: str = ""
    cds_info: Dict[str, str] = field(default_factory=dict)

    def to_json(self) -> str:
        return json.dumps(asdict(self), ensure_ascii=False, indent=2)

    @classmethod
    def from_json(cls, s: str) -> "EstadoRodada":
        d = json.loads(s)
        return cls(**d)


def _parse_dia(value) -> int | None:
    if value is None:
        return None
    m = _DIA.search(str(value))
    return int(m.group(1)) if m else None


def _parse_rodada(value) -> int | None:
    if value is None:
        return None
    m = _ROD.search(str(value))
    return int(m.group(1)) if m else None


def consolidar_estado(rodada_n: int, flamengo_path: Path,
                      cfg: Config,
                      estoque_mp_pdf: Dict[str, float],
                      estoque_pa_cd_pdf: Dict[str, Dict[str, int]],
                      historico_dre: List[float],
                      lead_table_path: Path | None = None) -> EstadoRodada:
    """Constrói o EstadoRodada para a rodada N.

    Args:
        rodada_n: número da rodada atual (e.g. 3)
        flamengo_path: caminho para o FLAMENGO.xlsm desta rodada
                       (contém SOL_TRANSP histórico R1..R(N-1))
        cfg: Config (parâmetros do jogo)
        estoque_mp_pdf: do PDF ESTOQUES (fim da rodada N-1)
        estoque_pa_cd_pdf: do PDF ESTOQUES
        historico_dre: lista de resultados acumulados R1..R(N-1)
        lead_table_path: caminho data/lead_times.json (default usa BASE/data/)
    """
    # Lookup leads
    lt_path = lead_table_path or BASE / "data" / "lead_times.json"
    lead_tab = json.loads(lt_path.read_text(encoding="utf-8"))

    def lt(modal, o, d):
        if o == d:
            return 0
        return lead_tab.get(modal, {}).get(o, {}).get(d)

    # Instalações
    inst = ler_instalacoes(flamengo_path)
    f1 = inst["fabricas"]["F1"]
    cds_info = {cd: d["cidade"] for cd, d in inst["cds"].items()}

    cap_mp_ton = {
        mp: f1["area_mp"][mp] * cfg.capacidades["pe_direito_deposito_m"] * cfg.densidades_mp[mp]
        for mp in ("MP1", "MP2", "MP3")
    }
    cap_pa_cd_un = {
        cd: {pa: int(d["area_pa"][pa] * cfg.capacidades["pe_direito_deposito_m"]
                     * cfg.densidades_pa[pa] / cfg.peso_un_ton[pa])
             for pa in ("PA1", "PA2", "PA3")}
        for cd, d in inst["cds"].items()
    }
    cap_min_dia = f1["maquinas"] * f1["turnos"] * 8 * 60

    # Parse SOL_TRANSP do FLAMENGO.xlsm. Cada linha Forn→F1 de rodada anterior
    # cuja data_chegada cai DENTRO da rodada N entra em mp_em_transito.
    wb = openpyxl.load_workbook(flamengo_path, keep_vba=True, data_only=True)
    ws = wb["SOL_TRANSP"]

    mp_em_transito: List[Dict[str, Any]] = []
    dia_inicio_rodada_n_abs = (rodada_n - 1) * 5 + 1  # e.g. R3 → abs 11
    dia_fim_rodada_n_abs = rodada_n * 5               # e.g. R3 → abs 15

    for r in range(5, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if not val:
            continue
        rod = _parse_rodada(val)
        if rod is None or rod >= rodada_n:
            continue
        origem_tipo = ws.cell(r, 2).value
        if (origem_tipo or "").strip() != "Fornecedor":
            continue
        origem_cid = ws.cell(r, 3).value
        dia_raw = _parse_dia(ws.cell(r, 4).value)
        if dia_raw is None:
            continue
        # Detecta absoluto vs relativo
        dia_abs_part = dia_raw if dia_raw > 5 else (rod - 1) * 5 + dia_raw
        modal = ws.cell(r, 5).value
        item = ws.cell(r, 6).value
        try:
            qtd = float(ws.cell(r, 7).value or 0)
        except (TypeError, ValueError):
            continue
        if not item or not item.startswith("MP"):
            continue
        lt_v = lt(modal, origem_cid, f1["cidade"])
        if lt_v is None:
            continue
        dia_abs_cheg = dia_abs_part + lt_v
        # Cai dentro da rodada N?
        if dia_inicio_rodada_n_abs <= dia_abs_cheg <= dia_fim_rodada_n_abs:
            dia_rel = dia_abs_cheg - (rodada_n - 1) * 5  # 1..5 dentro de N
            mp_em_transito.append({
                "dia_rel": dia_rel,
                "dia_abs": dia_abs_cheg,
                "mp": item,
                "qtd": qtd,
                "origem": origem_cid,
                "modal": modal,
                "lt": lt_v,
                "rodada_origem": rod,
            })

    # Ordena
    mp_em_transito.sort(key=lambda x: (x["dia_rel"], x["mp"]))

    return EstadoRodada(
        rodada=rodada_n,
        estoque_mp_ton=dict(estoque_mp_pdf),
        estoque_pa_cd=dict(estoque_pa_cd_pdf),
        mp_em_transito=mp_em_transito,
        historico_dre=list(historico_dre),
        cap_mp_ton=cap_mp_ton,
        cap_pa_cd_un=cap_pa_cd_un,
        cap_min_dia=cap_min_dia,
        fab_principal="F1",
        fab_cidade=f1["cidade"],
        cds_info=cds_info,
    )


def estado_r3_flamengo() -> EstadoRodada:
    """Helper: constrói o estado conhecido para Flamengo no início de R3."""
    cfg = Config.load(BASE)
    flam = BASE / "rodadas" / "rodada_3" / "FLAMENGO.xlsm"
    # Do PDF ESTOQUES R3 (fim R2 = Dia 10)
    estoque_mp = {"MP1": 78.98, "MP2": 50.36, "MP3": 48.14}
    estoque_pa = {cd: {"PA1": 0, "PA2": 0, "PA3": 0} for cd in ("CD1", "CD2")}
    historico = [-5_602_321.0, -1_128_560.0]  # DRE oficial (R2 ainda subestimado)
    return consolidar_estado(3, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r4_flamengo() -> EstadoRodada:
    """Helper: estado conhecido para Flamengo no início de R4 (= fim de R3).

    Fonte: rodadas/rodada_4/ESTOQUES_FLAMENGO.pdf (saldos no Dia 15).
      MP F1 Joinville: MP1≈0, MP2=0,02t, MP3=4,81t
      PA em CD: CD2 (Santos) PA2 = 105.106; restante 0.

    NOTA: além dos 105.106 PA2 em CD2, há ~359k PA2 que foram embarcados F1→CD
    nos Dias 14/15 de R3 (São Luís 70.948 + Santos 288.339) e estavam EM TRÂNSITO
    no fim de R3. Se o jogo os entregar como estoque utilizável em R4, a posição
    de PA2 sobe pra ~464k. Isso NÃO está incluído aqui (conservador: só o saldo
    confirmado no PDF). A verificar contra o realizado de R4.
    """
    cfg = Config.load(BASE)
    flam = BASE / "rodadas" / "rodada_4" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 0.0, "MP2": 0.02, "MP3": 4.81}
    estoque_pa = {
        "CD1": {"PA1": 0, "PA2": 0, "PA3": 0},          # São Luís
        "CD2": {"PA1": 0, "PA2": 105_106, "PA3": 0},    # Santos (buffer PA2 confirmado)
    }
    # DRE oficial realizada (rodada_4/DRE_FLAMENGO.pdf): R1, R2, R3
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0]
    return consolidar_estado(4, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r5_flamengo() -> EstadoRodada:
    """Estado de início de R5 (= fim de R4). R4 NÃO foi submetida (vendemos 0),
    então o estoque é o que carregou de R3 (o em-trânsito de PA2 chegou: 464k).

    Fonte: rodadas/rodada_5/ESTOQUES_FLAMENGO.pdf (saldos no Dia 20).
      MP F1: MP1=7,20t, MP2=12,02t, MP3=4,81t (o em-trânsito de R3 chegou; nada consumido)
      PA: CD1 (São Luís) PA2=70.948 ; CD2 (Santos) PA2=393.359 ; resto 0.
      MP em-trânsito p/ R5: NENHUM (R4 não comprou nada).
    R5 = PA1 @ R$69.
    """
    cfg = Config.load(BASE)
    flam = BASE / "rodadas" / "rodada_5" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 7.20, "MP2": 12.02, "MP3": 4.81}
    estoque_pa = {
        "CD1": {"PA1": 0, "PA2": 70_948, "PA3": 0},      # São Luís
        "CD2": {"PA1": 0, "PA2": 393_359, "PA3": 0},     # Santos (98,3% do cap PA2)
    }
    # R4 não submetida → resultado R4 ≈ só custos fixos + carregamento (negativo)
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_355_000.0]
    return consolidar_estado(5, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r6_flamengo() -> EstadoRodada:
    """Estado de início de R6 (= fim de R5). R5 FOI submetida — a previsão do
    solver cravou o estado inicial de R6 (MP1 66,4→66,41; MP2 38,7→38,71;
    MP3 23,6→23,57; PA idênticos), confirmando que o buffer de MP é REAL e a
    cascata da R4 foi quebrada (R6 começa abastecida).

    Fonte: rodadas/rodada_6/ESTOQUES_FLAMENGO.pdf (saldos no Dia 25).
      MP F1 Joinville: MP1=66,41t, MP2=38,71t, MP3=23,57t
      PA: CD1 (São Luís) PA2=70.948 ; CD2 (Santos) PA1=32.622, PA2=393.359, PA3=462.358.
      MP em-trânsito p/ R6: NENHUM — todos os pedidos de R5 (último Dia 24, lead≤3)
      chegam até o Dia 25, dentro de R5 (verificado contra lead_times).
    R6 = PA2 @ R$48 → a rodada que finalmente usa o buffer de 464k PA2 em estoque.
    """
    cfg = Config.load(BASE)
    flam = BASE / "rodadas" / "rodada_6" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 66.41, "MP2": 38.71, "MP3": 23.57}
    estoque_pa = {
        "CD1": {"PA1": 0, "PA2": 70_948, "PA3": 0},              # São Luís
        "CD2": {"PA1": 32_622, "PA2": 393_359, "PA3": 462_358},  # Santos
    }
    # DRE oficial realizada (rodadas/rodada_6/DRE_FLAMENGO.pdf): R1..R5.
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_356_169.0, 1_603_752.0]
    return consolidar_estado(6, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r7_flamengo() -> EstadoRodada:
    """Estado de início de R7 (= fim de R6). R6 FOI submetida — a previsão do solver
    cravou de novo o fim de R6 (MP1 50,3 MP2 50,4 MP3 18,0; PA idênticos), validado
    vs Estoques R7 (erro do resultado 0,03% após corrigir os 2 carregamentos).

    Fonte: rodadas/rodada_7/ESTOQUES_FLAMENGO.pdf (saldos no Dia 30).
      MP F1 Joinville: MP1=50,29t, MP2=50,39t, MP3=18,00t
      PA: CD1 (São Luís) PA1=50.872, PA2=156.404, PA3=160.000 ;
          CD2 (Santos)   PA1=32.622, PA2=0,       PA3=462.358.
      MP em-trânsito p\\ R7: NENHUM — o balanço de MP da R6 (ini+compra-consumo) fecha
      exatamente no estoque físico do Dia 30, logo toda MP comprada em R6 já chegou.
    R7 = PA2 @ R$44 (IND rodada_7). Carteira real R7 = 895.793 PA2.
    """
    cfg = Config.load(BASE)
    flam = BASE / "rodadas" / "rodada_7" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 50.29, "MP2": 50.39, "MP3": 18.00}
    estoque_pa = {
        "CD1": {"PA1": 50_872, "PA2": 156_404, "PA3": 160_000},  # São Luís
        "CD2": {"PA1": 32_622, "PA2": 0,       "PA3": 462_358},  # Santos
    }
    # DRE oficial realizada (rodadas/rodada_7/DRE_FLAMENGO.pdf): R1..R6.
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_356_169.0,
                 1_603_752.0, 38_569_647.0]
    return consolidar_estado(7, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r8_flamengo() -> EstadoRodada:
    """Estado de início de R8 (= fim de R7). R7 FOI submetida e o solver cravou o fim
    de R7: a previsão bateu o estoque real unidade-a-unidade (resultado 0,018% após
    calibrar o avião do frete a 11,6 — ver project_regra_frete_calibrada).

    Fonte: solver_v2/rodadas/rodada_8/ESTOQUES_FLAMENGO.pdf (saldos no Dia 35).
      MP F1 Joinville: MP1=49,75t, MP2=49,48t, MP3=41,63t
      PA: CD1 (São Luís) PA1=50.872, PA2=65.550,  PA3=160.000 ;
          CD2 (Santos)   PA1=32.622, PA2=162.334, PA3=462.358.
      MP em-trânsito p/ R8: NENHUM — toda MP comprada em R7 entregou até o Dia 35
      (verificado: entregas dos 11 caminhões caem nos dias 32-35).
    R8 = PA3 @ R$24 (IND rodada_8). Carteira real R8 = 1.335.398 PA3 (todos Dia 39).
    Há 622.358 PA3 já em estoque nos CDs — o over-build começa a escoar.
    """
    cfg = Config.load(BASE)
    flam = BASE / "solver_v2" / "rodadas" / "rodada_8" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 49.75, "MP2": 49.48, "MP3": 41.63}
    estoque_pa = {
        "CD1": {"PA1": 50_872, "PA2": 65_550,  "PA3": 160_000},  # São Luís
        "CD2": {"PA1": 32_622, "PA2": 162_334, "PA3": 462_358},  # Santos
    }
    # DRE oficial realizada (solver_v2/rodadas/rodada_8/DRE_FLAMENGO.pdf): R1..R7.
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_356_169.0,
                 1_603_752.0, 38_569_647.0, 28_520_764.0]
    return consolidar_estado(8, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r9_flamengo() -> EstadoRodada:
    """Estado de início de R9 (= fim de R8). R8 FOI submetida; a previsão do solver_v3
    cravou o estoque real (resultado 0,018% — ver análise R8→R9).

    Fonte: solver_v3/rodadas/rodada_9/ESTOQUES_FLAMENGO.pdf (saldos no Dia 40).
      MP F1 Joinville: MP1=79,50t, MP2=41,43t, MP3=24,60t
      PA: CD1 (São Luís) PA1=50.872, PA2=124.019, PA3=111.738 ;
          CD2 (Santos)   PA1=32.622, PA2=162.334, PA3=0.
      MP em-trânsito p/ R9: NENHUM — toda MP comprada em R8 entregou até o Dia 40
      (verificado: MP1 Manaus chega Dia 39-40, MP2 Cuiabá Dia 38, MP3 P.Alegre Dia 39).
      consolidar_estado confirma isso lendo o SOL_TRANSP do FLAMENGO.xlsm.
    R9 = PA2 @ R$55 (IND rodada_9). Carteira real R9 = 928.973 PA2 (todos Dia 43).
    """
    cfg = Config.load(BASE)
    flam = BASE / "solver_v3" / "rodadas" / "rodada_9" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 79.50, "MP2": 41.43, "MP3": 24.60}
    estoque_pa = {
        "CD1": {"PA1": 50_872, "PA2": 124_019, "PA3": 111_738},  # São Luís
        "CD2": {"PA1": 32_622, "PA2": 162_334, "PA3": 0},        # Santos
    }
    # DRE oficial realizada (solver_v3/rodadas/rodada_9/DRE_FLAMENGO.pdf): R1..R8.
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_356_169.0,
                 1_603_752.0, 38_569_647.0, 28_520_764.0, 24_541_201.0]
    return consolidar_estado(9, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r10_flamengo() -> EstadoRodada:
    """Estado de início de R10 (= fim de R9). R9 FOI submetida; previsão cravou
    (resultado 0,087% no plano enviado, após corrigir o frete c/ CT-e).

    Fonte: solver_v3/rodadas/rodada_10/ESTOQUES_FLAMENGO.pdf (saldos no Dia 45).
      MP F1 Joinville: MP1=116,66t, MP2=50,40t (cheio), MP3=27,37t
      PA: CD1 (São Luís) PA1=50.872, PA2=974,    PA3=111.738 ;
          CD2 (Santos)   PA1=32.622, PA2=150.789, PA3=0.
      MP em-trânsito p/ R10: consolidar_estado detecta automaticamente do SOL_TRANSP.
    R10 = PA3 @ R$27 (IND rodada_10). Carteira real R10 = 1.492.994 PA3 (todos Dia 47 = dia rel 2).
    """
    cfg = Config.load(BASE)
    flam = BASE / "solver_v3" / "rodadas" / "rodada_10" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 116.66, "MP2": 50.40, "MP3": 27.37}
    estoque_pa = {
        "CD1": {"PA1": 50_872, "PA2": 974,     "PA3": 111_738},  # São Luís
        "CD2": {"PA1": 32_622, "PA2": 150_789, "PA3": 0},        # Santos
    }
    # DRE oficial realizada (solver_v3/rodadas/rodada_10/DRE_FLAMENGO.pdf): R1..R9.
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_356_169.0,
                 1_603_752.0, 38_569_647.0, 28_520_764.0, 24_541_201.0, 40_749_742.0]
    return consolidar_estado(10, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r11_flamengo() -> EstadoRodada:
    """Estado de início de R11 (= fim de R10). R10 FOI submetida (plano PROTEGIDO: buffer
    400k PA2 em Santos). Resultado real 31,2M (previsão 0,043%); o bug dos 955 derrubou
    955 frascos de BH (NS 88,4%).

    Fonte: solver_v3/rodadas/rodada_11/ESTOQUES_FLAMENGO.pdf (saldos no Dia 50).
      MP F1 Joinville: MP1=7,27t (BAIXO!), MP2=0, MP3=0,49t
      PA: CD1 (São Luís) PA1=50.872, PA2=974,    PA3=1.740 ;
          CD2 (Santos)   PA1=32.622, PA2=399.685, PA3=0.
      MP em-trânsito p/ R11: consolidar_estado detecta do SOL_TRANSP.
    R11 = PA1 @ R$77 (IND rodada_11). Carteira real R11 = 381.544 PA1 (espalhada dias 52-55).
    ATENÇÃO: MP1 baixíssimo (7,27t) e PA1 consome 60g MP1/frasco → precisa comprar ~16t MP1;
    o buffer de 400k PA2 NÃO ajuda (rodada é PA1) e custa ~R$200k/rodada de carregamento.
    """
    cfg = Config.load(BASE)
    flam = BASE / "solver_v3" / "rodadas" / "rodada_11" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 7.27, "MP2": 0.0, "MP3": 0.49}
    estoque_pa = {
        "CD1": {"PA1": 50_872, "PA2": 974,     "PA3": 1_740},  # São Luís
        "CD2": {"PA1": 32_622, "PA2": 399_685, "PA3": 0},      # Santos
    }
    # DRE oficial realizada (solver_v3/rodadas/rodada_11/DRE_FLAMENGO.pdf): R1..R10.
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_356_169.0,
                 1_603_752.0, 38_569_647.0, 28_520_764.0, 24_541_201.0, 40_749_742.0, 31_227_724.0]
    return consolidar_estado(11, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r12_flamengo() -> EstadoRodada:
    """Estado de início de R12 (= fim de R11). R11 FOI submetida (multimodal + inter-rodada);
    resultado real 19,3M (previsão 0,35%), NS 100%. ÚLTIMA RODADA.

    Fonte: solver_v3/rodadas/rodada_12/ESTOQUES_FLAMENGO.pdf (saldos no Dia 55).
      MP F1 Joinville: MP1=0,12t, MP2=0,77t, MP3=2,10t (baixo — o grosso está EM TRÂNSITO)
      PA: CD1 (São Luís) PA1=4.513, PA2=974,    PA3=1.740 ;
          CD2 (Santos)   PA1=0,    PA2=399.685, PA3=0.
      MP em-trânsito p/ R12: o ~133t pré-pedido na R11 (inter-rodada) chega agora —
      consolidar_estado detecta do SOL_TRANSP (lê modal+lead corretamente).
    R12 = RODADA DUPLA: PA1 @ R$84 + PA2 @ R$44 (IND rodada_12). Carteiras: PA1 398.133, PA2 967.681.
    NOTA: última rodada → rodar com horizonte 0 (sem pré-pedir MP pra futuro inexistente).
    Os 400k de PA2 em Santos finalmente são usados (R12 tem demanda PA2).
    """
    cfg = Config.load(BASE)
    flam = BASE / "solver_v3" / "rodadas" / "rodada_12" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 0.12, "MP2": 0.77, "MP3": 2.10}
    estoque_pa = {
        "CD1": {"PA1": 4_513, "PA2": 974,     "PA3": 1_740},  # São Luís
        "CD2": {"PA1": 0,     "PA2": 399_685, "PA3": 0},      # Santos
    }
    # DRE oficial realizada (solver_v3/rodadas/rodada_12/DRE_FLAMENGO.pdf): R1..R11.
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_356_169.0, 1_603_752.0,
                 38_569_647.0, 28_520_764.0, 24_541_201.0, 40_749_742.0, 31_227_724.0, 19_296_696.0]
    return consolidar_estado(12, flam, cfg, estoque_mp, estoque_pa, historico)


def estado_r13_flamengo() -> EstadoRodada:
    """Estado de início de R13 (= fim de R12). R12 (dupla PA1+PA2) FOI submetida, mas o
    professor "trolou": a demanda real exigiu MUITO mais PA1 do que tínhamos em buffer
    (apostamos os 400k de buffer em PA2), então o PA1 vendeu só 52,8% (210.016 de ~398k)
    enquanto o PA2 fez 91,2%. NS empresa R12 = 80%. Resultado ~R$50,5M (vs 58,1M previsto).
    Fim de R12 ZEROU o estoque de PA (vendeu/escoou tudo) → entramos em R13 SEM buffer.

    Fonte: solver_v3/rodadas/rodada_13/ESTOQUES_FLAMENGO.pdf (saldos no Dia 60).
      MP F1 Joinville: MP1=21,62t (sobra — compramos MP1 demais p/ um PA1 que não vendeu),
                       MP2=0,00t, MP3=0,17t
      PA: CD1 (São Luís) PA3=1.740 (resto zero) ; CD2 (Santos) tudo zero.
      MP em-trânsito p/ R13: NENHUM — R12 rodou horizonte 0 (era "última rodada"), sem
      pré-pedido inter-rodada.
    R13 = RODADA TRIPLA: PA1 @ R$68 + PA2 @ R$44 + PA3 @ R$19 (IND rodada_13).
    Demanda total 1.996.187 (PA1 304.126 · PA2 696.733 · PA3 995.328); capacidade ~50.400
    min/sem só cobre ~84% e o Dia 1 pede 633.709 → NS será limitado por capacidade+buffer-zero.
    """
    cfg = Config.load(BASE)
    flam = BASE / "solver_v3" / "rodadas" / "rodada_13" / "FLAMENGO.xlsm"
    estoque_mp = {"MP1": 21.62, "MP2": 0.0, "MP3": 0.17}
    estoque_pa = {
        "CD1": {"PA1": 0, "PA2": 0, "PA3": 1_740},  # São Luís (único saldo de PA)
        "CD2": {"PA1": 0, "PA2": 0, "PA3": 0},       # Santos (zerado)
    }
    # DRE oficial realizada R1..R12. R12 ~50,5M reconstruído (a DRE da rodada_13 rolou a
    # janela e ficou inconsistente — "troll"; receita acum. crava R12 = R$56,49M de venda).
    historico = [-5_602_321.0, -11_554_929.0, 35_617_989.0, -1_356_169.0, 1_603_752.0,
                 38_569_647.0, 28_520_764.0, 24_541_201.0, 40_749_742.0, 31_227_724.0,
                 19_296_696.0, 50_459_441.0]
    return consolidar_estado(13, flam, cfg, estoque_mp, estoque_pa, historico)


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    if "--r6" in sys.argv:
        estado = estado_r6_flamengo()
    elif "--r4" in sys.argv:
        estado = estado_r4_flamengo()
    else:
        estado = estado_r3_flamengo()
    print(f"=== ESTADO R{estado.rodada} ===")
    print(f"Estoque MP: {estado.estoque_mp_ton}")
    print(f"Estoque PA: {estado.estoque_pa_cd}")
    print(f"Cap MP F1: {estado.cap_mp_ton}")
    print(f"Cap min/dia: {estado.cap_min_dia}")
    print(f"\nMP em-trânsito (de R1..R{estado.rodada-1} chegando em R{estado.rodada}):")
    if not estado.mp_em_transito:
        print("  (nenhuma)")
    for x in estado.mp_em_transito:
        print(f"  Dia rel {x['dia_rel']} (abs {x['dia_abs']}): {x['qtd']:.1f}t {x['mp']} "
              f"de {x['origem']} via {x['modal']} (lt={x['lt']}d, shipped R{x['rodada_origem']})")
    print(f"\nHistórico DRE: {estado.historico_dre}")
