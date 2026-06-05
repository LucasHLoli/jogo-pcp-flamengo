"""Compara qualidade HEURÍSTICA vs SOLVER (BUFFER) lendo as planilhas FLAMENGO.xlsm existentes.

Métricas avaliadas:
  - NS R3 (% OPs entregues no dia EXATO)
  - Receita / Custos / Lucro
  - Transportes usados (vs 220)
  - Utilização fábrica
  - Buffer PA2 R4 (CD2 + em-trânsito)
  - MP em-trânsito chegando em R4
  - Coerência (negativos, descartes, capacidades)
"""
from __future__ import annotations
import io
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from src.config import Config
from src.io_xlsm import ler_instalacoes
from src.planner_v3 import lead_dias, km_rota, custo_total_modal

cfg = Config.load(BASE)

# OPs R3 oficiais
OPS = [{"cidade": c, "pa": "PA3", "qtd": q, "dia_entrega": d} for c, q, d in [
    ("Belém", 20155, 5), ("Belo Horizonte", 70544, 3), ("Brasília", 117573, 3),
    ("Campinas", 56435, 2), ("Campo Grande", 23515, 3), ("Cuiabá", 28218, 3),
    ("Curitiba", 103464, 2), ("Fortaleza", 68528, 5), ("Goiânia", 65841, 3),
    ("João Pessoa", 40311, 4), ("Joinville", 28218, 2), ("Maceió", 40311, 4),
    ("Manaus", 20155, 5), ("Natal", 40311, 5), ("Porto Alegre", 103464, 2),
    ("Recife", 60466, 4), ("Ribeirão Preto", 47029, 2), ("Rio de Janeiro", 94059, 3),
    ("Salvador", 80622, 4), ("Santos", 47029, 2), ("São Luís", 20155, 5),
    ("São Paulo", 117573, 2), ("Uberlândia", 23515, 3), ("Vitória", 14109, 3),
    ("Vitória da Conquista", 12093, 4),
]]

PRECOS_R3 = {"PA1": 80, "PA2": 50, "PA3": 32}
ESTOQUE_MP_INI = {"MP1": 78.98, "MP2": 50.36, "MP3": 48.14}
EM_TRANSITO_R2 = [{"dia": 1, "mp": "MP1", "qtd": 8.7}]
FIX = 1_123_151
MAIOR_MP = {"MP1": 56000, "MP2": 22000, "MP3": 41000}
VEL = {"PA1": 15, "PA2": 30, "PA3": 60}
forn_min = {mp: min(cfg.fornecedores[mp], key=lambda x: x[1]) for mp in ("MP1", "MP2", "MP3")}

INFRA_FLAM = BASE / "rodadas" / "rodada_3" / "FLAMENGO.xlsm"
F1 = ler_instalacoes(INFRA_FLAM)["fabricas"]["F1"]
CDS = ler_instalacoes(INFRA_FLAM)["cds"]
cap_mp = {mp: F1["area_mp"][mp] * 2 * cfg.densidades_mp[mp] for mp in ("MP1", "MP2", "MP3")}
cap_pa_cd = {cd: {pa: int(d["area_pa"][pa] * 2 * cfg.densidades_pa[pa] / cfg.peso_un_ton[pa])
                  for pa in ("PA1", "PA2", "PA3")} for cd, d in CDS.items()}
CAP_MIN_DIA = F1["maquinas"] * F1["turnos"] * 8 * 60
cidade_to_cd = {d["cidade"]: cd for cd, d in CDS.items()}

_DIA = re.compile(r"Dia\s*(\d+)")
_ROD = re.compile(r"Rodada_(\d+)")


def carregar_plano(xlsm_path: Path) -> dict:
    """Lê SOL_TRANSP (R3) e OP_FABRICAS (R3) de um FLAMENGO.xlsm."""
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=True)
    ws_t = wb["SOL_TRANSP"]
    ws_o = wb["OP_FABRICAS"]

    sol_r3 = []
    for r in range(5, ws_t.max_row + 1):
        val = ws_t.cell(r, 1).value
        if not val:
            continue
        m = _ROD.search(str(val))
        if not m or int(m.group(1)) != 3:
            continue
        dia_str = str(ws_t.cell(r, 4).value)
        m2 = _DIA.search(dia_str)
        if not m2:
            continue
        dia_raw = int(m2.group(1))
        dia_rel = dia_raw - 10 if dia_raw > 5 else dia_raw
        if not (1 <= dia_rel <= 5):
            continue
        sol_r3.append({
            "Origem": ws_t.cell(r, 2).value,
            "Cidade": ws_t.cell(r, 3).value,
            "dia_rel": dia_rel,
            "Modal": ws_t.cell(r, 5).value,
            "Tipo": ws_t.cell(r, 6).value,
            "Qtde": float(ws_t.cell(r, 7).value or 0),
            "Destino": ws_t.cell(r, 8).value,
            "Cidade_Destino": ws_t.cell(r, 9).value,
        })

    prod_r3 = {}
    for d in range(1, 6):
        prod_r3[d] = {
            "PA1": int(ws_o.cell(6 + d, 2).value or 0),
            "PA2": int(ws_o.cell(6 + d, 3).value or 0),
            "PA3": int(ws_o.cell(6 + d, 4).value or 0),
        }
    return {"sol": sol_r3, "prod": prod_r3}


def calcular_indicadores(plano: dict, label: str) -> dict:
    """Calcula todos os indicadores de um plano R3."""
    sol = plano["sol"]
    prod = plano["prod"]

    # NS: chegadas no dia exato
    cdv = defaultdict(float)
    for row in sol:
        if row["Origem"] == "CD" and row["Destino"] == "Varejista":
            lt = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
            cdv[(row["Cidade_Destino"], row["Tipo"], row["dia_rel"] + lt)] += row["Qtde"]
    qty_total = sum(o["qtd"] for o in OPS)
    qty_atend = sum(o["qtd"] for o in OPS
                    if abs(cdv.get((o["cidade"], o["pa"], o["dia_entrega"]), 0) - o["qtd"]) < 1)
    ns_pct = qty_atend / qty_total * 100

    # Receita (apenas OPs entregues no dia exato)
    receita = sum(o["qtd"] * PRECOS_R3[o["pa"]] for o in OPS
                  if abs(cdv.get((o["cidade"], o["pa"], o["dia_entrega"]), 0) - o["qtd"]) < 1)

    # Custo MP
    custo_mp = sum(row["Qtde"] * forn_min[row["Tipo"]][1]
                   for row in sol if row["Origem"] == "Fornecedor")

    # Frete
    custo_frete = 0
    for row in sol:
        modal = row["Modal"]
        item = row["Tipo"]
        qtd = row["Qtde"]
        kv = km_rota(cfg, row["Cidade"], row["Cidade_Destino"], modal) or 0
        peso = qtd if item.startswith("MP") else qtd * cfg.peso_un_ton[item]
        custo_frete += custo_total_modal(cfg, modal, kv, peso, 1)

    # Estoque MP final F1 (simulado dia a dia)
    arrivals_mp = defaultdict(lambda: defaultdict(float))
    for it in EM_TRANSITO_R2:
        arrivals_mp[it["dia"]][it["mp"]] += it["qtd"]
    em_transito_fim_r3 = defaultdict(float)
    for row in sol:
        if row["Origem"] == "Fornecedor":
            lt = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
            dia_cheg = row["dia_rel"] + lt
            if 1 <= dia_cheg <= 5:
                arrivals_mp[dia_cheg][row["Tipo"]] += row["Qtde"]
            else:
                em_transito_fim_r3[row["Tipo"]] += row["Qtde"]
    stk_mp = dict(ESTOQUE_MP_INI)
    neg_mp = 0
    desc_mp = 0
    for d in range(1, 6):
        for mp in ("MP1", "MP2", "MP3"):
            pre = stk_mp[mp]
            arr = arrivals_mp[d].get(mp, 0)
            pos = pre + arr
            desc = max(0, pos - cap_mp[mp])
            pos = min(pos, cap_mp[mp])
            cons = sum(prod[d][pa] * cfg.BoM[pa][mp] / 1e6 for pa in ("PA1", "PA2", "PA3"))
            end = pos - cons
            if end < -0.01:
                neg_mp += 1
            if desc > 0.01:
                desc_mp += 1
            stk_mp[mp] = max(0, end)

    # Estoque PA CDs final
    arr_pa = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    sai_pa = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    em_transito_pa2_fim = 0.0
    for row in sol:
        if row["Origem"] == "Fábrica" and row["Destino"] == "CD":
            cd = cidade_to_cd[row["Cidade_Destino"]]
            lt = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
            dia_cheg = row["dia_rel"] + lt
            if 1 <= dia_cheg <= 5:
                arr_pa[dia_cheg][cd][row["Tipo"]] += row["Qtde"]
            elif row["Tipo"] == "PA2":
                em_transito_pa2_fim += row["Qtde"]
        elif row["Origem"] == "CD":
            cd = cidade_to_cd[row["Cidade"]]
            sai_pa[row["dia_rel"]][cd][row["Tipo"]] += row["Qtde"]

    stk_pa = {cd: {pa: 0 for pa in ("PA1", "PA2", "PA3")} for cd in CDS}
    neg_pa = 0
    over_pa = 0
    for d in range(1, 6):
        for cd in ("CD1", "CD2"):
            for pa in ("PA1", "PA2", "PA3"):
                pre = stk_pa[cd][pa]
                arr = arr_pa[d][cd].get(pa, 0)
                sai = sai_pa[d][cd].get(pa, 0)
                end = pre + arr - sai
                if end < -0.5:
                    neg_pa += 1
                if end > cap_pa_cd[cd][pa]:
                    over_pa += 1
                stk_pa[cd][pa] = max(0, end)

    # Carregamento (1% por dia × 5 dias = 5% custo carregamento sobre estoque médio; aqui usamos fim)
    carreg_mp = sum(stk_mp[mp] * MAIOR_MP[mp] * 0.01 for mp in ("MP1", "MP2", "MP3"))
    carreg_pa = sum(stk_pa[cd][pa] * PRECOS_R3[pa] * 0.01
                    for cd in CDS for pa in ("PA1", "PA2", "PA3"))

    custo_total = FIX + custo_mp + custo_frete + carreg_mp + carreg_pa
    lucro = receita - custo_total

    # Utilização fábrica
    min_por_dia = {d: sum(prod[d][pa] / VEL[pa] for pa in ("PA1", "PA2", "PA3")) for d in range(1, 6)}
    util_media = sum(min_por_dia.values()) / (5 * CAP_MIN_DIA) * 100

    # Cap modal violations
    cap_modal_ton = {"Avião": 1, "Caminhão": 24, "Navio": 100}
    cap_un_pa = {"Avião": {"PA1": 3333, "PA2": 4000, "PA3": 6666},
                 "Caminhão": {"PA1": 80000, "PA2": 96000, "PA3": 160000},
                 "Navio": {"PA1": 333333, "PA2": 400000, "PA3": 666666}}
    viol_modal = 0
    for row in sol:
        if row["Tipo"] in ("PA1", "PA2", "PA3"):
            if row["Qtde"] > cap_un_pa[row["Modal"]][row["Tipo"]] + 0.01:
                viol_modal += 1
        elif row["Tipo"].startswith("MP"):
            if row["Qtde"] > cap_modal_ton[row["Modal"]] + 0.01:
                viol_modal += 1

    # Cap fábrica violations
    viol_min = sum(1 for d in range(1, 6) if min_por_dia[d] > CAP_MIN_DIA + 1)

    return {
        "label": label,
        "ns_pct": ns_pct,
        "qty_atend": qty_atend,
        "qty_total": qty_total,
        "n_transp": len(sol),
        "receita": receita,
        "custo_mp": custo_mp,
        "custo_frete": custo_frete,
        "carreg_mp": carreg_mp,
        "carreg_pa": carreg_pa,
        "fix": FIX,
        "custo_total": custo_total,
        "lucro": lucro,
        "util_media_pct": util_media,
        "min_por_dia": min_por_dia,
        "stk_mp_fim": stk_mp,
        "em_transito_mp_fim": dict(em_transito_fim_r3),
        "stk_pa_fim": stk_pa,
        "em_transito_pa2_fim": em_transito_pa2_fim,
        "buffer_pa2_total": stk_pa["CD1"]["PA2"] + stk_pa["CD2"]["PA2"] + em_transito_pa2_fim,
        "pa2_produzido": sum(prod[d]["PA2"] for d in range(1, 6)),
        "neg_mp": neg_mp,
        "desc_mp": desc_mp,
        "neg_pa": neg_pa,
        "over_pa": over_pa,
        "viol_modal": viol_modal,
        "viol_min": viol_min,
    }


# =================== EXECUÇÃO ===================
print("=" * 100)
print("  COMPARAÇÃO DE QUALIDADE: HEURÍSTICA vs SOLVER")
print("=" * 100)

HEUR_PATH = BASE / "rodadas" / "rodada_3" / "FLAMENGO.xlsm"
SOLVER_PATH = BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO_BUFFER.xlsm"

print(f"\nLendo HEURÍSTICA:  {HEUR_PATH}")
plano_h = carregar_plano(HEUR_PATH)
print(f"  {len(plano_h['sol'])} linhas R3, {sum(plano_h['prod'][d]['PA1']+plano_h['prod'][d]['PA2']+plano_h['prod'][d]['PA3'] for d in range(1,6)):,} un produzidas")

print(f"\nLendo SOLVER:      {SOLVER_PATH}")
plano_s = carregar_plano(SOLVER_PATH)
print(f"  {len(plano_s['sol'])} linhas R3, {sum(plano_s['prod'][d]['PA1']+plano_s['prod'][d]['PA2']+plano_s['prod'][d]['PA3'] for d in range(1,6)):,} un produzidas")

ind_h = calcular_indicadores(plano_h, "HEURÍSTICA")
ind_s = calcular_indicadores(plano_s, "SOLVER")


def fmt_R(v): return f"R$ {v:>14,.0f}"
def fmt_pct(v): return f"{v:>7.1f}%"
def fmt_int(v): return f"{v:>15,.0f}"


def linha(label, val_h, val_s, fmt, melhor="maior"):
    """Imprime linha com marcador de vencedor."""
    sh, ss = fmt(val_h), fmt(val_s)
    if isinstance(val_h, (int, float)) and isinstance(val_s, (int, float)):
        diff = val_s - val_h
        if abs(diff) < 0.01:
            tag = "="
        elif (melhor == "maior" and diff > 0) or (melhor == "menor" and diff < 0):
            tag = "✅ SOLVER"
        else:
            tag = "✅ HEUR"
    else:
        tag = ""
    print(f"  {label:<32}{sh:>22}{ss:>22}  {tag}")


print(f"\n{'='*100}")
print(f"  {'INDICADOR':<32}{'HEURÍSTICA':>22}{'SOLVER (BUFFER)':>22}  VENCEDOR")
print(f"{'='*100}")

print("\n── QUALIDADE DE ENTREGA ──")
linha("NS R3 (% no dia exato)", ind_h["ns_pct"], ind_s["ns_pct"], fmt_pct, "maior")
linha("Frascos entregues no dia", ind_h["qty_atend"], ind_s["qty_atend"], fmt_int, "maior")
linha("Frascos planejados", ind_h["qty_total"], ind_s["qty_total"], fmt_int, "maior")

print("\n── DRE ──")
linha("Receita R3", ind_h["receita"], ind_s["receita"], fmt_R, "maior")
linha("Custo MP", ind_h["custo_mp"], ind_s["custo_mp"], fmt_R, "menor")
linha("Custo Frete", ind_h["custo_frete"], ind_s["custo_frete"], fmt_R, "menor")
linha("Carregamento MP", ind_h["carreg_mp"], ind_s["carreg_mp"], fmt_R, "menor")
linha("Carregamento PA", ind_h["carreg_pa"], ind_s["carreg_pa"], fmt_R, "menor")
linha("Custos fixos", ind_h["fix"], ind_s["fix"], fmt_R, "menor")
linha("Custo TOTAL", ind_h["custo_total"], ind_s["custo_total"], fmt_R, "menor")
linha("LUCRO R3", ind_h["lucro"], ind_s["lucro"], fmt_R, "maior")

print("\n── OPERAÇÃO ──")
linha("Transportes R3 (cap 220)", ind_h["n_transp"], ind_s["n_transp"], fmt_int, "menor")
linha("Utilização fábrica média", ind_h["util_media_pct"], ind_s["util_media_pct"], fmt_pct, "qualquer")

print("\n── BUFFER R4 (estratégia conservadora) ──")
linha("PA2 produzido em R3", ind_h["pa2_produzido"], ind_s["pa2_produzido"], fmt_int, "maior")
linha("PA2 em CDs fim R3", ind_h["stk_pa_fim"]["CD1"]["PA2"]+ind_h["stk_pa_fim"]["CD2"]["PA2"],
      ind_s["stk_pa_fim"]["CD1"]["PA2"]+ind_s["stk_pa_fim"]["CD2"]["PA2"], fmt_int, "maior")
linha("PA2 em-trânsito chegando R4", ind_h["em_transito_pa2_fim"], ind_s["em_transito_pa2_fim"], fmt_int, "maior")
linha("BUFFER PA2 TOTAL para R4", ind_h["buffer_pa2_total"], ind_s["buffer_pa2_total"], fmt_int, "maior")

print("\n── MP para R4 ──")
for mp in ("MP1", "MP2", "MP3"):
    linha(f"Estoque {mp} fim R3 (t)",
          ind_h["stk_mp_fim"][mp], ind_s["stk_mp_fim"][mp],
          lambda v: f"{v:>13.2f}t", "maior")
    linha(f"{mp} em-trânsito p/ R4 (t)",
          ind_h["em_transito_mp_fim"].get(mp, 0), ind_s["em_transito_mp_fim"].get(mp, 0),
          lambda v: f"{v:>13.2f}t", "maior")

print("\n── COERÊNCIA (violações = 0 é o ideal) ──")
linha("Cap modal violada (viagens)", ind_h["viol_modal"], ind_s["viol_modal"], fmt_int, "menor")
linha("Cap fábrica violada (dias)", ind_h["viol_min"], ind_s["viol_min"], fmt_int, "menor")
linha("MP F1 negativo (eventos)", ind_h["neg_mp"], ind_s["neg_mp"], fmt_int, "menor")
linha("MP F1 descartada (eventos)", ind_h["desc_mp"], ind_s["desc_mp"], fmt_int, "menor")
linha("PA CD negativo (eventos)", ind_h["neg_pa"], ind_s["neg_pa"], fmt_int, "menor")
linha("PA CD acima cap (eventos)", ind_h["over_pa"], ind_s["over_pa"], fmt_int, "menor")


# =================== PLACAR ===================
print(f"\n{'='*100}")
print(f"  PLACAR FINAL")
print(f"{'='*100}\n")

placar = {"SOLVER": 0, "HEUR": 0, "EMPATE": 0}
criterios = [
    ("NS R3", ind_h["ns_pct"], ind_s["ns_pct"], "maior"),
    ("Lucro R3", ind_h["lucro"], ind_s["lucro"], "maior"),
    ("Custo total", ind_h["custo_total"], ind_s["custo_total"], "menor"),
    ("Transportes", ind_h["n_transp"], ind_s["n_transp"], "menor"),
    ("Buffer PA2 R4", ind_h["buffer_pa2_total"], ind_s["buffer_pa2_total"], "maior"),
    ("Coerência (sum violações)",
     ind_h["viol_modal"]+ind_h["viol_min"]+ind_h["neg_mp"]+ind_h["desc_mp"]+ind_h["neg_pa"]+ind_h["over_pa"],
     ind_s["viol_modal"]+ind_s["viol_min"]+ind_s["neg_mp"]+ind_s["desc_mp"]+ind_s["neg_pa"]+ind_s["over_pa"],
     "menor"),
]
for nome, vh, vs, melhor in criterios:
    if abs(vh - vs) < 1e-6:
        placar["EMPATE"] += 1
        sym = "="
        venc = "empate"
    elif (melhor == "maior" and vs > vh) or (melhor == "menor" and vs < vh):
        placar["SOLVER"] += 1
        sym = "✅"
        venc = "SOLVER vence"
    else:
        placar["HEUR"] += 1
        sym = "✅"
        venc = "HEURÍSTICA vence"
    print(f"  {sym} {nome:<28}  {venc}")

print(f"\n  SOLVER:    {placar['SOLVER']} vitórias")
print(f"  HEUR:      {placar['HEUR']} vitórias")
print(f"  EMPATE:    {placar['EMPATE']}")

print(f"\n{'='*100}")
if placar["SOLVER"] > placar["HEUR"]:
    print(f"  🏆 VENCEDOR: SOLVER ({placar['SOLVER']}/{sum(placar.values())} critérios)")
elif placar["HEUR"] > placar["SOLVER"]:
    print(f"  🏆 VENCEDOR: HEURÍSTICA ({placar['HEUR']}/{sum(placar.values())} critérios)")
else:
    print(f"  ⚖️  EMPATE TÉCNICO")
print(f"{'='*100}")

# Diferença de lucro
diff_lucro = ind_s["lucro"] - ind_h["lucro"]
print(f"\n  Δ Lucro (SOLVER − HEUR): R$ {diff_lucro:+,.0f}")
print(f"  Δ NS:                    {ind_s['ns_pct']-ind_h['ns_pct']:+.1f} pp")
print(f"  Δ Buffer PA2 R4:         {ind_s['buffer_pa2_total']-ind_h['buffer_pa2_total']:+,.0f} frascos")
print(f"  Δ Transportes:           {ind_s['n_transp']-ind_h['n_transp']:+d}")
