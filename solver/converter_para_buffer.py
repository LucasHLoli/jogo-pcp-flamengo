"""Converte FLAMENGO horizon agressivo em versão BUFFER puro:
remove linhas CD→Varejo de PA2 em R3 (que seriam entregas pra R4).
Assim PA2 fica estocado nos CDs até R4 começar com OPs reais.
"""
from __future__ import annotations
import io
import re
import shutil
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except AttributeError:
    pass

import openpyxl
from src.io_xlsm import escrever_planos_de_df
import pandas as pd

# Lê FLAMENGO horizon
FLAMENGO_SOLVER = BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO.xlsm"

wb = openpyxl.load_workbook(FLAMENGO_SOLVER, keep_vba=True, data_only=True)
ws = wb["SOL_TRANSP"]

_ROD = re.compile(r"Rodada_(\d+)")
_DIA = re.compile(r"Dia\s*(\d+)")

# Lê todas as linhas R3
linhas_r3 = []
for r in range(5, ws.max_row + 1):
    val = ws.cell(r, 1).value
    if not val: continue
    m = _ROD.search(str(val))
    if not m or int(m.group(1)) != 3: continue
    dia_str = str(ws.cell(r, 4).value)
    dia_raw = int(_DIA.search(dia_str).group(1))
    dia_rel = dia_raw - 10 if dia_raw > 5 else dia_raw
    linhas_r3.append({
        "Rodada": "Rodada_3",
        "Origem": ws.cell(r, 2).value,
        "Cidade": ws.cell(r, 3).value,
        "Dia da Coleta": f"Dia {dia_rel}",
        "Modal": ws.cell(r, 5).value,
        "Tipo do Produto": ws.cell(r, 6).value,
        "Qtde": float(ws.cell(r, 7).value or 0),
        "Destino": ws.cell(r, 8).value,
        "Cidade_Destino": ws.cell(r, 9).value,
    })

print(f"Linhas R3 originais: {len(linhas_r3)}")

# REMOVER: CD→Varejo de PA2 (esses são entregas R4 antecipadas)
linhas_removidas = []
linhas_mantidas = []
for row in linhas_r3:
    if row["Origem"] == "CD" and row["Destino"] == "Varejista" and row["Tipo do Produto"] == "PA2":
        linhas_removidas.append(row)
    else:
        linhas_mantidas.append(row)

print(f"\nLinhas REMOVIDAS (CD→Varejo PA2):")
total_pa2_movido = 0
for r in linhas_removidas:
    qty = r["Qtde"]
    total_pa2_movido += qty
    print(f"  {r['Cidade']} {r['Dia da Coleta']} {r['Modal']} PA2 {qty:.0f} → {r['Cidade_Destino']}")
print(f"\nTotal PA2 que vai FICAR NO CD: {total_pa2_movido:,.0f} frascos")
print(f"Linhas R3 finais: {len(linhas_mantidas)}")

# OP_FABRICAS (mesma — não muda produção)
ws_op = wb["OP_FABRICAS"]
df_op_rows = []
for r in range(7, 12):
    df_op_rows.append({
        "Dia": f"Dia {r - 6}",
        "PA1": int(ws_op.cell(r, 2).value or 0),
        "PA2": int(ws_op.cell(r, 3).value or 0),
        "PA3": int(ws_op.cell(r, 4).value or 0),
    })

df_sol = pd.DataFrame(linhas_mantidas, columns=[
    "Rodada", "Origem", "Cidade", "Dia da Coleta", "Modal",
    "Tipo do Produto", "Qtde", "Destino", "Cidade_Destino"
])
df_op = pd.DataFrame(df_op_rows)

# Salva BUFFER version
OUT = BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO_BUFFER.xlsm"
# Cria cópia do template (rodada_2 sem R3)
shutil.copy(BASE / "rodadas" / "rodada_2" / "FLAMENGO.xlsm", OUT)
n = escrever_planos_de_df(OUT, df_sol, df_op, rodada_n=3)
print(f"\n✅ FLAMENGO_BUFFER.xlsm criado em: {OUT}")
print(f"   {n} linhas R3 (vs {len(linhas_r3)} original)")

# Também atualiza FLAMENGO.xlsm e FLAMENGO_SOLVER.xlsm
shutil.copy(OUT, BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO.xlsm")
shutil.copy(OUT, BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO_SOLVER.xlsm")
escrever_planos_de_df(BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO.xlsm",
                       df_sol, df_op, rodada_n=3)
escrever_planos_de_df(BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO_SOLVER.xlsm",
                       df_sol, df_op, rodada_n=3)
print(f"✅ FLAMENGO.xlsm e FLAMENGO_SOLVER.xlsm também atualizados")

# Mescla histórico R1..R_atual automaticamente nos arquivos finais
from solver.mesclar_historico import mesclar_historico
mesclar_historico(rodada_alvo=3, base=BASE)
