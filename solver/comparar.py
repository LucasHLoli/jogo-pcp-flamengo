"""Comparativo 3-vias: HEURÍSTICA / SOLVER MILP (R3 only) / SOLVER HORIZONTE (R3+R4 forecast).

Gera Comparativo.xlsx em solver/rodadas/rodada_N/ + relatório terminal.
"""
from __future__ import annotations
import argparse
import io
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from src.config import Config

from solver.state import estado_r3_flamengo
from solver.milp_horizon import resolver_horizonte
from solver.milp import resolver_rodada
from solver.forecast_r4 import forecast_ops_r4
from solver.solve import ops_r3


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--rodada", type=int, default=3)
    p.add_argument("--time_limit", type=float, default=120)
    args = p.parse_args()

    cfg = Config.load(BASE)
    estado = estado_r3_flamengo()
    ops3 = ops_r3()
    ops4 = forecast_ops_r4(args.rodada, 0.40)
    precos_r3 = {"PA1": 80, "PA2": 50, "PA3": 32}
    precos_r4 = {"PA1": 80, "PA2": 50, "PA3": 25}

    print("=== RODANDO AS 3 VERSÕES ===\n")

    # --- VERSÃO 1: SOLVER R3 (single) ---
    print("[1/3] Solver MILP single-rodada (R3 apenas)...")
    res_single = resolver_rodada(
        estado=estado, ops=ops3, cfg=cfg,
        pa_proxima_rodada="PA2", buffer_pa_proxima_min=79113,
        preco_pa_rodada=precos_r3,
        ns_min=0.80, objetivo="max_lucro",
        time_limit_s=args.time_limit, verbose=False,
    )

    # --- VERSÃO 2: SOLVER HORIZONTE R3+R4 ---
    print("[2/3] Solver MILP horizonte (R3 + R4 forecast)...")
    res_horizon = resolver_horizonte(
        estado_r3=estado, ops_r3=ops3, ops_r4_forecast=ops4, cfg=cfg,
        precos_r3=precos_r3, precos_r4=precos_r4,
        ns_min=0.80, time_limit_s=args.time_limit, verbose=False,
    )

    # --- VERSÃO 3: HEURÍSTICA ---
    print("[3/3] Heurística (planner_v3)...")
    from src.planner_v3 import planejar_v3
    from src.io_xlsm import ler_instalacoes
    flam = BASE / "rodadas" / f"rodada_{args.rodada}" / "FLAMENGO.xlsm"
    inst = ler_instalacoes(flam)
    res_heur = planejar_v3(
        rodada_n=args.rodada, ops_rodada=ops3,
        estoque_inicial_mp_ton=estado.estoque_mp_ton,
        estoque_inicial_pa_cd=estado.estoque_pa_cd,
        mp_em_transito_chegando=[{"dia": x["dia_rel"], "mp": x["mp"], "qtd": x["qtd"]}
                                  for x in estado.mp_em_transito],
        cfg=cfg, instalacoes=inst,
        pa_proxima_rodada="PA2", buffer_pa_proxima=79113,
        compras_mp_extra_para_r_mais_1={"MP1": 48, "MP3": 48},
    )

    # ==== CALCULAR LUCROS REAIS (custos via custo_total_modal) ====
    from src.planner_v3 import custo_total_modal, km_rota
    from collections import defaultdict
    forn_min = {mp: min(cfg.fornecedores[mp], key=lambda x: x[1]) for mp in ("MP1", "MP2", "MP3")}
    maior_mp = {"MP1": 56000, "MP2": 22000, "MP3": 41000}
    FIX = 1_123_151  # custos fixos por rodada

    def calc_real(df_sol, df_op, est_mp_fim, est_pa_fim, ops_atend, precos):
        """Recalcula custos reais (frete sem linearização)."""
        receita = sum(int(op["qtd"]) * precos[op["pa"]] for op in ops_atend)
        custo_mp = sum(float(row["Qtde"]) * forn_min[row["Tipo do Produto"]][1]
                       for _, row in df_sol.iterrows() if row["Origem"] == "Fornecedor")
        custo_frete = 0
        for _, row in df_sol.iterrows():
            modal = row["Modal"]; item = row["Tipo do Produto"]; qtd = float(row["Qtde"])
            kv = km_rota(cfg, row["Cidade"], row["Cidade_Destino"], modal) or 0
            peso = qtd if item.startswith("MP") else qtd * cfg.peso_un_ton[item]
            custo_frete += custo_total_modal(cfg, modal, kv, peso, 1)
        carreg_mp = sum(est_mp_fim.get(mp, 0) * maior_mp[mp] * 0.01 for mp in ("MP1", "MP2", "MP3"))
        carreg_pa = sum(est_pa_fim.get(cd, {}).get(pa, 0) * precos[pa] * 0.01
                        for cd in ("CD1", "CD2") for pa in ("PA1", "PA2", "PA3"))
        custo_total = FIX + custo_mp + custo_frete + carreg_mp + carreg_pa
        return receita, custo_mp, custo_frete, carreg_mp + carreg_pa, custo_total, receita - custo_total

    # SINGLE
    ops_atend_single = [r for r in res_single.ops_atendidas]
    s_rec, s_mp, s_fr, s_car, s_ct, s_lucro = calc_real(
        res_single.df_sol_transp, res_single.df_op_fabricas,
        res_single.estoque_mp_final, res_single.estoque_pa_cd_final,
        ops_atend_single, precos_r3,
    )

    # HORIZON (só R3)
    h_rec_r3, h_mp_r3, h_fr_r3, h_car_r3, h_ct_r3, h_lucro_r3 = calc_real(
        res_horizon.df_sol_transp_r3, res_horizon.df_op_fabricas_r3,
        res_horizon.estoque_mp_fim_r3, res_horizon.estoque_pa_cd_fim_r3,
        res_horizon.ops_atend_r3, precos_r3,
    )
    h_rec_r4, h_mp_r4, h_fr_r4, h_car_r4, h_ct_r4, h_lucro_r4 = calc_real(
        res_horizon.df_sol_transp_r4, res_horizon.df_op_fabricas_r4,
        res_horizon.estoque_mp_fim_r4, res_horizon.estoque_pa_cd_fim_r4,
        res_horizon.ops_atend_r4, precos_r4,
    )

    # HEURÍSTICA
    he_atend = [r for r in res_heur["resumo"]["rotas_op"] if r.get("alocada")]
    he_atend_dict = [{"pa": r["pa"], "qtd": r["qtd"]} for r in he_atend]
    he_est_pa_fim = {cd: {pa: 0 for pa in ("PA1","PA2","PA3")} for cd in ("CD1", "CD2")}
    he_est_pa_fim["CD2"]["PA2"] = res_heur["resumo"]["buffer_pa_proxima_alocado"]
    he_rec, he_mp, he_fr, he_car, he_ct, he_lucro = calc_real(
        res_heur["df_sol_transp"], res_heur["df_op_fabricas"],
        res_heur["resumo"]["estoque_mp_final_ton"], he_est_pa_fim,
        he_atend_dict, precos_r3,
    )

    # ==== TABELA COMPARATIVA ====
    print(f"\n{'='*100}")
    print(f"  COMPARATIVO TRÊS-VIAS — R3 + projeção R4")
    print(f"{'='*100}\n")
    print(f"{'Indicador':<28}{'HEURÍSTICA':>22}{'SOLVER R3-only':>22}{'SOLVER HORIZONTE':>22}")
    print("-" * 100)

    he_ns = sum(int(r['qtd']) for r in he_atend) / sum(o['qtd'] for o in ops3) * 100
    s_ns = sum(int(r['qtd']) for r in res_single.ops_atendidas) / sum(o['qtd'] for o in ops3) * 100

    print(f"{'NS R3':<28}{f'{he_ns:.1f}%':>22}{f'{s_ns:.1f}%':>22}{f'{res_horizon.ns_r3_pct:.1f}%':>22}")
    print(f"{'NS R4 (forecast)':<28}{'—':>22}{'— (não modela R4)':>22}{f'{res_horizon.ns_r4_pct:.1f}%':>22}")
    print()
    print(f"{'Receita R3':<28}{f'R$ {he_rec:>14,.0f}':>22}{f'R$ {s_rec:>14,.0f}':>22}{f'R$ {h_rec_r3:>14,.0f}':>22}")
    print(f"{'Receita R4 esperada':<28}{'—':>22}{'—':>22}{f'R$ {h_rec_r4:>14,.0f}':>22}")
    print()
    print(f"{'Compra MP R3':<28}{f'R$ {he_mp:>14,.0f}':>22}{f'R$ {s_mp:>14,.0f}':>22}{f'R$ {h_mp_r3:>14,.0f}':>22}")
    print(f"{'Compra MP R4':<28}{'—':>22}{'—':>22}{f'R$ {h_mp_r4:>14,.0f}':>22}")
    print(f"{'Frete R3 (total)':<28}{f'R$ {he_fr:>14,.0f}':>22}{f'R$ {s_fr:>14,.0f}':>22}{f'R$ {h_fr_r3:>14,.0f}':>22}")
    print(f"{'Frete R4 (total)':<28}{'—':>22}{'—':>22}{f'R$ {h_fr_r4:>14,.0f}':>22}")
    print(f"{'Carregamento R3':<28}{f'R$ {he_car:>14,.0f}':>22}{f'R$ {s_car:>14,.0f}':>22}{f'R$ {h_car_r3:>14,.0f}':>22}")
    print(f"{'Carregamento R4':<28}{'—':>22}{'—':>22}{f'R$ {h_car_r4:>14,.0f}':>22}")
    print()
    print(f"{'Custo total R3':<28}{f'R$ {he_ct:>14,.0f}':>22}{f'R$ {s_ct:>14,.0f}':>22}{f'R$ {h_ct_r3:>14,.0f}':>22}")
    print(f"{'Custo total R4':<28}{'—':>22}{'—':>22}{f'R$ {h_ct_r4:>14,.0f}':>22}")
    print()
    print(f"{'LUCRO R3':<28}{f'R$ {he_lucro:>14,.0f}':>22}{f'R$ {s_lucro:>14,.0f}':>22}{f'R$ {h_lucro_r3:>14,.0f}':>22}")
    print(f"{'LUCRO R4 esperado':<28}{'? (sem R4)':>22}{'? (negativo, sem MP)':>22}{f'R$ {h_lucro_r4:>14,.0f}':>22}")
    print(f"{'LUCRO R3+R4 (horizonte)':<28}{'?':>22}{'?':>22}{f'R$ {h_lucro_r3+h_lucro_r4:>14,.0f}':>22}")
    print()
    he_transp = len(res_heur["df_sol_transp"])
    he_mp1 = res_heur["resumo"]["estoque_mp_final_ton"]["MP1"]
    he_mp3 = res_heur["resumo"]["estoque_mp_final_ton"]["MP3"]
    he_buf = res_heur["resumo"]["buffer_pa_proxima_alocado"]
    he_pa2 = res_heur["resumo"]["producao_total_por_pa"]["PA2"]
    s_mp1 = res_single.estoque_mp_final["MP1"]
    s_mp3 = res_single.estoque_mp_final["MP3"]
    s_buf = res_single.estoque_pa_cd_final["CD2"]["PA2"]
    s_pa2 = res_single.df_op_fabricas["PA2"].sum()
    ho_mp1 = res_horizon.estoque_mp_fim_r3["MP1"]
    ho_mp3 = res_horizon.estoque_mp_fim_r3["MP3"]
    ho_buf = res_horizon.estoque_pa_cd_fim_r3["CD2"]["PA2"]
    ho_pa2_r3 = res_horizon.df_op_fabricas_r3["PA2"].sum()
    ho_pa2_r4 = res_horizon.df_op_fabricas_r4["PA2"].sum()

    print(f"{'Transportes R3':<28}{f'{he_transp}/220':>22}{f'{res_single.n_transportes}/220':>22}{f'{res_horizon.n_transp_r3}/220':>22}")
    print(f"{'Transportes R4':<28}{'—':>22}{'—':>22}{f'{res_horizon.n_transp_r4}/220':>22}")
    print()
    print(f"{'Estoque MP1 fim R3':<28}{f'{he_mp1:.1f}t':>22}{f'{s_mp1:.1f}t':>22}{f'{ho_mp1:.1f}t':>22}")
    print(f"{'Estoque MP3 fim R3':<28}{f'{he_mp3:.1f}t':>22}{f'{s_mp3:.1f}t':>22}{f'{ho_mp3:.1f}t':>22}")
    print(f"{'Buffer PA2 R4 (CD2)':<28}{f'{he_buf:,}':>22}{f'{s_buf:,}':>22}{f'{ho_buf:,}':>22}")
    print()
    print(f"{'PA2 produzido R3':<28}{f'{he_pa2:,}':>22}{f'{s_pa2:,}':>22}{f'{ho_pa2_r3:,}':>22}")
    print(f"{'PA2 produzido R4':<28}{'—':>22}{'—':>22}{f'{ho_pa2_r4:,}':>22}")
    print()
    print("ANÁLISE:")
    print(f"  • Heurística e Solver R3-only NÃO modelam R4 → R4 vai precisar de MP nova (lead 3d Manaus = risco).")
    print(f"  • Solver HORIZONTE produz PA2 nos dias livres de R3 (4-5) e envia via Navio chegando R4 dia 1-2.")
    print(f"  • Mesmo o R3 isolado tem lucro PIOR no horizonte (R$ {h_lucro_r3-s_lucro:+,.0f}), mas o R4 garantido vale +R$ {h_lucro_r4:,.0f}.")
    print(f"  • RECOMENDADO: usar SOLVER HORIZONTE — submete R3 com R4 já planejado.")

    # ==== EXCEL ====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparativo_3vias"

    FILL_HDR = PatternFill("solid", fgColor="1F4E78")
    FILL_OK = PatternFill("solid", fgColor="C6EFCE")
    FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
    FILL_WIN = PatternFill("solid", fgColor="FFD966")
    FILL_INFO = PatternFill("solid", fgColor="DDEBF7")
    FONT_HDR = Font(bold=True, color="FFFFFF", size=12)
    FONT_BOLD = Font(bold=True)

    ws.cell(1, 1).value = f"COMPARATIVO 3-VIAS — R{args.rodada} + R{args.rodada+1} (forecast)"
    ws.merge_cells("A1:E1")
    ws.cell(1, 1).font = FONT_HDR; ws.cell(1, 1).fill = FILL_HDR
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    headers = ["Indicador", "HEURÍSTICA", "SOLVER R3-only", "SOLVER HORIZONTE", "Vencedor"]
    for j, h in enumerate(headers):
        c = ws.cell(3, j + 1)
        c.value = h
        c.font = FONT_BOLD; c.fill = FILL_INFO
        c.alignment = Alignment(horizontal="center")

    rows_data = [
        ("NS R3", he_ns, s_ns, res_horizon.ns_r3_pct, "%"),
        ("NS R4 forecast", None, None, res_horizon.ns_r4_pct, "%"),
        ("Receita R3", he_rec, s_rec, h_rec_r3, "$"),
        ("Receita R4 esperada", None, None, h_rec_r4, "$"),
        ("Compra MP R3", he_mp, s_mp, h_mp_r3, "$"),
        ("Compra MP R4", None, None, h_mp_r4, "$"),
        ("Frete R3", he_fr, s_fr, h_fr_r3, "$"),
        ("Frete R4", None, None, h_fr_r4, "$"),
        ("Carreg R3", he_car, s_car, h_car_r3, "$"),
        ("Carreg R4", None, None, h_car_r4, "$"),
        ("Custo total R3", he_ct, s_ct, h_ct_r3, "$"),
        ("Custo total R4", None, None, h_ct_r4, "$"),
        ("LUCRO R3", he_lucro, s_lucro, h_lucro_r3, "$"),
        ("LUCRO R4 esperado", None, None, h_lucro_r4, "$"),
        ("LUCRO R3+R4 (horizonte)", None, None, h_lucro_r3 + h_lucro_r4, "$"),
        ("Transportes R3", len(res_heur["df_sol_transp"]), res_single.n_transportes,
         res_horizon.n_transp_r3, "#"),
        ("Transportes R4", None, None, res_horizon.n_transp_r4, "#"),
        ("MP1 fim R3 (ton)", res_heur["resumo"]["estoque_mp_final_ton"]["MP1"],
         res_single.estoque_mp_final["MP1"], res_horizon.estoque_mp_fim_r3["MP1"], "t"),
        ("MP3 fim R3 (ton)", res_heur["resumo"]["estoque_mp_final_ton"]["MP3"],
         res_single.estoque_mp_final["MP3"], res_horizon.estoque_mp_fim_r3["MP3"], "t"),
        ("PA2 produzido R3 (un)", res_heur["resumo"]["producao_total_por_pa"]["PA2"],
         res_single.df_op_fabricas["PA2"].sum(), res_horizon.df_op_fabricas_r3["PA2"].sum(), "un"),
    ]

    r = 4
    for label, h, s, ho, tipo in rows_data:
        ws.cell(r, 1).value = label
        ws.cell(r, 2).value = h
        ws.cell(r, 3).value = s
        ws.cell(r, 4).value = ho
        if tipo == "$":
            for c in (2, 3, 4):
                if isinstance(ws.cell(r, c).value, (int, float)):
                    ws.cell(r, c).number_format = 'R$ #,##0'
        elif tipo == "%":
            for c in (2, 3, 4):
                if isinstance(ws.cell(r, c).value, (int, float)):
                    ws.cell(r, c).number_format = '0.0"%"'
        # Vencedor
        vals = [(h, "H"), (s, "S"), (ho, "HO")]
        vals = [v for v in vals if isinstance(v[0], (int, float))]
        if vals and "LUCRO" in label or "Receita" in label or "NS" in label:
            best = max(vals, key=lambda x: x[0])
            ws.cell(r, 5).value = best[1]
            ws.cell(r, 5).fill = FILL_OK
        elif vals and ("Custo" in label or "Compra" in label or "Frete" in label or "Carreg" in label):
            best = min(vals, key=lambda x: x[0])
            ws.cell(r, 5).value = best[1]
            ws.cell(r, 5).fill = FILL_OK
        # Destacar linhas de lucro
        if "LUCRO" in label:
            for c in range(1, 6):
                ws.cell(r, c).font = FONT_BOLD
        r += 1

    # Legenda
    r += 2
    ws.cell(r, 1).value = "LEGENDA: H=Heurística | S=Solver R3-only | HO=Solver Horizonte (R3+R4)"
    ws.cell(r, 1).font = FONT_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    notas = [
        "Heurística: src/planner_v3.py — gulosa, sem visão de R4.",
        "Solver R3-only: solver/milp.py — ótima R3, mas deixa estoque vazio → R4 difícil.",
        "Solver Horizonte: solver/milp_horizon.py — modela R3+R4 conjuntamente com forecast HW PA2.",
        "Lucro R4 da heurística e solver R3-only é DESCONHECIDO (não modelado).",
        "Solver Horizonte produz PA2 nos dias livres de R3 e envia Navio chegando R4 dia 1-2.",
    ]
    for n in notas:
        ws.cell(r, 1).value = n
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        r += 1

    # auto width
    for col in ws.columns:
        try:
            l = col[0].column_letter
        except AttributeError:
            continue
        m = 12
        for c in col:
            if c.value is not None:
                m = max(m, min(60, len(str(c.value)) + 2))
        ws.column_dimensions[l].width = m

    OUT = BASE / "solver" / "rodadas" / f"rodada_{args.rodada}" / "Comparativo.xlsx"
    wb.save(OUT)
    print(f"\nComparativo.xlsx salvo em: {OUT}")


if __name__ == "__main__":
    main()
