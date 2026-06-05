"""Mescla todas as rodadas anteriores + a rodada atual num único FLAMENGO_SOLVER.xlsm.

Uso:
    python solver/mesclar_historico.py --rodada 3   # produz R1+R2+R3 mesclados
    python solver/mesclar_historico.py --rodada 4   # produz R1+R2+R3+R4 mesclados

Também pode ser chamado por outros scripts via `mesclar_historico(rodada=N)`.

Para cada rodada r ∈ {1..N}, procura as linhas em fontes na ordem de preferência:
  1) solver/rodadas/rodada_{r}/FLAMENGO_BUFFER.xlsm        (nossa saída solver pura)
  2) solver/rodadas/rodada_{r}/FLAMENGO_SOLVER.xlsm        (nossa saída solver mesclada anterior)
  3) rodadas/rodada_{r+1}/FLAMENGO.xlsm                    (template do jogo p/ rodada seguinte = histórico de r)
  4) solver/rodadas/rodada_{r+1}/FLAMENGO.xlsm             (template solver da rodada seguinte)
  5) rodadas/rodada_{r}/FLAMENGO.xlsm                      (template do jogo da própria rodada r — pode ter histórico anterior)

Para a rodada N (alvo), pega só de fontes 1) ou 2) (saída do solver atual).
Saída sempre vai para:
  solver/rodadas/rodada_N/FLAMENGO_SOLVER.xlsm
  solver/rodadas/rodada_N/FLAMENGO.xlsm
"""
from __future__ import annotations
import argparse
import io
import re
import shutil
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

_ROD = re.compile(r"Rodada_(\d+)")


def _ler_rodada(xlsm: Path, rodada: int) -> list[list]:
    """Lê linhas SOL_TRANSP de uma rodada específica (9 colunas)."""
    try:
        wb = openpyxl.load_workbook(xlsm, keep_vba=True, data_only=True)
    except Exception:
        return []
    if "SOL_TRANSP" not in wb.sheetnames:
        return []
    ws = wb["SOL_TRANSP"]
    linhas = []
    for r in range(5, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        m = _ROD.search(str(v))
        if not m or int(m.group(1)) != rodada:
            continue
        linhas.append([ws.cell(r, c).value for c in range(1, 10)])
    return linhas


def _fontes_para(rodada: int, alvo: int, base: Path) -> list[Path]:
    """Lista de fontes possíveis, ordenada por preferência."""
    if rodada == alvo:
        # Para a rodada alvo, pega exclusivamente das saídas do solver
        return [
            base / "solver" / "rodadas" / f"rodada_{rodada}" / "FLAMENGO_BUFFER.xlsm",
            base / "solver" / "rodadas" / f"rodada_{rodada}" / "FLAMENGO_SOLVER.xlsm",
            base / "solver" / "rodadas" / f"rodada_{rodada}" / "FLAMENGO.xlsm",
        ]
    # Para rodadas históricas: prioriza output solver, depois templates do jogo
    return [
        base / "solver" / "rodadas" / f"rodada_{rodada}" / "FLAMENGO_BUFFER.xlsm",
        base / "solver" / "rodadas" / f"rodada_{rodada}" / "FLAMENGO_SOLVER.xlsm",
        base / "rodadas" / f"rodada_{rodada + 1}" / "FLAMENGO.xlsm",
        base / "solver" / "rodadas" / f"rodada_{rodada + 1}" / "FLAMENGO.xlsm",
        base / "rodadas" / f"rodada_{rodada}" / "FLAMENGO.xlsm",
    ]


def coletar(rodada: int, alvo: int, base: Path) -> tuple[list[list], Path | None]:
    """Coleta linhas de uma rodada, tentando fontes na ordem de preferência."""
    for fonte in _fontes_para(rodada, alvo, base):
        if not fonte.exists():
            continue
        linhas = _ler_rodada(fonte, rodada)
        if linhas:
            return linhas, fonte
    return [], None


def mesclar_historico(rodada_alvo: int, base: Path = BASE, verbose: bool = True) -> Path:
    """Mescla R1..R_{rodada_alvo} num FLAMENGO_SOLVER.xlsm e retorna o path final."""
    pasta_alvo = base / "solver" / "rodadas" / f"rodada_{rodada_alvo}"
    template = pasta_alvo / "FLAMENGO_BUFFER.xlsm"
    if not template.exists():
        # fallback: pega o FLAMENGO.xlsm da pasta alvo
        template = pasta_alvo / "FLAMENGO.xlsm"
    if not template.exists():
        raise FileNotFoundError(f"Não encontrei template em {pasta_alvo}")

    out_solver = pasta_alvo / "FLAMENGO_SOLVER.xlsm"
    out_flamengo = pasta_alvo / "FLAMENGO.xlsm"

    if verbose:
        print(f"\n[mesclar_historico] Rodada alvo: R{rodada_alvo}")
        print(f"  Template base: {template.name}")

    # Coleta linhas de cada rodada 1..N
    todas = {}
    for r in range(1, rodada_alvo + 1):
        linhas, fonte = coletar(r, rodada_alvo, base)
        todas[r] = linhas
        if verbose:
            fname = fonte.relative_to(base) if fonte else "—"
            print(f"  R{r}: {len(linhas):>4} linhas  ← {fname}")

    total = sum(len(v) for v in todas.values())
    if verbose:
        print(f"  TOTAL: {total} linhas\n")

    # Copia template e regrava SOL_TRANSP completo
    shutil.copy(template, out_solver)
    wb = openpyxl.load_workbook(out_solver, keep_vba=True)
    ws = wb["SOL_TRANSP"]

    for r in range(5, ws.max_row + 1):
        for c in range(1, 10):
            ws.cell(r, c).value = None

    r = 5
    for rodada in range(1, rodada_alvo + 1):
        for linha in todas[rodada]:
            for c, val in enumerate(linha, start=1):
                ws.cell(r, c).value = val
            r += 1

    wb.save(out_solver)
    shutil.copy(out_solver, out_flamengo)

    if verbose:
        print(f"  ✅ FLAMENGO_SOLVER.xlsm  → {out_solver}")
        print(f"  ✅ FLAMENGO.xlsm        → {out_flamengo}")

    return out_solver


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--rodada", type=int, required=True, help="Rodada alvo (1..N)")
    args = p.parse_args()
    mesclar_historico(args.rodada)


if __name__ == "__main__":
    main()
