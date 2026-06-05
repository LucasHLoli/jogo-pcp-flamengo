"""Valida o R3 do SOLVER contra TODAS as regras do jogo.

Lê solver/rodadas/rodada_3/FLAMENGO.xlsm e simula tudo dia a dia.
Reporta qualquer violação encontrada.

Regras verificadas:
  R1. PA chega no DIA EXATO solicitado (antes/depois = descartado)
  R2. PA sai da F1 no MESMO dia que é produzido
  R3. MP chegando sem espaço = descartada (cap MP)
  R4. Total transportes R3 ≤ 220
  R5. Capacidade da fábrica respeitada (10.080 min/dia)
  R6. Capacidade modal respeitada (cap/viagem)
  R7. PA no CD nunca negativo (CD tem o PA que despacha)
  R8. CD ≤ cap (sem descarte de PA no CD)
  R9. MP em F1 nunca negativo (sempre há MP pra produzir)
  R10. MP comprada chega no fornecedor mais barato
  R11. Production matches CD shipments (Σ prod = Σ envio F1→CD)
  R12. Stock MP F1 incluindo em-trânsito de R2 está coerente
"""
from __future__ import annotations
import io
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from src.config import Config
from src.io_xlsm import ler_instalacoes
from src.planner_v3 import lead_dias, km_rota, custo_total_modal

cfg = Config.load(BASE)
lead_tab = json.loads((BASE / "data" / "lead_times.json").read_text(encoding="utf-8"))

# OPs R3 oficiais
OPS = [{"cidade": c, "pa": "PA3", "qtd": q, "dia_entrega": d} for c, q, d in [
    ("Belém", 20155, 5), ("Belo Horizonte", 70544, 3), ("Brasília", 117573, 3),
    ("Campinas", 56435, 2), ("Campo Grande", 23515, 3), ("Cuiabá", 28218, 3),
    ("Curitiba", 103464, 2), ("Fortaleza", 68528, 5), ("Goiânia", 65841, 3),
    ("João Pessoa", 40311, 4), ("Joinville", 28218, 2), ("Maceió", 40311, 4),
    ("Manaus", 20155, 5), ("Natal", 40311, 5), ("Porto Alegre", 103464, 2),
    ("Recife", 60466, 4), ("Ribeirão Preto", 47029, 2), ("Rio de Janeiro", 94059, 3),
    ("Salvador", 80622, 4), ("Santos", 47029, 2), ("São Luís", 20155, 5),
    ("São Paulo", 117573, 2), ("Uberlândia", 23515, 3), ("Vitória", 14109, 3),
    ("Vitória da Conquista", 12093, 4),
]]

# Estado inicial R3 (do PDF Estoques R3 = fim de R2)
ESTOQUE_MP_INI = {"MP1": 78.98, "MP2": 50.36, "MP3": 48.14}
EM_TRANSITO = [{"dia": 1, "mp": "MP1", "qtd": 8.7}]  # MP1 Manaus shipped R2 Dia 8

# Forn mais barato
FORN_MIN = {mp: min(cfg.fornecedores[mp], key=lambda x: x[1]) for mp in ("MP1", "MP2", "MP3")}

# Cap
F1 = ler_instalacoes(BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO.xlsm")["fabricas"]["F1"]
CDS = ler_instalacoes(BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO.xlsm")["cds"]
cap_mp = {mp: F1["area_mp"][mp] * 2 * cfg.densidades_mp[mp] for mp in ("MP1", "MP2", "MP3")}
cap_pa_cd = {cd: {pa: int(d["area_pa"][pa] * 2 * cfg.densidades_pa[pa] / cfg.peso_un_ton[pa])
                  for pa in ("PA1", "PA2", "PA3")} for cd, d in CDS.items()}
CAP_MIN_DIA = F1["maquinas"] * F1["turnos"] * 8 * 60
fab_cidade = F1["cidade"]
cidade_to_cd = {d["cidade"]: cd for cd, d in CDS.items()}
VEL = {"PA1": 15, "PA2": 30, "PA3": 60}


# ===== LER FLAMENGO DO SOLVER =====
FLAMENGO = BASE / "solver" / "rodadas" / "rodada_3" / "FLAMENGO.xlsm"
print(f"Lendo {FLAMENGO}...\n")
wb = openpyxl.load_workbook(FLAMENGO, keep_vba=True, data_only=True)
ws_t = wb["SOL_TRANSP"]
ws_o = wb["OP_FABRICAS"]

_DIA = re.compile(r"Dia\s*(\d+)")
_ROD = re.compile(r"Rodada_(\d+)")

# Linhas R3
sol_r3 = []
for r in range(5, ws_t.max_row + 1):
    val = ws_t.cell(r, 1).value
    if not val: continue
    m = _ROD.search(str(val))
    if not m or int(m.group(1)) != 3: continue
    dia_str = str(ws_t.cell(r, 4).value)
    m2 = _DIA.search(dia_str)
    if not m2: continue
    dia_raw = int(m2.group(1))
    dia_rel = dia_raw - 10 if dia_raw > 5 else dia_raw  # R3 = abs 11..15
    if not (1 <= dia_rel <= 5): continue
    sol_r3.append({
        "Origem": ws_t.cell(r, 2).value,
        "Cidade": ws_t.cell(r, 3).value,
        "dia_rel": dia_rel,
        "Modal": ws_t.cell(r, 5).value,
        "Tipo": ws_t.cell(r, 6).value,
        "Qtde": float(ws_t.cell(r, 7).value or 0),
        "Destino": ws_t.cell(r, 8).value,
        "Cidade_Destino": ws_t.cell(r, 9).value,
    })

# OP_FABRICAS R3
prod_r3 = {}
for d in range(1, 6):
    prod_r3[d] = {
        "PA1": int(ws_o.cell(6 + d, 2).value or 0),
        "PA2": int(ws_o.cell(6 + d, 3).value or 0),
        "PA3": int(ws_o.cell(6 + d, 4).value or 0),
    }


# ============ EXECUÇÃO DA VALIDAÇÃO ============
problemas = []
sucessos = []


def check(nome, ok, detail=""):
    if ok:
        sucessos.append((nome, detail))
        print(f"  ✅ {nome}: {detail}")
    else:
        problemas.append((nome, detail))
        print(f"  ❌ {nome}: {detail}")


print("=" * 80)
print(f"  VALIDANDO SOLVER R3 — {len(sol_r3)} linhas + 5 dias OP_FABRICAS")
print("=" * 80)

# ===== R4: TRANSPORTES ≤ 220 =====
print(f"\n[R4] Total de transportes ≤ 220:")
check("Total transp R3", len(sol_r3) <= 220, f"{len(sol_r3)} / 220")

# ===== R5: CAP FÁBRICA =====
print(f"\n[R5] Cap fábrica ≤ 10.080 min/dia:")
total_min_falhas = 0
for d in range(1, 6):
    min_usado = sum(prod_r3[d][pa] / VEL[pa] for pa in ("PA1", "PA2", "PA3"))
    ok = min_usado <= CAP_MIN_DIA + 1
    print(f"  Dia {d}: {min_usado:.0f} min usados / {CAP_MIN_DIA} cap ({min_usado/CAP_MIN_DIA*100:.1f}%) {'✅' if ok else '❌'}")
    if not ok: total_min_falhas += 1
check("Cap min/dia respeitada (todos os dias)", total_min_falhas == 0, f"{total_min_falhas} dias violados")

# ===== R6: CAP MODAL POR VIAGEM =====
print(f"\n[R6] Cap modal por viagem respeitada:")
cap_modal_ton = {"Avião": 1, "Caminhão": 24, "Navio": 100}
cap_modal_un_pa = {"Avião": {"PA1": 3333, "PA2": 4000, "PA3": 6666},
                    "Caminhão": {"PA1": 80000, "PA2": 96000, "PA3": 160000},
                    "Navio": {"PA1": 333333, "PA2": 400000, "PA3": 666666}}
violacoes_modal = 0
for row in sol_r3:
    modal = row["Modal"]; item = row["Tipo"]; qtd = row["Qtde"]
    if item in ("PA1", "PA2", "PA3"):
        cap = cap_modal_un_pa[modal][item]
        if qtd > cap + 0.01:
            print(f"  ❌ Viagem {modal} {item} qty={qtd:.0f} > cap {cap}")
            violacoes_modal += 1
    elif item.startswith("MP"):
        cap = cap_modal_ton[modal]
        if qtd > cap + 0.01:
            print(f"  ❌ Viagem {modal} {item} qty={qtd:.2f}t > cap {cap}t")
            violacoes_modal += 1
check("Cap modal por viagem", violacoes_modal == 0, f"{violacoes_modal} violações")

# ===== R2: PA SAI DA F1 NO MESMO DIA DA PRODUÇÃO =====
print(f"\n[R2] PA sai F1 no mesmo dia que é produzido:")
prod_vs_envio = []
for d in range(1, 6):
    for pa in ("PA1", "PA2", "PA3"):
        prod = prod_r3[d][pa]
        enviado = sum(row["Qtde"] for row in sol_r3
                      if row["dia_rel"] == d and row["Origem"] == "Fábrica"
                      and row["Destino"] == "CD" and row["Tipo"] == pa)
        diff = prod - enviado
        if prod > 0 or enviado > 0:
            ok = abs(diff) <= max(prod * 0.01, 5)
            print(f"  Dia {d} {pa}: produzido {prod:>7,} | enviado F1→CD {enviado:>7,.0f} | diff {diff:+,.0f} {'✅' if ok else '❌'}")
            prod_vs_envio.append(ok)
check("Σ produção = Σ envio F1→CD por dia", all(prod_vs_envio),
      f"{sum(1 for x in prod_vs_envio if not x)} discrepâncias")

# ===== R1: PA CHEGA NO DIA EXATO =====
print(f"\n[R1] Cada OP chega no varejo no DIA EXATO:")
cdv_chegando_no_dia = defaultdict(float)  # (cidade, pa, dia_chegada) → qty
for row in sol_r3:
    if row["Origem"] == "CD" and row["Destino"] == "Varejista":
        lt = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
        dia_cheg = row["dia_rel"] + lt
        cdv_chegando_no_dia[(row["Cidade_Destino"], row["Tipo"], dia_cheg)] += row["Qtde"]
n_fora_dia = 0
for op in OPS:
    qty_no_dia = cdv_chegando_no_dia.get((op["cidade"], op["pa"], op["dia_entrega"]), 0)
    if abs(qty_no_dia - op["qtd"]) > 1:
        print(f"  ❌ {op['cidade']:<22} {op['pa']} dia={op['dia_entrega']}: planejado {op['qtd']:,} entregue no dia {qty_no_dia:,.0f}")
        n_fora_dia += 1
check("Todas as 25 OPs chegam no dia exato", n_fora_dia == 0, f"{n_fora_dia} fora do dia")

# ===== R9, R3: MP em F1 dia a dia =====
print(f"\n[R3, R9] Estoque MP F1 dia a dia:")
arrivals_mp = defaultdict(lambda: defaultdict(float))
for it in EM_TRANSITO:
    arrivals_mp[it["dia"]][it["mp"]] += it["qtd"]
for row in sol_r3:
    if row["Origem"] == "Fornecedor":
        lt = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
        dia_cheg = row["dia_rel"] + lt
        if 1 <= dia_cheg <= 5:
            arrivals_mp[dia_cheg][row["Tipo"]] += row["Qtde"]
stk = dict(ESTOQUE_MP_INI)
neg_mp = 0
desc_mp = 0
print(f"  {'Dia':<4}{'MP':<5}{'Pre':>9}{'Arr':>8}{'Pos':>9}{'Desc':>7}{'Cons':>8}{'End':>9}")
for d in range(1, 6):
    for mp in ("MP1", "MP2", "MP3"):
        pre = stk[mp]
        arr = arrivals_mp[d].get(mp, 0)
        pos = pre + arr
        desc = max(0, pos - cap_mp[mp])
        pos = min(pos, cap_mp[mp])
        cons = sum(prod_r3[d][pa] * cfg.BoM[pa][mp] / 1e6 for pa in ("PA1", "PA2", "PA3"))
        end = pos - cons
        flag = ""
        if end < -0.01:
            flag = "❌ NEG"; neg_mp += 1
        if desc > 0.01:
            flag += f" ⚠️ DESC {desc:.2f}t"; desc_mp += 1
        print(f"  D{d}  {mp:<5}{pre:>9.2f}{arr:>8.2f}{pos:>9.2f}{desc:>7.2f}{cons:>8.2f}{end:>9.2f} {flag}")
        stk[mp] = max(0, end)
check("MP F1 nunca negativo", neg_mp == 0, f"{neg_mp} eventos")
check("MP F1 nunca excede cap (descarte = 0)", desc_mp == 0, f"{desc_mp} eventos descarte")

# ===== R7, R8: PA nos CDs dia a dia =====
print(f"\n[R7, R8] Estoque PA CDs dia a dia:")
arrivals_pa_cd = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
saidas_cd = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
for row in sol_r3:
    if row["Origem"] == "Fábrica" and row["Destino"] == "CD":
        cd = cidade_to_cd[row["Cidade_Destino"]]
        lt = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
        dia_cheg = row["dia_rel"] + lt
        if 1 <= dia_cheg <= 5:
            arrivals_pa_cd[dia_cheg][cd][row["Tipo"]] += row["Qtde"]
    elif row["Origem"] == "CD":
        cd = cidade_to_cd[row["Cidade"]]
        saidas_cd[row["dia_rel"]][cd][row["Tipo"]] += row["Qtde"]
stk_pa = {cd: {pa: 0 for pa in ("PA1", "PA2", "PA3")} for cd in CDS}
neg_pa = 0
over_pa = 0
print(f"  {'Dia':<4}{'CD':<5}{'PA':<5}{'Pre':>8}{'Arr':>10}{'Sai':>10}{'End':>10}{'Cap':>10}")
for d in range(1, 6):
    for cd in ("CD1", "CD2"):
        for pa in ("PA1", "PA2", "PA3"):
            pre = stk_pa[cd][pa]
            arr = arrivals_pa_cd[d][cd].get(pa, 0)
            sai = saidas_cd[d][cd].get(pa, 0)
            end = pre + arr - sai
            flag = ""
            if end < -0.5:
                flag = "❌ NEG"; neg_pa += 1
            if end > cap_pa_cd[cd][pa]:
                flag += f" ❌ >CAP"; over_pa += 1
            if pre + arr + sai != 0:
                print(f"  D{d}  {cd:<5}{pa:<5}{pre:>8.0f}{arr:>10.0f}{sai:>10.0f}{end:>10.0f}{cap_pa_cd[cd][pa]:>10} {flag}")
            stk_pa[cd][pa] = max(0, end)
check("PA CD nunca negativo (CD tem o que despacha)", neg_pa == 0, f"{neg_pa} eventos")
check("PA CD nunca excede cap CD", over_pa == 0, f"{over_pa} eventos")

# ===== R10: MP comprado do fornecedor mais barato =====
print(f"\n[R10] MP comprada do fornecedor mais barato:")
n_fornec_caro = 0
for row in sol_r3:
    if row["Origem"] == "Fornecedor":
        forn_real = row["Cidade"]
        mp = row["Tipo"]
        forn_certo = FORN_MIN[mp][0]
        if forn_real != forn_certo:
            print(f"  ❌ {mp} comprado de {forn_real} (esperado: {forn_certo})")
            n_fornec_caro += 1
check("MP do fornecedor mais barato", n_fornec_caro == 0, f"{n_fornec_caro} violações")

# ===== NS =====
print(f"\n[NS] Nível de serviço:")
qty_total = sum(o["qtd"] for o in OPS)
qty_atend = sum(o["qtd"] for o in OPS
                if abs(cdv_chegando_no_dia.get((o["cidade"], o["pa"], o["dia_entrega"]), 0) - o["qtd"]) < 1)
ns_pct = qty_atend / qty_total * 100
print(f"  NS = {qty_atend:,} / {qty_total:,} = {ns_pct:.1f}%")
check("NS ≥ 80%", ns_pct >= 80, f"{ns_pct:.1f}%")
check("NS = 100% (ideal)", ns_pct == 100, f"{ns_pct:.1f}%")

# ===== RESUMO FINAL =====
print(f"\n{'=' * 80}")
print(f"  RESUMO DA VALIDAÇÃO")
print(f"{'=' * 80}")
print(f"  ✅ Aprovados:  {len(sucessos)}")
print(f"  ❌ Falhas:     {len(problemas)}")
print()
if problemas:
    print("PROBLEMAS:")
    for n, d in problemas:
        print(f"  ❌ {n}: {d}")
else:
    print("🎉 PLANO 100% VALIDADO!")

print()
print(f"  Frascos entregues no dia exato: {qty_atend:,} / {qty_total:,}")
print(f"  Buffer PA2 R4 (CD2): {stk_pa['CD2']['PA2']:,} frascos")
print(f"  Buffer PA2 R4 (CD1): {stk_pa['CD1']['PA2']:,} frascos")
print(f"  Estoque MP final F1: MP1={stk['MP1']:.2f}t  MP2={stk['MP2']:.2f}t  MP3={stk['MP3']:.2f}t")
