"""SanityCheck do SOLVER (separado da heurística).

Lê:
  solver/rodadas/rodada_N/FLAMENGO.xlsm  ← gerado pelo solver
Gera:
  solver/rodadas/rodada_N/SanityCheck.xlsm

Mesma lógica do scripts/gera_sanity_check.py mas SEM re-rodar planner —
usa a SOL_TRANSP e OP_FABRICAS efetivamente escritas no Excel do solver.
"""
from __future__ import annotations
import argparse
import io
import json
import math
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from src.config import Config
from src.io_xlsm import ler_instalacoes
from src.planner_v3 import lead_dias, km_rota, custo_total_modal

from solver.state import estado_r3_flamengo

FILL_HDR = PatternFill("solid", fgColor="1F4E78")
FILL_SUB = PatternFill("solid", fgColor="9BC2E6")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
FILL_INFO = PatternFill("solid", fgColor="DDEBF7")
FONT_HDR = Font(bold=True, color="FFFFFF", size=12)
FONT_SUB = Font(bold=True, size=11)
FONT_BOLD = Font(bold=True)

_DIA = re.compile(r"Dia\s*(\d+)")
_ROD = re.compile(r"Rodada_(\d+)")


def title(ws, row, ncols, text):
    ws.cell(row, 1).value = text
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row, 1).font = FONT_HDR
    ws.cell(row, 1).fill = FILL_HDR
    ws.cell(row, 1).alignment = Alignment(horizontal="center")


def header(ws, row, headers, fill=FILL_INFO):
    for j, h in enumerate(headers):
        c = ws.cell(row, j + 1)
        c.value = h
        c.font = FONT_BOLD
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def money(c):
    c.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def pct(c):
    c.number_format = "0.0%"


def auto_width(ws, mx=50):
    for col in ws.columns:
        try:
            l = col[0].column_letter
        except AttributeError:
            continue
        m = 10
        for c in col:
            if c.value is not None:
                m = max(m, min(mx, len(str(c.value)) + 2))
        ws.column_dimensions[l].width = m


# OPs R3
OPS_R3 = [{"cidade": c, "pa": "PA3", "qtd": q, "dia_entrega": d} for c, q, d in [
    ("Belém", 20155, 5), ("Belo Horizonte", 70544, 3), ("Brasília", 117573, 3),
    ("Campinas", 56435, 2), ("Campo Grande", 23515, 3), ("Cuiabá", 28218, 3),
    ("Curitiba", 103464, 2), ("Fortaleza", 68528, 5), ("Goiânia", 65841, 3),
    ("João Pessoa", 40311, 4), ("Joinville", 28218, 2), ("Maceió", 40311, 4),
    ("Manaus", 20155, 5), ("Natal", 40311, 5), ("Porto Alegre", 103464, 2),
    ("Recife", 60466, 4), ("Ribeirão Preto", 47029, 2), ("Rio de Janeiro", 94059, 3),
    ("Salvador", 80622, 4), ("Santos", 47029, 2), ("São Luís", 20155, 5),
    ("São Paulo", 117573, 2), ("Uberlândia", 23515, 3), ("Vitória", 14109, 3),
    ("Vitória da Conquista", 12093, 4),
]]


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--rodada", type=int, default=3)
    args = p.parse_args()

    cfg = Config.load(BASE)
    estado = estado_r3_flamengo()

    FLAMENGO = BASE / "solver" / "rodadas" / f"rodada_{args.rodada}" / "FLAMENGO.xlsm"
    if not FLAMENGO.exists():
        print(f"ERRO: {FLAMENGO} não encontrado. Rode 'python solver/solve.py --rodada {args.rodada}' antes.")
        return

    print(f"Lendo SOL_TRANSP de {FLAMENGO}...")
    wb_src = openpyxl.load_workbook(FLAMENGO, keep_vba=True, data_only=True)
    ws_t = wb_src["SOL_TRANSP"]
    ws_o = wb_src["OP_FABRICAS"]
    inst = ler_instalacoes(FLAMENGO)
    F1 = inst["fabricas"]["F1"]
    CDS = inst["cds"]
    fab_cidade = F1["cidade"]
    cds_info = {cd: d["cidade"] for cd, d in CDS.items()}
    cidade_to_cd = {v: k for k, v in cds_info.items()}

    # Parse SOL_TRANSP R3
    sol_rows = []
    for r in range(5, ws_t.max_row + 1):
        v = ws_t.cell(r, 1).value
        if not v or _ROD.search(str(v)) is None: continue
        if int(_ROD.search(str(v)).group(1)) != args.rodada: continue
        dia_str = str(ws_t.cell(r, 4).value)
        m = _DIA.search(dia_str)
        if not m: continue
        dia_raw = int(m.group(1))
        # Detecta absoluto (>5) ou relativo
        dia_rel = dia_raw - (args.rodada - 1) * 5 if dia_raw > 5 else dia_raw
        sol_rows.append({
            "Rodada": v,
            "Origem": ws_t.cell(r, 2).value,
            "Cidade": ws_t.cell(r, 3).value,
            "Dia da Coleta": f"Dia {dia_rel}",
            "Modal": ws_t.cell(r, 5).value,
            "Tipo do Produto": ws_t.cell(r, 6).value,
            "Qtde": float(ws_t.cell(r, 7).value or 0),
            "Destino": ws_t.cell(r, 8).value,
            "Cidade_Destino": ws_t.cell(r, 9).value,
        })
    df_sol = pd.DataFrame(sol_rows)

    # Parse OP_FABRICAS
    df_op_rows = []
    for r in range(7, 12):
        df_op_rows.append({
            "Dia": f"Dia {r - 6}",
            "PA1": int(ws_o.cell(r, 2).value or 0),
            "PA2": int(ws_o.cell(r, 3).value or 0),
            "PA3": int(ws_o.cell(r, 4).value or 0),
        })
    df_op = pd.DataFrame(df_op_rows)

    print(f"  SOL_TRANSP R{args.rodada}: {len(df_sol)} linhas")
    print(f"  OP_FABRICAS: PA3 total = {df_op['PA3'].sum()}, PA2 total = {df_op['PA2'].sum()}")

    # --- SIMULAÇÃO DIA A DIA ---
    cap_mp = {mp: F1["area_mp"][mp] * 2 * cfg.densidades_mp[mp] for mp in ("MP1", "MP2", "MP3")}
    cap_pa_cd = {cd: {pa: int(d["area_pa"][pa] * 2 * cfg.densidades_pa[pa] / cfg.peso_un_ton[pa])
                      for pa in ("PA1", "PA2", "PA3")}
                 for cd, d in CDS.items()}
    cap_min_dia = F1["maquinas"] * F1["turnos"] * 8 * 60

    arrivals_mp = defaultdict(lambda: defaultdict(float))
    arrivals_pa_cd = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    saidas_cd = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    # MP em-trânsito (das rodadas anteriores)
    for x in estado.mp_em_transito:
        arrivals_mp[x["dia_rel"]][x["mp"]] += x["qtd"]

    for _, row in df_sol.iterrows():
        dia_part = int(_DIA.search(str(row["Dia da Coleta"])).group(1))
        modal = row["Modal"]; item = row["Tipo do Produto"]; qtd = float(row["Qtde"])
        o, d = row["Cidade"], row["Cidade_Destino"]
        lt = lead_dias(cfg, o, d, modal) or 0
        dia_cheg = dia_part + lt
        if row["Origem"] == "Fornecedor" and 1 <= dia_cheg <= 5:
            arrivals_mp[dia_cheg][item] += qtd
        elif row["Origem"] == "Fábrica" and row["Destino"] == "CD":
            cd_id = cidade_to_cd[d]
            if 1 <= dia_cheg <= 5:
                arrivals_pa_cd[dia_cheg][cd_id][item] += qtd
        elif row["Origem"] == "CD" and row["Destino"] == "Varejista":
            cd_id = cidade_to_cd[o]
            saidas_cd[dia_part][cd_id][item] += qtd

    # Simulação
    sim_mp = {}
    sim_pa = {cd: {} for cd in CDS}
    stk_mp = dict(estado.estoque_mp_ton)
    stk_pa = {cd: dict(estado.estoque_pa_cd[cd]) for cd in CDS}

    for dia in range(1, 6):
        op_row = df_op[df_op["Dia"] == f"Dia {dia}"].iloc[0]
        prod = {pa: int(op_row[pa]) for pa in ("PA1", "PA2", "PA3")}
        sim_mp[dia] = {}
        for mp in ("MP1", "MP2", "MP3"):
            pre = stk_mp[mp]
            arr = arrivals_mp[dia].get(mp, 0)
            pos = pre + arr
            desc = max(0, pos - cap_mp[mp])
            pos = min(pos, cap_mp[mp])
            cons = sum(prod[pa] * cfg.BoM[pa][mp] / 1e6 for pa in ("PA1", "PA2", "PA3"))
            end = pos - cons
            sim_mp[dia][mp] = {"pre": pre, "arr": arr, "pos": pos, "desc": desc,
                                "cons": cons, "end": end}
            stk_mp[mp] = max(0, end)
        for cd in CDS:
            sim_pa[cd][dia] = {}
            for pa in ("PA1", "PA2", "PA3"):
                pre = stk_pa[cd][pa]
                arr = arrivals_pa_cd[dia][cd].get(pa, 0)
                sai = saidas_cd[dia][cd].get(pa, 0)
                end = pre + arr - sai
                sim_pa[cd][dia][pa] = {"pre": pre, "arr": arr, "sai": sai, "end": end}
                stk_pa[cd][pa] = max(0, end)

    estoque_mp_final = {mp: sim_mp[5][mp]["end"] for mp in ("MP1", "MP2", "MP3")}
    estoque_pa_final = {cd: {pa: int(sim_pa[cd][5][pa]["end"]) for pa in ("PA1", "PA2", "PA3")}
                        for cd in CDS}

    # Custos
    forn_min = {mp: min(cfg.fornecedores[mp], key=lambda x: x[1]) for mp in ("MP1", "MP2", "MP3")}
    custo_compra = sum(float(row["Qtde"]) * forn_min[row["Tipo do Produto"]][1]
                       for _, row in df_sol.iterrows() if row["Origem"] == "Fornecedor")
    custo_frete = 0.0
    for _, row in df_sol.iterrows():
        modal = row["Modal"]; item = row["Tipo do Produto"]; qtd = float(row["Qtde"])
        km = km_rota(cfg, row["Cidade"], row["Cidade_Destino"], modal) or 0
        peso = qtd if item.startswith("MP") else qtd * cfg.peso_un_ton[item]
        custo_frete += custo_total_modal(cfg, modal, km, peso, 1)
    maior_mp = {"MP1": 56000, "MP2": 22000, "MP3": 41000}
    preco_pa = {"PA1": 80, "PA2": 50, "PA3": 32}
    carreg_mp = sum(estoque_mp_final[mp] * maior_mp[mp] * 0.01 for mp in ("MP1", "MP2", "MP3"))
    carreg_pa = sum(estoque_pa_final[cd][pa] * preco_pa[pa] * 0.01
                    for cd in CDS for pa in ("PA1", "PA2", "PA3"))
    FIX = {"parcela_terr": 506968, "parcela_maq": 415567, "parcela_mo": 84,
           "manut_fab": 1313, "sal_op": 450, "custo_prod": 172086, "manut_cd": 26683}
    fix_total = sum(FIX.values())

    # Atendimento (CD → Varejo qty por (cidade, pa))
    cdv_por_op = defaultdict(float)
    for _, row in df_sol.iterrows():
        if row["Origem"] == "CD":
            dia_part = int(_DIA.search(str(row["Dia da Coleta"])).group(1))
            modal = row["Modal"]
            lt = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], modal) or 0
            dia_cheg = dia_part + lt
            cdv_por_op[(row["Cidade_Destino"], row["Tipo do Produto"], dia_cheg)] += float(row["Qtde"])

    atendidas = []
    descartadas = []
    for op in OPS_R3:
        key = (op["cidade"], op["pa"], op["dia_entrega"])
        qtd_dia_exato = cdv_por_op.get(key, 0)
        if abs(qtd_dia_exato - op["qtd"]) < 1:
            atendidas.append(op)
        else:
            descartadas.append({**op, "qtd_no_dia": qtd_dia_exato})

    qtd_atend = sum(o["qtd"] for o in atendidas)
    total_qtd = sum(o["qtd"] for o in OPS_R3)
    ns_pct = qtd_atend / total_qtd * 100
    receita = qtd_atend * 32.00
    custo_total = fix_total + custo_compra + custo_frete + carreg_mp + carreg_pa
    resultado = receita - custo_total

    # ============ CRIA WB ============
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 00 RESUMO
    ws = wb.create_sheet("00_RESUMO_SOLVER")
    title(ws, 1, 5, f"RESUMO EXECUTIVO — SOLVER MILP — RODADA {args.rodada}")
    header(ws, 3, ["Indicador", "Valor", "Detalhe", "Fonte"])
    kpis = [
        ("🤖 Origem do plano", "SOLVER MILP (python-mip CBC)", "", "solver/milp.py"),
        ("🎯 NS R3", f"{len(atendidas)}/{len(OPS_R3)} ({ns_pct:.1f}%)", f"{qtd_atend:,} frascos no dia exato", "FLAMENGO.xlsm CD→V"),
        ("💰 Receita R3", f"R$ {receita:,.0f}", f"{qtd_atend:,} × R$ 32,00", "Σ qtds × preço"),
        ("💸 Custo total", f"R$ {custo_total:,.0f}", "Fixos + MP + Frete + Carreg", "calculado"),
        ("  Custos fixos", f"R$ {fix_total:,.0f}", "parcelas + manut + EE", "DRE oficial"),
        ("  Compra MP", f"R$ {custo_compra:,.0f}", f"Σ qty × R$/ton fornec mais barato", "Forn→F1 linhas"),
        ("  Frete", f"R$ {custo_frete:,.0f}", "Σ frete por viagem", "viagem por viagem"),
        ("  Carregamento", f"R$ {carreg_mp+carreg_pa:,.0f}", "MP × maior_preço × 1% + PA × preço × 1%", "estoque final × 1%"),
        ("📈 LUCRO R3", f"R$ {resultado:,.0f}", "Receita - Custos", "↑"),
        ("⏱️ Utilização", f"{sum(int(df_op[df_op['Dia']==f'Dia {d}'].iloc[0][pa])/[15,30,60][i] for d in range(1,6) for i,pa in enumerate(['PA1','PA2','PA3']))/50400*100:.1f}%", f"Σ min usados / 50.400", "OP_FABRICAS"),
        ("🚚 Transportes", f"{len(df_sol)}/220", "≤ 220 = OK", "SOL_TRANSP count"),
        ("📦 MP final F1 (MP1)", f"{estoque_mp_final['MP1']:.2f} t", f"cap {cap_mp['MP1']:.1f}", "simulação dia a dia"),
        ("📦 MP final F1 (MP2)", f"{estoque_mp_final['MP2']:.2f} t", f"cap {cap_mp['MP2']:.1f}", "simulação dia a dia"),
        ("📦 MP final F1 (MP3)", f"{estoque_mp_final['MP3']:.2f} t", f"cap {cap_mp['MP3']:.1f}", "simulação dia a dia"),
        ("🏪 PA2 buffer CD2", f"{estoque_pa_final['CD2']['PA2']:,}", f"cap {cap_pa_cd['CD2']['PA2']:,}", "simulação dia a dia"),
    ]
    for i, (m, v, d, f) in enumerate(kpis):
        r = 4 + i
        ws.cell(r, 1).value = m
        ws.cell(r, 2).value = v
        ws.cell(r, 3).value = d
        ws.cell(r, 4).value = f
        ws.cell(r, 2).font = FONT_BOLD
        # Destaque para linhas-chave
        if "LUCRO" in m or "NS" in m:
            for c in (1, 2, 3, 4):
                ws.cell(r, c).fill = FILL_OK
        elif i == 0:
            for c in (1, 2, 3, 4):
                ws.cell(r, c).fill = FILL_SUB

    # 01 OPs detalhe
    ws = wb.create_sheet("01_OPs_R3")
    title(ws, 1, 7, f"OPs R3 — validação dia exato")
    header(ws, 3, ["#", "Cidade", "PA", "Qtd planejada", "Dia entrega", "Qtd entregue no dia", "✅/❌"])
    for i, op in enumerate(OPS_R3, start=4):
        ws.cell(i, 1).value = i - 3
        ws.cell(i, 2).value = op["cidade"]
        ws.cell(i, 3).value = op["pa"]
        ws.cell(i, 4).value = op["qtd"]
        ws.cell(i, 5).value = op["dia_entrega"]
        key = (op["cidade"], op["pa"], op["dia_entrega"])
        qtd_no_dia = cdv_por_op.get(key, 0)
        ws.cell(i, 6).value = round(qtd_no_dia)
        ok = abs(qtd_no_dia - op["qtd"]) < 1
        ws.cell(i, 7).value = "✅" if ok else "❌"
        ws.cell(i, 7).fill = FILL_OK if ok else FILL_BAD

    # 02 Transportes
    ws = wb.create_sheet("02_Transportes_R3")
    title(ws, 1, 10, f"TRANSPORTES R3 — {len(df_sol)} linhas")
    header(ws, 3, ["#", "Origem", "Cidade", "Dia (rel)", "Modal", "Item", "Qtd", "Destino", "Cidade dest", "Lead time"])
    for i, row in df_sol.iterrows():
        r = 4 + i
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = row["Origem"]
        ws.cell(r, 3).value = row["Cidade"]
        ws.cell(r, 4).value = row["Dia da Coleta"]
        ws.cell(r, 5).value = row["Modal"]
        ws.cell(r, 6).value = row["Tipo do Produto"]
        ws.cell(r, 7).value = row["Qtde"]
        ws.cell(r, 8).value = row["Destino"]
        ws.cell(r, 9).value = row["Cidade_Destino"]
        ws.cell(r, 10).value = lead_dias(cfg, row["Cidade"], row["Cidade_Destino"], row["Modal"]) or 0
    auto_width(ws)

    # 03 Produção
    ws = wb.create_sheet("03_Producao")
    title(ws, 1, 8, "PRODUÇÃO POR DIA + MP CONSUMIDO")
    header(ws, 3, ["Dia", "PA1", "PA2", "PA3", "Min usados", "MP1 cons (t)", "MP2 cons (t)", "MP3 cons (t)"])
    for i, row in df_op.iterrows():
        r = 4 + i
        d = int(row["Dia"].split()[-1])
        pa1, pa2, pa3 = int(row["PA1"]), int(row["PA2"]), int(row["PA3"])
        ws.cell(r, 1).value = f"Dia {d}"
        ws.cell(r, 2).value = pa1
        ws.cell(r, 3).value = pa2
        ws.cell(r, 4).value = pa3
        ws.cell(r, 5).value = pa1/15 + pa2/30 + pa3/60
        ws.cell(r, 6).value = round((pa1*60 + pa2*75 + pa3*75) / 1e6, 4)
        ws.cell(r, 7).value = round((pa1*90 + pa2*125 + pa3*30) / 1e6, 4)
        ws.cell(r, 8).value = round((pa1*150 + pa2*50 + pa3*45) / 1e6, 4)
    auto_width(ws)

    # 04 Estoque MP dia a dia
    ws = wb.create_sheet("04_Estoque_MP_Dia")
    title(ws, 1, 9, "ESTOQUE MP F1 — DIA A DIA (Solver)")
    header(ws, 3, ["Dia", "MP", "Pré", "+Arr", "Pos", "Desc", "Consumo", "End", "Cap"])
    r = 4
    for d in range(1, 6):
        for mp in ("MP1", "MP2", "MP3"):
            info = sim_mp[d][mp]
            ws.cell(r, 1).value = f"Dia {d}"
            ws.cell(r, 2).value = mp
            ws.cell(r, 3).value = round(info["pre"], 3)
            ws.cell(r, 4).value = round(info["arr"], 3)
            ws.cell(r, 5).value = round(info["pos"], 3)
            ws.cell(r, 6).value = round(info["desc"], 3)
            ws.cell(r, 7).value = round(info["cons"], 3)
            ws.cell(r, 8).value = round(info["end"], 3)
            ws.cell(r, 9).value = round(cap_mp[mp], 2)
            if info["desc"] > 0.01: ws.cell(r, 6).fill = FILL_BAD
            if info["end"] < -0.01: ws.cell(r, 8).fill = FILL_BAD
            r += 1
        r += 1
    auto_width(ws)

    # 05 Estoque PA CD dia a dia
    ws = wb.create_sheet("05_Estoque_PA_Dia")
    title(ws, 1, 8, "ESTOQUE PA NOS CDs — DIA A DIA (Solver)")
    header(ws, 3, ["Dia", "CD", "PA", "Pré", "+Arr", "−Saída", "End", "Cap"])
    r = 4
    for d in range(1, 6):
        for cd in ("CD1", "CD2"):
            for pa in ("PA1", "PA2", "PA3"):
                info = sim_pa[cd][d][pa]
                ws.cell(r, 1).value = f"Dia {d}"
                ws.cell(r, 2).value = f"{cd} {CDS[cd]['cidade']}"
                ws.cell(r, 3).value = pa
                ws.cell(r, 4).value = round(info["pre"])
                ws.cell(r, 5).value = round(info["arr"])
                ws.cell(r, 6).value = round(info["sai"])
                ws.cell(r, 7).value = round(info["end"])
                ws.cell(r, 8).value = cap_pa_cd[cd][pa]
                r += 1
        r += 1
    auto_width(ws)

    # 06 Estoque final
    ws = wb.create_sheet("06_Estoque_FimR3")
    title(ws, 1, 6, "POSIÇÃO FINAL R3 (Dia 15)")
    header(ws, 3, ["MP/PA", "Local", "Estoque", "Cap", "Custo carreg R$", "Ocup"])
    r = 4
    for mp in ("MP1", "MP2", "MP3"):
        ws.cell(r, 1).value = mp
        ws.cell(r, 2).value = "F1 Joinville"
        ws.cell(r, 3).value = round(estoque_mp_final[mp], 3)
        ws.cell(r, 4).value = round(cap_mp[mp], 2)
        ws.cell(r, 5).value = round(estoque_mp_final[mp] * maior_mp[mp] * 0.01, 2)
        ws.cell(r, 6).value = f"=C{r}/D{r}"
        pct(ws.cell(r, 6))
        money(ws.cell(r, 5))
        r += 1
    r += 1
    for cd in ("CD1", "CD2"):
        for pa in ("PA1", "PA2", "PA3"):
            ws.cell(r, 1).value = pa
            ws.cell(r, 2).value = f"{cd} {CDS[cd]['cidade']}"
            ws.cell(r, 3).value = estoque_pa_final[cd][pa]
            ws.cell(r, 4).value = cap_pa_cd[cd][pa]
            ws.cell(r, 5).value = round(estoque_pa_final[cd][pa] * preco_pa[pa] * 0.01, 2)
            ws.cell(r, 6).value = f"=C{r}/D{r}"
            pct(ws.cell(r, 6))
            money(ws.cell(r, 5))
            r += 1
    auto_width(ws)

    # 07 DRE
    ws = wb.create_sheet("07_DRE")
    title(ws, 1, 4, "DRE — Rodada 3 (Solver)")
    header(ws, 3, ["Linha", "Fórmula", "Valor R$", "Como é calculado"])
    rows = [
        ("Receita PA3", f"{qtd_atend:,} × R$ 32", receita, "Σ qtd entregue × preço"),
        ("(-) Parcela terrenos", "fixo", -FIX["parcela_terr"], "DRE oficial"),
        ("(-) Parcela máquinas", "fixo", -FIX["parcela_maq"], "DRE oficial"),
        ("(-) Contratação MO", "fixo", -FIX["parcela_mo"], "DRE oficial"),
        ("(-) Manut fábricas", "fixo", -FIX["manut_fab"], "DRE oficial"),
        ("(-) Salário operários", "fixo", -FIX["sal_op"], "DRE oficial"),
        ("(-) Custo produção", "fixo", -FIX["custo_prod"], "DRE oficial"),
        ("(-) Manut CDs", "fixo", -FIX["manut_cd"], "DRE oficial"),
        ("(-) Compra MP", "Σ qty×preço fornec", -custo_compra, "fornecedor mais barato"),
        ("(-) Frete", "Σ frete por viagem", -custo_frete, "regra ocup ≥80% ou peso"),
        ("(-) Carreg MP", "Σ MP final × maior_preço × 1%", -carreg_mp, "MP1 Belém 56k, MP2 VdC 22k, MP3 Joi 41k"),
        ("(-) Carreg PA", "Σ PA final × preço tabela × 1%", -carreg_pa, "PA2 R$50, PA3 R$25"),
        ("= RESULTADO R3", "Receita - Custos", resultado, "↑"),
    ]
    for i, (l, f, v, c) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1).value = l
        ws.cell(r, 2).value = f
        ws.cell(r, 3).value = v
        ws.cell(r, 4).value = c
        money(ws.cell(r, 3))
        if "RESULTADO" in l:
            for cc in range(1, 5):
                ws.cell(r, cc).font = FONT_BOLD
                ws.cell(r, cc).fill = FILL_OK
    auto_width(ws)

    # 08 Checks
    ws = wb.create_sheet("08_Checks")
    title(ws, 1, 5, "SANITY CHECKS")
    header(ws, 3, ["#", "Checagem", "Valor", "Esperado", "Status"])

    n_neg_mp = sum(1 for d in range(1, 6) for mp in ("MP1", "MP2", "MP3") if sim_mp[d][mp]["end"] < -0.01)
    n_desc_mp = sum(1 for d in range(1, 6) for mp in ("MP1", "MP2", "MP3") if sim_mp[d][mp]["desc"] > 0.01)
    n_neg_pa = sum(1 for cd in CDS for d in range(1, 6) for pa in ("PA1","PA2","PA3") if sim_pa[cd][d][pa]["end"] < -0.5)
    n_over_pa = sum(1 for cd in CDS for d in range(1, 6) for pa in ("PA1","PA2","PA3") if sim_pa[cd][d][pa]["end"] > cap_pa_cd[cd][pa])
    max_min = max(sum(int(df_op[df_op["Dia"]==f"Dia {d}"].iloc[0][pa])/[15,30,60][i] for i,pa in enumerate(("PA1","PA2","PA3"))) for d in range(1,6))

    checks = [
        ("Transportes ≤ 220", len(df_sol), "≤ 220", len(df_sol) <= 220),
        ("Min usados ≤ Cap/dia", round(max_min), f"≤ {cap_min_dia}", max_min <= cap_min_dia + 1),
        ("OPs entregues no dia exato", f"{len(atendidas)}/{len(OPS_R3)}", "25", len(atendidas) == 25),
        ("Estoque MP nunca neg (dia a dia)", f"{n_neg_mp} dias", "0", n_neg_mp == 0),
        ("Estoque MP nunca excede cap (descarte=0)", f"{n_desc_mp} eventos", "0", n_desc_mp == 0),
        ("Estoque PA CD nunca negativo", f"{n_neg_pa} dias", "0", n_neg_pa == 0),
        ("Estoque PA CD ≤ cap CD", f"{n_over_pa} excede", "0", n_over_pa == 0),
        ("MP1 final ≤ cap", round(estoque_mp_final["MP1"], 2), round(cap_mp["MP1"], 2),
         estoque_mp_final["MP1"] <= cap_mp["MP1"] + 0.01),
        ("MP2 final ≤ cap", round(estoque_mp_final["MP2"], 2), round(cap_mp["MP2"], 2),
         estoque_mp_final["MP2"] <= cap_mp["MP2"] + 0.01),
        ("MP3 final ≤ cap", round(estoque_mp_final["MP3"], 2), round(cap_mp["MP3"], 2),
         estoque_mp_final["MP3"] <= cap_mp["MP3"] + 0.01),
    ]
    for i, (c, v, e, ok) in enumerate(checks):
        r = 4 + i
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = c
        ws.cell(r, 3).value = v
        ws.cell(r, 4).value = e
        ws.cell(r, 5).value = "✅ OK" if ok else "❌ FALHA"
        ws.cell(r, 5).font = FONT_BOLD
        ws.cell(r, 5).fill = FILL_OK if ok else FILL_BAD
    auto_width(ws)

    OUT = BASE / "solver" / "rodadas" / f"rodada_{args.rodada}" / "SanityCheck_Solver.xlsm"
    wb.save(OUT)
    print(f"\nOK: SanityCheck_Solver.xlsm criado em {OUT}")
    print(f"\nRESUMO R3 (SOLVER):")
    print(f"  NS:                  {ns_pct:.1f}% ({len(atendidas)}/{len(OPS_R3)})")
    print(f"  Receita:             R$ {receita:,.0f}")
    print(f"  Custos fixos:        R$ {fix_total:,.0f}")
    print(f"  Custo compra MP:     R$ {custo_compra:,.0f}")
    print(f"  Custo frete:         R$ {custo_frete:,.0f}")
    print(f"  Custo carregamento:  R$ {carreg_mp + carreg_pa:,.0f}")
    print(f"  Custo TOTAL:         R$ {custo_total:,.0f}")
    print(f"  RESULTADO R3:        R$ {resultado:,.0f}")
    print(f"  Estoque MP final:    MP1={estoque_mp_final['MP1']:.2f}t  MP2={estoque_mp_final['MP2']:.2f}t  MP3={estoque_mp_final['MP3']:.2f}t")
    print(f"  Buffer PA2 R4:       {estoque_pa_final['CD2']['PA2']:,} (Santos)")


if __name__ == "__main__":
    main()
