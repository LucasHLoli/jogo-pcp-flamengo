"""Leitura e escrita de FLAMENGO.xlsm, Rodada N.xlsm e OP_Rodada_N.xlsx."""
from __future__ import annotations
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
import pandas as pd

from src.domain import OP, TransitItem, PlanoProducao, PlanoTransporte  # noqa: F401


_DIA_RE = re.compile(r"Dia\s*(\d+)")
_RODADA_RE = re.compile(r"Rodada_?(\d+)")


def _parse_dia(value) -> int | None:
    if value is None:
        return None
    s = str(value).strip()
    m = _DIA_RE.search(s)
    return int(m.group(1)) if m else None


def _parse_rodada(value) -> int | None:
    if value is None:
        return None
    s = str(value).strip()
    m = _RODADA_RE.search(s)
    return int(m.group(1)) if m else None


def ler_instalacoes(path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    ws = wb["INSTALAÇÕES"]
    empresa = ws.cell(1, 3).value
    fabricas = {}
    for row in (8, 9):
        nome = ws.cell(row, 1).value
        cidade = ws.cell(row, 2).value
        if not cidade:
            continue
        fabricas[nome] = {
            "cidade": cidade,
            "maquinas": int(ws.cell(row, 3).value or 0),
            "turnos": int(ws.cell(row, 4).value or 0),
            "mo": int(ws.cell(row, 5).value or 0),
            "area_mp": {
                "MP1": float(ws.cell(row, 6).value or 0),
                "MP2": float(ws.cell(row, 7).value or 0),
                "MP3": float(ws.cell(row, 8).value or 0),
            },
        }
    cds = {}
    for row in (12, 13, 14, 15):
        nome = ws.cell(row, 1).value
        cidade = ws.cell(row, 2).value
        if not cidade:
            continue
        cds[nome] = {
            "cidade": cidade,
            "area_pa": {
                "PA1": float(ws.cell(row, 3).value or 0),
                "PA2": float(ws.cell(row, 4).value or 0),
                "PA3": float(ws.cell(row, 5).value or 0),
            },
            "area_total": float(ws.cell(row, 6).value or 0),
        }
    return {"empresa": empresa, "fabricas": fabricas, "cds": cds}


def calcular_rod_dia_chegada(rod_part: int, dia_part: int, lead_dias: int) -> Tuple[int, int]:
    """Regra fechada (spec §6.2). dia_part é relativo (1-5)."""
    total = dia_part + lead_dias
    rod_cheg = rod_part + (total - 1) // 5
    dia_cheg = ((total - 1) % 5) + 1
    return rod_cheg, dia_cheg


def dia_relativo_para_absoluto(rodada_n: int, dia_relativo: int) -> int:
    """Converte dia relativo da rodada (1-5) → dia absoluto do jogo (1-75).
    R1 Dia 1 = Dia 1 absoluto, R2 Dia 1 = Dia 6, R3 Dia 1 = Dia 11, etc."""
    return (rodada_n - 1) * 5 + dia_relativo


def dia_absoluto_para_relativo(rodada_n: int, dia_absoluto: int) -> int:
    """Converte dia absoluto → relativo da rodada."""
    return dia_absoluto - (rodada_n - 1) * 5


def ler_sol_transp(path: Path, rodada: int | None = None) -> List[TransitItem]:
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    ws = wb["SOL_TRANSP"]
    items: List[TransitItem] = []
    for r in range(5, ws.max_row + 1):
        rod = _parse_rodada(ws.cell(r, 1).value)
        if rod is None:
            break
        if rodada is not None and rod != rodada:
            continue
        origem_tipo = ws.cell(r, 2).value
        origem_cidade = ws.cell(r, 3).value
        dia_part_abs = _parse_dia(ws.cell(r, 4).value)
        modal = ws.cell(r, 5).value
        item = ws.cell(r, 6).value
        qtd = ws.cell(r, 7).value
        destino_tipo = ws.cell(r, 8).value
        destino_cidade = ws.cell(r, 9).value
        lead = ws.cell(r, 10).value
        if dia_part_abs is None or modal is None or item is None or qtd is None:
            continue
        try:
            lead_int = int(float(lead)) if lead is not None else 0
        except (TypeError, ValueError):
            lead_int = 0
        # Converte dia absoluto da planilha (Dia 6-10 pra R2) → relativo (1-5) interno
        dia_part = dia_absoluto_para_relativo(rod, dia_part_abs)
        # Se já estiver em formato relativo (Dia 1-5 pra R1, ou plano legado), mantém
        if dia_part < 1 or dia_part > 5:
            dia_part = dia_part_abs  # fallback: assume que veio relativo
        rod_cheg, dia_cheg = calcular_rod_dia_chegada(rod, dia_part, lead_int)
        items.append(TransitItem(
            rod_part=rod, dia_part=dia_part,
            rod_cheg=rod_cheg, dia_cheg=dia_cheg,
            origem_tipo=origem_tipo, origem_cidade=origem_cidade,
            destino_tipo=destino_tipo, destino_cidade=destino_cidade,
            modal=modal, item=item, qtd=float(qtd),
        ))
    return items


def ler_op_rodada(path: Path) -> List[OP]:
    path = Path(path)
    if not path.exists():
        return []
    df = pd.read_excel(path)
    return [
        OP(rodada=int(r["Rodada"]), cidade=str(r["Cidade"]), pa=str(r["PA"]),
           qtd=int(r["Qtd"]), dia_entrega=int(r["Dia_Entrega"]))
        for _, r in df.iterrows()
    ]


def escrever_plano(
    path: Path,
    planos_transporte: List["PlanoTransporte"],
    planos_producao: List["PlanoProducao"],
    rodada_n: int,
) -> None:
    """Escreve SOL_TRANSP e OP_FABRICAS preservando VBA e fórmulas.

    SOL_TRANSP: limpa colunas A-I das linhas da rodada_n e regrava.
    OP_FABRICAS: atualiza bloco F1 (linhas 7-11, colunas B-D).
    """
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=False)
    ws = wb["SOL_TRANSP"]

    # 1) Limpar linhas A-I desta rodada (preserva J-Z fórmulas)
    for r in range(5, ws.max_row + 1):
        rod = _parse_rodada(ws.cell(r, 1).value)
        if rod == rodada_n:
            for c in range(1, 10):
                ws.cell(r, c).value = None

    # 2) Encontrar primeira linha livre para a rodada N
    linha_alvo = 5
    while True:
        val = ws.cell(linha_alvo, 1).value
        if val is None or str(val).strip() == "":
            break
        linha_alvo += 1

    # 3) Escrever planos
    for plano in planos_transporte:
        ws.cell(linha_alvo, 1).value = f"Rodada_{plano.rodada}"
        ws.cell(linha_alvo, 2).value = plano.origem_tipo
        ws.cell(linha_alvo, 3).value = plano.origem_cidade
        ws.cell(linha_alvo, 4).value = f"Dia {plano.dia_coleta}"
        ws.cell(linha_alvo, 5).value = plano.modal
        ws.cell(linha_alvo, 6).value = plano.item
        ws.cell(linha_alvo, 7).value = plano.qtd
        ws.cell(linha_alvo, 8).value = plano.destino_tipo
        ws.cell(linha_alvo, 9).value = plano.destino_cidade
        linha_alvo += 1

    # 4) OP_FABRICAS bloco F1: linhas 7..11, colunas B=PA1, C=PA2, D=PA3
    ws_op = wb["OP_FABRICAS"]
    pa_to_col = {"PA1": 2, "PA2": 3, "PA3": 4}
    for r in range(7, 12):
        for c in (2, 3, 4):
            ws_op.cell(r, c).value = 0.0
    for r in range(17, 22):
        for c in (2, 3, 4):
            ws_op.cell(r, c).value = 0.0
    dia_to_linha_f1 = {d: 6 + d for d in range(1, 6)}
    for plano in planos_producao:
        if plano.fabrica != "F1":
            continue
        r = dia_to_linha_f1[plano.dia]
        c = pa_to_col[plano.pa]
        ws_op.cell(r, c).value = int(plano.qtd)
    ws_op.cell(4, 6).value = f"Rodada_{rodada_n}"

    wb.save(path)


def escrever_planos_de_df(
    path: Path,
    df_sol_transp,
    df_op_fabricas,
    rodada_n: int,
    fabrica: str = "F1",
) -> int:
    """Escreve direto a partir dos DataFrames retornados pelo planner_manual,
    preservando VBA, fórmulas (colunas J-Z do SOL_TRANSP) e demais abas intactas.

    Args:
        path: caminho do FLAMENGO.xlsm (será modificado in-place).
        df_sol_transp: DataFrame com colunas (Rodada, Origem, Cidade, Dia da Coleta,
            Modal, Tipo do Produto, Qtde, Destino, Cidade_Destino).
        df_op_fabricas: DataFrame com colunas (Dia, PA1, PA2, PA3) — 5 linhas.
        rodada_n: número da rodada (para limpar linhas antigas dessa rodada).
        fabrica: 'F1' ou 'F2'. Default 'F1'.

    Returns:
        Número de linhas escritas no SOL_TRANSP.
    """
    wb = openpyxl.load_workbook(path, keep_vba=True, data_only=False)
    ws = wb["SOL_TRANSP"]

    # 1) Limpar linhas A-I da rodada_n (preserva J-Z fórmulas)
    for r in range(5, ws.max_row + 1):
        rod = _parse_rodada(ws.cell(r, 1).value)
        if rod == rodada_n:
            for c in range(1, 10):
                ws.cell(r, c).value = None

    # 2) Achar primeira linha livre (após linhas de rodadas anteriores)
    linha_alvo = 5
    while True:
        val = ws.cell(linha_alvo, 1).value
        if val is None or str(val).strip() == "":
            break
        linha_alvo += 1

    # 3) Escrever cada linha do DataFrame
    # IMPORTANTE: converte "Dia X" relativo (1-5) → absoluto do jogo
    # R1 Dia 1 → Dia 1; R2 Dia 1 → Dia 6; R3 Dia 1 → Dia 11; etc.
    n_linhas = 0
    for _, row in df_sol_transp.iterrows():
        dia_str = str(row["Dia da Coleta"])
        m_dia = _DIA_RE.search(dia_str)
        if m_dia:
            dia_rel = int(m_dia.group(1))
            # Se vier > 5, já está em absoluto; senão converte
            if 1 <= dia_rel <= 5:
                dia_abs = dia_relativo_para_absoluto(rodada_n, dia_rel)
                dia_str = f"Dia {dia_abs}"
        ws.cell(linha_alvo, 1).value = str(row["Rodada"])
        ws.cell(linha_alvo, 2).value = str(row["Origem"])
        ws.cell(linha_alvo, 3).value = str(row["Cidade"])
        ws.cell(linha_alvo, 4).value = dia_str
        ws.cell(linha_alvo, 5).value = str(row["Modal"])
        ws.cell(linha_alvo, 6).value = str(row["Tipo do Produto"])
        # Qtde: int se for MP (toneladas) ou int se for PA (unidades).
        # Como o jogo aceita float, deixamos como número.
        try:
            ws.cell(linha_alvo, 7).value = float(row["Qtde"])
        except (TypeError, ValueError):
            ws.cell(linha_alvo, 7).value = row["Qtde"]
        ws.cell(linha_alvo, 8).value = str(row["Destino"])
        ws.cell(linha_alvo, 9).value = str(row["Cidade_Destino"])
        linha_alvo += 1
        n_linhas += 1

    # 4) OP_FABRICAS: zera bloco F1 (linhas 7..11) e F2 (17..21), escreve do df
    ws_op = wb["OP_FABRICAS"]
    pa_to_col = {"PA1": 2, "PA2": 3, "PA3": 4}
    bloco_linhas_f1 = range(7, 12)
    bloco_linhas_f2 = range(17, 22)
    for r in bloco_linhas_f1:
        for c in (2, 3, 4):
            ws_op.cell(r, c).value = 0.0
    for r in bloco_linhas_f2:
        for c in (2, 3, 4):
            ws_op.cell(r, c).value = 0.0

    # Mapeia Dia 1..5 → linhas 7..11 (F1) ou 17..21 (F2)
    base_linha = 6 if fabrica == "F1" else 16  # Dia 1 = base+1
    for _, row in df_op_fabricas.iterrows():
        dia_str = str(row["Dia"]).strip()
        # extrai número do "Dia X"
        try:
            dia_num = int(dia_str.split()[-1])
        except (ValueError, IndexError):
            continue
        if not (1 <= dia_num <= 5):
            continue
        r = base_linha + dia_num
        for pa in ("PA1", "PA2", "PA3"):
            if pa in row:
                ws_op.cell(r, pa_to_col[pa]).value = int(row[pa])

    # Atualiza referência de rodada
    ws_op.cell(4, 6).value = f"Rodada_{rodada_n}"

    wb.save(path)
    return n_linhas
